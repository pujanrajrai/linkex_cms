from django.contrib import messages
from .form import LedgerForm
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse, QueryDict
from django.utils.safestring import mark_safe
from django.db.models import Q
import json
import re
from django.shortcuts import get_object_or_404, render, redirect
from decorators import has_roles
from finance.models.ledger import Ledger
from accounts.models import Agency
from datetime import datetime
from django.views import View
from django.utils.decorators import method_decorator
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import pytz
from finance.models.invoice import Invoice


@login_required
@has_roles(['admin', 'agencyuser'])
def ledger_list(request):
    requested_html = re.search(r'^text/html', request.META.get('HTTP_ACCEPT'))
    if not requested_html:
        post_dict = QueryDict(request.POST.urlencode(), mutable=True)
        param_start = max(int(request.POST.get('start', 0)), 0)
        param_limit = max(int(request.POST.get('length', 10)), 1)

        column_map = {
            1: 'company__name',
            2: 'agency__company_name',
            5: 'particular',
            8: 'entry_type',
            9: 'reference',
            10: 'reference_no',
            11: 'remarks',
        }

        object_list = Ledger.objects.select_related('agency', 'company').all()

        if request.user.role == 'agencyuser':
            object_list = object_list.filter(agency=request.user.agency)

        # Apply filters from GET parameters (for export consistency)
        agency_id = request.GET.get('agency') or request.POST.get('agency')
        company_id = request.GET.get('company') or request.POST.get('company')
        from_date_str = request.GET.get(
            'from_date') or request.POST.get('from_date')
        to_date_str = request.GET.get('to_date') or request.POST.get('to_date')

        if agency_id:
            object_list = object_list.filter(agency_id=agency_id)
        if company_id:
            object_list = object_list.filter(company_id=company_id)
        if from_date_str:
            try:
                from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
                object_list = object_list.filter(created_at__gte=from_date)
            except ValueError:
                pass
        if to_date_str:
            try:
                from datetime import time
                to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
                to_date_end = datetime.combine(to_date, time.max)
                object_list = object_list.filter(created_at__lte=to_date_end)
            except ValueError:
                pass

        count = object_list.count()

        # Global search
        search_value = request.POST.get('search[value]', '')
        if search_value:
            object_list = object_list.filter(
                Q(agency__company_name__icontains=search_value) |
                Q(company__name__icontains=search_value) |
                Q(particular__icontains=search_value) |
                Q(reference__icontains=search_value) |
                Q(reference_no__icontains=search_value) |
                Q(remarks__icontains=search_value)
            )

        # Ordering
        order_column_index = int(request.POST.get('order[0][column]', 0))
        order_dir = request.POST.get('order[0][dir]', 'desc')
        order_field = column_map.get(order_column_index, 'created_at')
        if order_dir == "asc":
            order_field = '-' + order_field
        else:
            order_field = order_field

        for index, field in column_map.items():
            column_search_value = request.POST.get(
                f'columns[{index}][search][value]', '').strip()
            if column_search_value:
                kwargs = {f"{field}__icontains": column_search_value}
                object_list = object_list.filter(**kwargs)
        # Always add created_at as secondary ordering for consistency
        object_list = object_list.order_by(
            order_field, '-created_at')[param_start:param_start + param_limit]

        filtered_total = object_list.count()
        row = list(object_list.values(
            'id', 'agency__company_name', 'company__name', 'ledger_type',
            'entry_type', 'particular', 'amount', 'company_balance',
            'balance', 'reference', 'reference_no', 'remarks', 'created_at'
        ))

        context = {
            'draw': post_dict.get('draw'),
            'recordsTotal': count,
            'recordsFiltered': filtered_total,
            'data': row,
        }

        data = mark_safe(json.dumps(context, indent=4,
                         sort_keys=True, default=str))
        return HttpResponse(data, content_type='application/json')

    # Handle filter form for HTML requests
    from finance.forms import LedgerFilterForm
    form = LedgerFilterForm(request.GET, user=request.user)

    context = {
        "title": "Ledger",
        "form": form,
        "agency": request.GET.get('agency', ''),
        "company": request.GET.get('company', ''),
        "from_date": request.GET.get('from_date', ''),
        "to_date": request.GET.get('to_date', '')
    }
    return render(request, "finance/ledger/list.html", context)


@login_required
def get_agency_balance(request, agency_pk):
    agency = get_object_or_404(Agency, pk=agency_pk)
    leadger = Ledger.objects.filter(agency=agency).first()
    return JsonResponse({
        'last_balance': leadger.get_last_balance(agency=agency)
    })


@login_required
@has_roles(['admin', 'agencyuser'])
def ledger_entry(request):
    if request.method == 'POST':
        form = LedgerForm(request.POST)
        if form.is_valid():
            Ledger.objects.create(
                agency=form.cleaned_data['agency'],
                company=form.cleaned_data['company'],
                ledger_type=form.cleaned_data['ledger_type'],
                particular=form.cleaned_data['particular'],
                amount=form.cleaned_data['amount'],
                remarks=form.cleaned_data['remarks'],
                entry_type="LEDGER ENTRY"
            )
            messages.success(request, 'Ledger entry created successfully')
            return redirect('finance:pages:ledger:list')
    else:
        form = LedgerForm()
    context = {
        'form': form,
        "title": "Ledger Entry"
    }
    return render(request, "finance/ledger/create.html", context)


@method_decorator(login_required, name='dispatch')
@method_decorator(has_roles(['admin']), name='dispatch')
class ExportLedgerExcel(View):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Ledger.xlsx"'

        # Get filter parameters
        agency_id = request.GET.get('agency')
        company_id = request.GET.get('company')
        from_date_str = request.GET.get('from_date')
        to_date_str = request.GET.get('to_date')

        # Parse dates properly
        from_date = self.parse_date(from_date_str) if from_date_str else None
        to_date = self.parse_date(to_date_str) if to_date_str else None

        # Build queryset
        ledgers = Ledger.objects.select_related('agency', 'company').all()

        # Filter by user role
        if request.user.role == 'agencyuser':
            ledgers = ledgers.filter(agency=request.user.agency)

        # Apply filters
        if agency_id:
            ledgers = ledgers.filter(agency_id=agency_id)
        if company_id:
            ledgers = ledgers.filter(company_id=company_id)
        if from_date:
            ledgers = ledgers.filter(created_at__gte=from_date)
        if to_date:
            # Add time to make it end of day
            from datetime import time
            to_date_end = datetime.combine(to_date, time.max)
            ledgers = ledgers.filter(created_at__lte=to_date_end)

        # Order by creation date
        ledgers = ledgers.order_by('-created_at')

        # Create workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Ledger"

        headers = [
            "Date", "Agency", "Company", "Debit", "Credit",
            "Particular", "Company Balance", "Balance", "Entry Type",
            "Reference", "Reference No", "Remarks"
        ]

        # Add headers
        worksheet.append(headers)

        # Define border style - darker borders
        dark_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        # Style headers
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        data_alignment = Alignment(
            horizontal='center', vertical='center', indent=1)

        # Apply header styling
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = dark_border

        # Track column widths for auto-sizing
        column_widths = {}

        # Initialize column widths with header lengths
        for col_num, header in enumerate(headers, 1):
            column_widths[col_num] = len(str(header))

        row_num = 2  # Start from row 2 (after headers)

        invoice = True

        for ledger in ledgers:
            # check if invoice is active
            if ledger.reference == 'INVOICE':
                try:
                    invoice = Invoice.objects.get(id=ledger.reference_no)
                    if not invoice.is_active:
                        invoice = False
                except Exception as e:
                    invoice = True
            # Prepare row data
            debit_amount = ledger.amount if ledger.ledger_type == 'DEBIT' else ''
            credit_amount = ledger.amount if ledger.ledger_type == 'CREDIT' else ''

            row_data = [
                ledger.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                ledger.agency.company_name,
                ledger.company.name,
                debit_amount,
                credit_amount,
                ledger.particular,
                ledger.company_balance,
                ledger.balance,
                ledger.entry_type,
                ledger.reference or '',
                ledger.reference_no or '',
                ledger.remarks or ''
            ]

            # Add row data to worksheet
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = dark_border
                cell.alignment = data_alignment

                # Update column width tracking
                if value is not None:
                    cell_length = len(str(value))
                    if col_num not in column_widths or cell_length > column_widths[col_num]:
                        column_widths[col_num] = cell_length
            if not invoice:
                for col_num in range(1, len(row_data) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(
                        start_color='FF9999', end_color='FF9999', fill_type='solid')
            row_num += 1

        # Auto-adjust column widths
        for col_num, width in column_widths.items():
            column_letter = get_column_letter(col_num)
            # Add some padding to the width and set a reasonable maximum
            adjusted_width = min(width + 3, 100)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(response)
        return response

    def parse_date(self, date_str):
        """Parse date string with multiple format attempts"""
        if not date_str:
            return None

        # List of possible date formats to try
        date_formats = [
            '%Y-%m-%d',      # 2024-01-15
            '%m/%d/%Y',      # 01/15/2024
            '%d/%m/%Y',      # 15/01/2024
            '%d-%m-%Y',      # 15-01-2024
            '%Y/%m/%d',      # 2024/01/15
            '%b. %d, %Y',    # Jan. 15, 2024
            '%B %d, %Y',     # January 15, 2024
        ]

        for date_format in date_formats:
            try:
                return datetime.strptime(date_str, date_format)
            except ValueError:
                continue

        # If no format worked, log the error and return None
        print(f"Could not parse date: {date_str}")
        return None
