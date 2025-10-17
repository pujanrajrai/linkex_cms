from django.shortcuts import render, get_object_or_404
from django.utils.safestring import mark_safe
from django.db.models import Q
from django.http import HttpResponse, QueryDict
from django.contrib.auth.decorators import login_required
import json
import re
from awb.models import AWBAPIResponse
from decorators import has_roles
from openpyxl import Workbook
from datetime import datetime


@login_required
@has_roles(['admin'])
def api_responses_list(request):
    """
    View to handle AWB API responses list with server-side DataTables processing
    """
    requested_html = re.search(r'^text/html', request.META.get('HTTP_ACCEPT'))

    if not requested_html:
        # Handle DataTables AJAX request
        post_dict = QueryDict(request.POST.urlencode(), mutable=True)
        param_start = max(int(request.POST.get('start', 0)), 0)
        param_limit = max(int(request.POST.get('length', 10)), 1)

        # Column mapping for ordering
        column_map = {
            0: None,  # Serial number
            1: 'awb__awbno',  # AWB Number
            2: 'vendor',  # Vendor
            3: 'is_success',  # Status
            4: 'created_at',  # Created At
            5: None,  # Reference No
            6: None,  # Actions
        }

        # Get all API responses
        object_list = AWBAPIResponse.objects.select_related('awb').all()
        count = object_list.count()

        # Handle search
        search_value = request.POST.get('search[value]', '')
        if search_value:
            object_list = object_list.filter(
                Q(awb__awbno__icontains=search_value) |
                Q(vendor__icontains=search_value) |
                Q(request_url__icontains=search_value)
            )

        # Handle date filters
        from_date = request.POST.get('from_date', '')
        to_date = request.POST.get('to_date', '')
        if from_date:
            object_list = object_list.filter(created_at__date__gte=from_date)
        if to_date:
            object_list = object_list.filter(created_at__date__lte=to_date)

        # Handle vendor filter
        vendor_filter = request.POST.get('vendor_filter', '')
        if vendor_filter:
            object_list = object_list.filter(vendor=vendor_filter)

        # Handle status filter
        status_filter = request.POST.get('status_filter', '')
        if status_filter:
            if status_filter == 'success':
                object_list = object_list.filter(is_success=True)
            elif status_filter == 'failure':
                # Failed without reference number
                failed_list = object_list.filter(is_success=False)
                # Filter out those with reference numbers
                filtered_ids = []
                for api_response in failed_list:
                    ref_no = api_response.reference_no
                    if not (ref_no and ref_no != 'N/A' and ref_no.strip()):
                        filtered_ids.append(api_response.id)
                object_list = object_list.filter(id__in=filtered_ids)
            elif status_filter == 'failure_with_ref':
                # Failed with reference number
                failed_list = object_list.filter(is_success=False)
                # Filter those with reference numbers
                filtered_ids = []
                for api_response in failed_list:
                    ref_no = api_response.reference_no
                    if ref_no and ref_no != 'N/A' and ref_no.strip():
                        filtered_ids.append(api_response.id)
                object_list = object_list.filter(id__in=filtered_ids)

        # Handle ordering
        order_idx = int(request.POST.get('order[0][column]', 4))
        order_dir = request.POST.get('order[0][dir]', 'desc')
        order_field = column_map.get(order_idx)
        if order_field:
            object_list = object_list.order_by(
                order_field if order_dir == 'asc' else f'-{order_field}'
            )
        else:
            object_list = object_list.order_by('-created_at')

        # Apply pagination
        paginated_list = object_list[param_start:param_limit + param_start]
        filtered_total = object_list.count()

        # Prepare data for DataTables
        data = []
        for api_response in paginated_list:
            # Determine status type for filtering
            if api_response.is_success:
                status_badge = "Success"
                status_class = "bg-green-100 text-green-800"
                status_type = "success"
            else:
                # Check if failure has reference number
                ref_no = api_response.reference_no
                if ref_no and ref_no != 'N/A' and ref_no.strip():
                    status_badge = "Failed with Ref"
                    status_class = "bg-orange-100 text-orange-800"
                    status_type = "failure_with_ref"
                else:
                    status_badge = "Failed"
                    status_class = "bg-red-100 text-red-800"
                    status_type = "failure"

            data.append({
                'id': api_response.id,
                'awb_no': api_response.awb.awbno if api_response.awb else 'N/A',
                'vendor': api_response.vendor,
                'status': f'<span class="px-2 py-1 rounded-full text-xs font-medium {status_class}">{status_badge}</span>',
                'status_type': status_type,  # For filtering
                'created_at': api_response.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(api_response, 'created_at') else 'N/A',
                'reference_no': api_response.reference_no or 'N/A',
                'request_url': api_response.request_url[:50] + '...' if api_response.request_url and len(api_response.request_url) > 50 else (api_response.request_url or 'N/A'),
                'full_request_url': api_response.request_url or 'N/A',
                'response': api_response.response,
                'payload': api_response.payload,
                'has_label': bool(api_response.label),
                'has_pdf': bool(api_response.pdf),
                'updated_at': api_response.updated_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(api_response, 'updated_at') else 'N/A',
            })

        context = {
            'draw': post_dict.get('draw'),
            'recordsTotal': count,
            'recordsFiltered': filtered_total,
            'data': data,
        }

        return HttpResponse(
            mark_safe(json.dumps(context, indent=4,
                      sort_keys=True, default=str)),
            content_type='application/json'
        )

    # Handle regular HTML request
    return render(request, 'awb/api_responses/list.html', {
        "title": "AWB API Responses",
        "current": "awb_api_responses"
    })


@login_required
@has_roles(['admin'])
def export_api_responses(request):
    # export to excel
    # get all api responses
    api_responses = AWBAPIResponse.objects.select_related('awb').all()

    # Handle date filters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    if from_date:
        api_responses = api_responses.filter(created_at__date__gte=from_date)
    if to_date:
        api_responses = api_responses.filter(created_at__date__lte=to_date)

    # Handle vendor filter
    vendor = request.GET.get('vendor')
    if vendor:
        api_responses = api_responses.filter(vendor=vendor)

    # Handle status filter
    status = request.GET.get('status')
    if status:
        # ""
        # "success"
        # "failure"
        # "failure_with_ref"
        if status == "success":
            api_responses = api_responses.filter(is_success=True)
        elif status == "failure":
            # Failed without reference number - need to filter in Python
            failed_responses = api_responses.filter(is_success=False)
            filtered_ids = []
            for api_response in failed_responses:
                ref_no = api_response.reference_no
                if not (ref_no and ref_no != 'N/A' and ref_no.strip()):
                    filtered_ids.append(api_response.id)
            api_responses = api_responses.filter(id__in=filtered_ids)
        elif status == "failure_with_ref":
            # Failed with reference number - need to filter in Python
            failed_responses = api_responses.filter(is_success=False)
            filtered_ids = []
            for api_response in failed_responses:
                ref_no = api_response.reference_no
                if ref_no and ref_no != 'N/A' and ref_no.strip():
                    filtered_ids.append(api_response.id)
            api_responses = api_responses.filter(id__in=filtered_ids)
        else:
            api_responses = api_responses.filter(is_success=False)

    # export to excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="api_responses.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "API Responses"

    # Define headers
    headers = [
        "Created At",
        "Reference No",
        "AWB Number",
        "Vendor",
        "Status",
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Write data
    for row_num, api_response in enumerate(api_responses, 2):
        worksheet.cell(row=row_num, column=1, value=api_response.created_at.strftime(
            '%Y-%m-%d %H:%M:%S') if api_response.created_at else '')
        worksheet.cell(row=row_num, column=2,
                       value=api_response.reference_no)
        worksheet.cell(row=row_num, column=3,
                       value=api_response.awb.awbno if api_response.awb else 'N/A')
        worksheet.cell(row=row_num, column=4, value=api_response.vendor)
        worksheet.cell(row=row_num, column=5,
                       value='Success' if api_response.is_success else 'Failure')

    # Save workbook to response
    workbook.save(response)
    return response
