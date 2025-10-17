from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import Workbook
from django.http import HttpResponse
from io import BytesIO
import inflect


def convert_to_words_with_cents(amount):
    """Convert numeric amount to words with cents"""
    p = inflect.engine()

    whole_number = int(amount)
    cents = round((amount - whole_number) * 100)

    whole_number_in_words = p.number_to_words(
        whole_number).replace(",", "").upper()
    total_amount_in_words = ""
    if whole_number > 0:
        total_amount_in_words += whole_number_in_words
    if cents > 0:
        cents_in_words = p.number_to_words(cents).replace(",", "").upper()
        total_amount_in_words += f" AND {cents_in_words} CENTS"
    total_amount_in_words += " USD ONLY"
    total_amount_in_words = total_amount_in_words.upper()
    # remove any special characters ,*/-#$
    total_amount_in_words = total_amount_in_words.replace(
        ",", " ").replace("*", " ").replace("/", " ").replace("-", " ").replace("#", " ").replace("$", " ").replace(".", " ").replace("  ", " ")
    return total_amount_in_words


def create_invoice_excel(invoice_data, filename=None):
    """
    Create standardized invoice Excel file using common design.

    Args:
        invoice_data (dict): Dictionary containing all necessary invoice data
        filename (str): Optional filename for the response

    Returns:
        HttpResponse: Excel file response
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # === Define Styles ===
    bold = Font(bold=True)
    bold_large = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    large_pt_side = Side(style="thin", color="000000")
    thick_border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=large_pt_side,
        bottom=large_pt_side
    )

    # Row tracker
    row = 1

    # === 1) Main Header ===
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws.cell(row=row, column=1, value="INVOICE & PACKING LIST")
    cell.font = bold_large
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thick_border
    # Add borders to merged cells
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = thick_border
    row += 1

    # === 2) Top Info Section ===
    # Row 1
    ws.cell(
        row=row,
        column=1,
        value="COUNTRY OF ORIGIN: NEPAL"
    ).border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=large_pt_side,
        bottom=Side(style=None)
    )
    ws.merge_cells(f'A{row}:C{row}')

    ws.cell(row=row, column=4,
            value=f"ACTUAL WEIGHT :{invoice_data.get('actual_weight', '')}").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=large_pt_side,
        bottom=Side(style=None)
    )
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # Row 2 - Invoice Date
    ws.cell(row=row, column=1,
            value=f"INVOICE DATE. : {invoice_data.get('invoice_date', '')}").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=Side(style=None)
    )
    ws.merge_cells(f'A{row}:C{row}')

    ws.cell(row=row, column=4,
            value=f"TOTAL PIECES :{invoice_data.get('total_pieces', '')}").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=Side(style=None)
    )
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # Row 3 - Invoice Number
    ws.cell(row=row, column=1,
            value=f"INVOICE NO: {invoice_data.get('invoice_no', '')}").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=Side(style=None)
    )
    ws.merge_cells(f'A{row}:C{row}')

    ws.cell(row=row, column=4, value="").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=Side(style=None)
    )
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # Row 4 - Empty row
    ws.cell(row=row, column=1, value="").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=large_pt_side
    )
    ws.merge_cells(f'A{row}:C{row}')

    ws.cell(row=row, column=4, value="").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=large_pt_side
    )
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # === 3) Shipper/Consignee Section ===
    # Headers
    ws.cell(row=row, column=1, value="SHIPPER").font = bold
    ws.cell(row=row, column=1).border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=large_pt_side,
        bottom=Side(style=None)
    )
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=4, value="CONSIGNEE").font = bold
    ws.cell(row=row, column=4).border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=large_pt_side,
        bottom=Side(style=None)
    )
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # Shipper and Consignee Details
    shipper_info = invoice_data.get('shipper_info', [])
    consignee_info = invoice_data.get('consignee_info', [])

    for s_info, c_info in zip(shipper_info, consignee_info):
        cell_s = ws.cell(row=row, column=1, value=s_info)
        cell_s.border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=Side(style=None),
            bottom=Side(style=None)
        )
        cell_s.alignment = left
        ws.merge_cells(f'A{row}:C{row}')

        cell_c = ws.cell(row=row, column=4, value=c_info)
        cell_c.border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=Side(style=None),
            bottom=Side(style=None)
        )
        cell_c.alignment = left
        ws.merge_cells(f'D{row}:H{row}')
        row += 1

    # Empty row
    ws.cell(row=row, column=1, value="").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=large_pt_side
    )
    ws.merge_cells(f'A{row}:C{row}')

    ws.cell(row=row, column=4, value="").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=large_pt_side
    )
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # === 4) Items Table ===
    # Table Headers
    headers = ["BOXES", "SR NO", "DESCRIPTION", "HS CODE",
               "UNIT TYPE", "QUANTITY", "UNIT RATE", "AMOUNT (USD)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = bold
        cell.alignment = center
        cell.border = thick_border
    row += 1

    # Table Content
    sr_no = 1
    box_data = invoice_data.get('boxes', [])

    for box in box_data:
        box_items = box.get('items', [])
        start_row = row
        box_number = box.get('box_number', 'BOX')

        # Write box number once
        box_cell = ws.cell(row=start_row, column=1, value=f"BOX{box_number}")
        box_cell.alignment = center
        box_cell.border = thick_border

        for item in box_items:
            # Item details
            data = [
                None,  # Box number handled separately
                sr_no,
                item.get('description', ''),
                item.get('hs_code', ''),
                item.get('unit_type', ''),
                item.get('quantity', ''),
                item.get('unit_rate', ''),
                item.get('amount', '')
            ]

            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.alignment = center
                cell.border = thick_border

            sr_no += 1
            row += 1

        # Merge box number cells if there are multiple items
        if len(box_items) > 1:
            ws.merge_cells(start_row=start_row, start_column=1,
                           end_row=row-1, end_column=1)

    # === 5) Auto-adjust Column Widths and Row Heights ===
    # Auto-fit all column widths
    for column in range(1, 9):  # Columns A through H
        max_length = 0
        column_letter = get_column_letter(column)

        # Check all cells in the column
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=column)
            try:
                if cell.value:
                    # Handle different types of values and calculate proper width
                    cell_value = str(cell.value).strip()

                    # For wrapped text, consider line breaks
                    if '\n' in cell_value:
                        lines = cell_value.split('\n')
                        cell_length = max(len(line) for line in lines)
                    else:
                        cell_length = len(cell_value)

                    max_length = max(max_length, cell_length)
            except:
                continue

        # Set width with padding but no fixed limits
        adjusted_width = max_length + 3 if max_length > 0 else 10
        ws.column_dimensions[column_letter].width = adjusted_width

    # Auto-fit row heights
    for row_num in range(1, ws.max_row + 1):
        max_height = 0  # Start with 0 to let content determine height

        for col_num in range(1, 9):  # Check all columns in each row
            cell = ws.cell(row=row_num, column=col_num)

            if cell.value:
                try:
                    cell_value = str(cell.value).strip()

                    # Calculate height based on text length and column width
                    column_letter = get_column_letter(col_num)
                    column_width = ws.column_dimensions[column_letter].width or 10

                    # Estimate lines needed based on text length and column width
                    if len(cell_value) > column_width:
                        estimated_lines = max(
                            1, len(cell_value) // int(column_width))
                        # Add extra lines for explicit line breaks
                        if '\n' in cell_value:
                            estimated_lines += cell_value.count('\n')

                        # Calculate height (approximately 15 points per line)
                        calculated_height = estimated_lines * 15
                        max_height = max(max_height, calculated_height)

                    # Handle explicit line breaks
                    elif '\n' in cell_value:
                        line_count = cell_value.count('\n') + 1
                        calculated_height = line_count * 15
                        max_height = max(max_height, calculated_height)
                    else:
                        # Single line content
                        max_height = max(max_height, 15)

                except:
                    continue

        # Set the row height based on content (minimum 15 for readability)
        if max_height > 0:
            ws.row_dimensions[row_num].height = max(15, max_height)
        else:
            ws.row_dimensions[row_num].height = 15

    # === 6) Add Footer Section ===
    # Calculate totals
    total_quantity = invoice_data.get('total_quantity', 0)
    grand_total = invoice_data.get('grand_total', 0)

    # Add Total Quantity row
    ws.cell(row=row, column=1, value="Total Quantity").border = thick_border
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=6, value=total_quantity).border = thick_border
    ws.cell(row=row, column=7, value="Grand Total").border = thick_border
    ws.cell(row=row, column=8, value=grand_total).border = thick_border
    row += 1

    grandtotal_in_words = convert_to_words_with_cents(grand_total)

    # Add Total Amount row
    ws.cell(row=row, column=1, value=grandtotal_in_words).border = thick_border
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=7,
            value=f"Total:{grand_total}").border = thick_border
    ws.merge_cells(f'G{row}:H{row}')
    row += 1

    # Add Notes Section
    ws.cell(row=row, column=1, value="NOTES").border = thick_border
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=4, value="SIGNATURE / STAMP").border = thick_border
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # Add Declaration Note
    declaration_cell = ws.cell(
        row=row,
        column=1,
        value="WE DECLARE THAT THE ABOVE MENTIONED GOODS ARE MADE IN NEPAL AND OTHER DESCRIPTIONS ARE TRUE."
    )
    declaration_cell.border = thick_border
    declaration_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(f'A{row}:C{row}')

    signature_cell = ws.cell(row=row, column=4, value="")
    signature_cell.border = thick_border
    signature_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # Final adjustment for the last few rows that were added after the main auto-fit
    for row_num in range(row - 4, row):
        if row_num > 0:
            max_height = 0
            for col_num in range(1, 9):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    try:
                        cell_value = str(cell.value).strip()
                        column_letter = get_column_letter(col_num)
                        column_width = ws.column_dimensions[column_letter].width or 10

                        if len(cell_value) > column_width:
                            estimated_lines = max(
                                1, len(cell_value) // int(column_width))
                            calculated_height = estimated_lines * 15
                            max_height = max(max_height, calculated_height)
                        elif '\n' in cell_value:
                            line_count = cell_value.count('\n') + 1
                            calculated_height = line_count * 15
                            max_height = max(max_height, calculated_height)
                        else:
                            max_height = max(max_height, 15)
                    except:
                        continue

            if max_height > 0:
                ws.row_dimensions[row_num].height = max(15, max_height)
            else:
                ws.row_dimensions[row_num].height = 15

    # Manual column width adjustments
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 48

    # === Page Setup: IMPROVED horizontal centering ===
    # Set page orientation to portrait (default)
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    # Enable fit to page
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToPage = True

    # IMPORTANT: Center horizontally on the printed page
    ws.print_options.horizontalCentered = True
    # Only center horizontally, not vertically
    ws.print_options.verticalCentered = False

    # Alternative method - also set page_setup horizontal centering
    ws.page_setup.horizontalCentered = True

    # Set custom margins (in inches)
    ws.page_margins = PageMargins(
        left=0.2,
        right=0.2,
        top=0.2,
        bottom=0.2,
        header=0.2,
        footer=0.2
    )

    # Export
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename={filename or "invoice.xlsx"}'
    return response


def create_multi_sheet_invoice_excel(invoices_data, filename=None):
    """
    Create Excel file with multiple sheets for multiple invoices.

    Args:
        invoices_data (list): List of invoice data dictionaries
        filename (str): Optional filename for the response

    Returns:
        HttpResponse: Excel file response
    """
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Define styles (same as single sheet)
    bold = Font(bold=True)
    bold_large = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    large_pt_side = Side(style="thin", color="000000")
    thick_border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=large_pt_side,
        bottom=large_pt_side
    )

    # Create a sheet for each invoice
    for idx, invoice_data in enumerate(invoices_data):
        sheet_name = f"AWB_{invoice_data.get('invoice_no', f'Invoice_{idx+1}')}"
        ws = wb.create_sheet(title=sheet_name)

        # Apply the same logic as create_invoice_excel but to this specific worksheet
        _populate_invoice_sheet(
            ws, invoice_data, bold, bold_large, center, left, thick_border, large_pt_side)

    # Export
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename={filename or "invoices.xlsx"}'
    return response


def _populate_invoice_sheet(ws, invoice_data, bold, bold_large, center, left, thick_border, large_pt_side):
    """Helper function to populate a single worksheet with invoice data."""
    row = 1

    # === 1) Main Header ===
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws.cell(row=row, column=1, value="INVOICE & PACKING LIST")
    cell.font = bold_large
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thick_border
    # Add borders to merged cells
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = thick_border
    row += 1

    # === 2) Top Info Section ===
    # Row 1
    ws.cell(
        row=row,
        column=1,
        value="COUNTRY OF ORIGIN: NEPAL"
    ).border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=large_pt_side,
        bottom=Side(style=None)
    )
    ws.merge_cells(f'A{row}:C{row}')

    ws.cell(row=row, column=4,
            value=f"ACTUAL WEIGHT :{invoice_data.get('actual_weight', '')}").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=large_pt_side,
        bottom=Side(style=None)
    )
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # Row 2 - Invoice Date
    ws.cell(row=row, column=1,
            value=f"INVOICE DATE. : {invoice_data.get('invoice_date', '')}").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=Side(style=None)
    )
    ws.merge_cells(f'A{row}:C{row}')

    ws.cell(row=row, column=4,
            value=f"TOTAL PIECES :{invoice_data.get('total_pieces', '')}").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=Side(style=None)
    )
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # Row 3 - Invoice Number
    ws.cell(row=row, column=1,
            value=f"INVOICE NO: {invoice_data.get('invoice_no', '')}").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=Side(style=None)
    )
    ws.merge_cells(f'A{row}:C{row}')

    ws.cell(row=row, column=4, value="").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=Side(style=None)
    )
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # Row 4 - Empty row
    ws.cell(row=row, column=1, value="").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=large_pt_side
    )
    ws.merge_cells(f'A{row}:C{row}')

    ws.cell(row=row, column=4, value="").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=large_pt_side
    )
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # === 3) Shipper/Consignee Section ===
    # Headers
    ws.cell(row=row, column=1, value="SHIPPER").font = bold
    ws.cell(row=row, column=1).border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=large_pt_side,
        bottom=Side(style=None)
    )
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=4, value="CONSIGNEE").font = bold
    ws.cell(row=row, column=4).border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=large_pt_side,
        bottom=Side(style=None)
    )
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # Shipper and Consignee Details
    shipper_info = invoice_data.get('shipper_info', [])
    consignee_info = invoice_data.get('consignee_info', [])

    for s_info, c_info in zip(shipper_info, consignee_info):
        cell_s = ws.cell(row=row, column=1, value=s_info)
        cell_s.border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=Side(style=None),
            bottom=Side(style=None)
        )
        cell_s.alignment = left
        ws.merge_cells(f'A{row}:C{row}')

        cell_c = ws.cell(row=row, column=4, value=c_info)
        cell_c.border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=Side(style=None),
            bottom=Side(style=None)
        )
        cell_c.alignment = left
        ws.merge_cells(f'D{row}:H{row}')
        row += 1

    # Empty row
    ws.cell(row=row, column=1, value="").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=large_pt_side
    )
    ws.merge_cells(f'A{row}:C{row}')

    ws.cell(row=row, column=4, value="").border = Border(
        left=large_pt_side,
        right=large_pt_side,
        top=Side(style=None),
        bottom=large_pt_side
    )
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # === 4) Items Table ===
    # Table Headers
    headers = ["BOXES", "SR NO", "DESCRIPTION", "HS CODE",
               "UNIT TYPE", "QUANTITY", "UNIT RATE", "AMOUNT (USD)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = bold
        cell.alignment = center
        cell.border = thick_border
    row += 1

    # Table Content
    sr_no = 1
    box_data = invoice_data.get('boxes', [])

    for box in box_data:
        box_items = box.get('items', [])
        start_row = row
        box_number = box.get('box_number', 'BOX')

        # Write box number once
        box_cell = ws.cell(row=start_row, column=1, value=f"BOX{box_number}")
        box_cell.alignment = center
        box_cell.border = thick_border

        for item in box_items:
            # Item details
            data = [
                None,  # Box number handled separately
                sr_no,
                item.get('description', ''),
                item.get('hs_code', ''),
                item.get('unit_type', ''),
                item.get('quantity', ''),
                item.get('unit_rate', ''),
                item.get('amount', '')
            ]

            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.alignment = center
                cell.border = thick_border

            sr_no += 1
            row += 1

        # Merge box number cells if there are multiple items
        if len(box_items) > 1:
            ws.merge_cells(start_row=start_row, start_column=1,
                           end_row=row-1, end_column=1)

    # === 5) Auto-adjust Column Widths and Row Heights ===
    # Auto-fit all column widths
    for column in range(1, 9):  # Columns A through H
        max_length = 0
        column_letter = get_column_letter(column)

        # Check all cells in the column
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=column)
            try:
                if cell.value:
                    # Handle different types of values and calculate proper width
                    cell_value = str(cell.value).strip()

                    # For wrapped text, consider line breaks
                    if '\n' in cell_value:
                        lines = cell_value.split('\n')
                        cell_length = max(len(line) for line in lines)
                    else:
                        cell_length = len(cell_value)

                    max_length = max(max_length, cell_length)
            except:
                continue

        # Set width with padding but no fixed limits
        adjusted_width = max_length + 3 if max_length > 0 else 10
        ws.column_dimensions[column_letter].width = adjusted_width

    # Auto-fit row heights
    for row_num in range(1, ws.max_row + 1):
        max_height = 0  # Start with 0 to let content determine height

        for col_num in range(1, 9):  # Check all columns in each row
            cell = ws.cell(row=row_num, column=col_num)

            if cell.value:
                try:
                    cell_value = str(cell.value).strip()

                    # Calculate height based on text length and column width
                    column_letter = get_column_letter(col_num)
                    column_width = ws.column_dimensions[column_letter].width or 10

                    # Estimate lines needed based on text length and column width
                    if len(cell_value) > column_width:
                        estimated_lines = max(
                            1, len(cell_value) // int(column_width))
                        # Add extra lines for explicit line breaks
                        if '\n' in cell_value:
                            estimated_lines += cell_value.count('\n')

                        # Calculate height (approximately 15 points per line)
                        calculated_height = estimated_lines * 15
                        max_height = max(max_height, calculated_height)

                    # Handle explicit line breaks
                    elif '\n' in cell_value:
                        line_count = cell_value.count('\n') + 1
                        calculated_height = line_count * 15
                        max_height = max(max_height, calculated_height)
                    else:
                        # Single line content
                        max_height = max(max_height, 15)

                except:
                    continue

        # Set the row height based on content (minimum 15 for readability)
        if max_height > 0:
            ws.row_dimensions[row_num].height = max(15, max_height)
        else:
            ws.row_dimensions[row_num].height = 15

    # === 6) Add Footer Section ===
    # Calculate totals
    total_quantity = invoice_data.get('total_quantity', 0)
    grand_total = invoice_data.get('grand_total', 0)

    # Add Total Quantity row
    ws.cell(row=row, column=1, value="Total Quantity").border = thick_border
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=6, value=total_quantity).border = thick_border
    ws.cell(row=row, column=7, value="Grand Total").border = thick_border
    ws.cell(row=row, column=8, value=grand_total).border = thick_border
    row += 1

    grandtotal_in_words = convert_to_words_with_cents(grand_total)

    # Add Total Amount row
    ws.cell(row=row, column=1, value=grandtotal_in_words).border = thick_border
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=7,
            value=f"Total:{grand_total}").border = thick_border
    ws.merge_cells(f'G{row}:H{row}')
    row += 1

    # Add Notes Section
    ws.cell(row=row, column=1, value="NOTES").border = thick_border
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=4, value="SIGNATURE / STAMP").border = thick_border
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # Add Declaration Note
    declaration_cell = ws.cell(
        row=row,
        column=1,
        value="WE DECLARE THAT THE ABOVE MENTIONED GOODS ARE MADE IN NEPAL AND OTHER DESCRIPTIONS ARE TRUE."
    )
    declaration_cell.border = thick_border
    declaration_cell.alignment = Alignment(vertical="center")
    ws.merge_cells(f'A{row}:C{row}')

    signature_cell = ws.cell(row=row, column=4, value="")
    signature_cell.border = thick_border
    signature_cell.alignment = Alignment(vertical="center")
    ws.merge_cells(f'D{row}:H{row}')
    row += 1

    # Final adjustment for the last few rows that were added after the main auto-fit
    for row_num in range(row - 4, row):
        if row_num > 0:
            max_height = 0
            for col_num in range(1, 9):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    try:
                        cell_value = str(cell.value).strip()
                        column_letter = get_column_letter(col_num)
                        column_width = ws.column_dimensions[column_letter].width or 10

                        if len(cell_value) > column_width:
                            estimated_lines = max(
                                1, len(cell_value) // int(column_width))
                            calculated_height = estimated_lines * 15
                            max_height = max(max_height, calculated_height)
                        elif '\n' in cell_value:
                            line_count = cell_value.count('\n') + 1
                            calculated_height = line_count * 15
                            max_height = max(max_height, calculated_height)
                        else:
                            max_height = max(max_height, 15)
                    except:
                        continue

            if max_height > 0:
                ws.row_dimensions[row_num].height = max(15, max_height)
            else:
                ws.row_dimensions[row_num].height = 15

    # Manual column width adjustments
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 48

    # === Page Setup: IMPROVED horizontal centering ===
    # Set page orientation to portrait (default)
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    # Enable fit to page
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToPage = True

    # IMPORTANT: Center horizontally on the printed page
    ws.print_options.horizontalCentered = True
    # Only center horizontally, not vertically
    ws.print_options.verticalCentered = False

    # Alternative method - also set page_setup horizontal centering
    ws.page_setup.horizontalCentered = True

    # Set custom margins (in inches)
    ws.page_margins = PageMargins(
        left=0.2,
        right=0.2,
        top=0.2,
        bottom=0.2,
        header=0.2,
        footer=0.2
    )
