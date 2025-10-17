from hub.models import VendorLoginCred
import requests
from finance.models import Ledger
import base64
from django.db.models import Exists, OuterRef, Q
from awb.models import verify_awb_validation
from django.db.models import Sum, Prefetch
from openpyxl import Workbook
from openpyxl.styles import Font
from .forms import AWBStatusForm, UpdateBoxAWBNumberForm, UpdateForwardingNumberForm
from io import BytesIO
from django import forms
import pandas as pd
from django.utils import timezone
from datetime import timedelta
from django.db import models
from django.db.models import Count, Sum
from itertools import chain
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse
from hub.models.run import RunAWB
from django.core.exceptions import ValidationError
from .utils import AWBInvoiceExporter
from django.views.decorators.csrf import csrf_exempt
from django.forms import formset_factory, modelformset_factory
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import AWBForm, ConsigneeForm, ConsignorForm, BoxDetailsForm, BoxItemForm, BoxUpdateForm, BoxItemUpdateForm, ForwardinNumberUploadForm
from django.http import JsonResponse, HttpResponse, QueryDict
from django.db.models import Q
from awb.models import UnitType, BoxDetails, BoxItem, AWBAPIResponse
from accounts.models import Agency, DividingFactor, Company
from accounts.models import Agency, DividingFactor, Country

from django.shortcuts import get_object_or_404
from django.forms.models import model_to_dict
from django.db import transaction
from awb.models import AWBDetail, Consignee, Consignor
from django.shortcuts import render, get_object_or_404
from django.utils.safestring import mark_safe
import json
import re
from decorators import has_roles
from django.contrib.auth.decorators import login_required
from awb.models import AWBStatus
from .forms import AWBUpdateForm


@login_required
@has_roles(['admin', 'agencyuser'])
def print_awb(request, awb_no):
    awb = get_object_or_404(AWBDetail, awbno=awb_no)
    if request.user.role == 'agencyuser' and awb.agency != request.user.agency:
        messages.error(request, "You don't have permission to print this AWB.")
        return redirect("awb:pages:awb:list")
    consignee = Consignee.objects.get(awb=awb)
    consignor = Consignor.objects.get(awb=awb)
    box_details = BoxDetails.objects.filter(awb=awb)
    box_items = BoxItem.objects.filter(box__awb=awb)
    print(box_items)
    context = {
        "awb": awb,
        "consignee": consignee,
        "consignor": consignor,
        "box_details": box_details,
        "box_items": box_items,
        "title": "Shipment"
    }
    return render(request, 'awb/awb/print.html', context)


@login_required
@has_roles(['admin', 'agencyuser'])
def print_label1(request, awb_no):
    awb = get_object_or_404(AWBDetail, awbno=awb_no)
    base64_label1 = awb.label_1

    if not base64_label1:
        messages.error(request, "Label1 not found")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    if awb.vendor.code == "SGUS":
        # convertinto pdf
        pdf_label1 = base64.b64decode(base64_label1)
        return HttpResponse(pdf_label1, content_type='application/pdf')

    elif awb.vendor.code != "SGCA":
        label1 = base64.b64decode(base64_label1)
        return HttpResponse(label1, content_type='image/png')

    else:
        html_content = base64.b64decode(base64_label1).decode('utf-8')
        return HttpResponse(html_content, content_type='text/html')


@login_required
@has_roles(['admin', 'agencyuser'])
def print_box_label1(request, awb_no):
    awb = get_object_or_404(AWBDetail, awbno=awb_no)
    box_details = BoxDetails.objects.filter(awb=awb)
    box_items = BoxItem.objects.filter(box__awb=awb)
    context = {
        "awb": awb,
        "box_details": box_details,
        "box_items": box_items,
        "title": "Box Label"
    }
    return render(request, 'awb/awb/box_label1.html', context)


@login_required
@has_roles(['admin', 'agencyuser'])
def print_ubx_label(request, awb_no):
    awb = get_object_or_404(AWBDetail, awbno=awb_no)
    box_details = BoxDetails.objects.filter(awb=awb)
    vendor = awb.vendor.code
    service_code = awb.service.code
    if (vendor == "UBX" and service_code.strip() == "DPD") or (vendor == "UBX" and service_code.strip() == "DPDG") or (vendor == "UBX" and service_code.strip() == "DHL"):
        base64_pdf_label = awb.label_1
        pdf_label = base64.b64decode(base64_pdf_label)
        return HttpResponse(pdf_label, content_type='application/pdf')
    else:
        context = {
            "awb": awb,
            "box_details": box_details
        }
        return render(request, 'awb/awb/box_label1.html', context)


@login_required
@has_roles(['admin', 'agencyuser'])
def print_courierx_label(request, awb_no):

    awb = get_object_or_404(AWBDetail, awbno=awb_no)
    service = awb.service.name

    if service == "UPS":
        box_details = BoxDetails.objects.filter(awb=awb)
        context = {
            "awb": awb,
            "box_details": box_details
        }
        return render(request, 'awb/awb/box_label1.html', context)

    else:
        url = "https://ideal.couriex.com/Service.svc/AirwaybillPDFFormat"
        cred = VendorLoginCred.objects.get(vendor__code="COURIERX")
        payload = {
            "AirwayBillNumber": f"{awb.reference_number}",
            "Country": "AE",
            "RequestUser": "ideal",
            "UserName": cred.username,
            "Password": cred.password,
            "AccountNo": cred.additional_cred1,
            "PrintType": "A4"
        }
        response = requests.post(url, json=payload, timeout=60)
        if response.status_code == 200:
            base64_label1 = response.json()["ReportDoc"]
            label1 = base64.b64decode(base64_label1)
            return HttpResponse(label1, content_type='application/pdf')
        elif response.status_code == 400:
            messages.error(
                request, "Error generating label. Please check the AWB number and try again.")
            return redirect(request.META.get('HTTP_REFERER', '/'))
        else:
            messages.error(
                request, "Error generating label. Please check the AWB number and try again.")
            return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
@has_roles(['admin', 'agencyuser'])
def verify_awb(request, awb_no):
    try:
        role = request.user.role
        if role == 'agencyuser':
            agency = request.user.agency
            if not agency.can_verify_awb:
                messages.error(
                    request, "You don't have permission to verify this AWB.")
                return redirect(request.META.get('HTTP_REFERER', '/'))
        awb = get_object_or_404(AWBDetail, awbno=awb_no)
        verify_awb_validation(awb)
        awb.is_verified = True
        awb.save()
        messages.success(request, "AWB verified successfully!")
    except Exception as e:
        messages.error(request, "Error verifying AWB: " + str(e))
    return redirect(request.META.get('HTTP_REFERER'))


@login_required
@has_roles(['admin'])
def unverify_awb(request, awb_no):
    try:
        awb = get_object_or_404(AWBDetail, awbno=awb_no)
        if RunAWB.objects.filter(awb=awb).exists():
            messages.error(request, "Cannot unverify AWB that is in run.")
            return redirect(request.META.get('HTTP_REFERER', '/'))
        awb.is_verified = False
        awb.save()
        messages.success(request, "AWB unverified successfully!")
    except ValidationError as e:
        # Check if AWB is in a run
        run_awb = RunAWB.everything.select_related(
            'run').filter(awb__awbno=awb_no).first()

        if run_awb:
            run_url = reverse('hub:pages:run:add_awb', kwargs={
                              'run_id': run_awb.run.id})
            error_msg = mark_safe(
                f"Cannot unverify AWB that is in <a href='{run_url}' class='underline text-white-600 hover:text-white-800' target='_blank'>Run #{run_awb.run.run_no}</a>")
            messages.error(request, error_msg)
        else:
            messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
    return redirect(request.META.get('HTTP_REFERER'))


@login_required
@has_roles(['admin', 'agencyuser'])
def shipment_detail(request, awb_no):
    # try:
    awb = (
        AWBDetail.objects
        .select_related(
            # Assuming AWB has FK to these
            'consignee', 'consignor', 'agency'
        )
        .prefetch_related(
            Prefetch('boxdetails', queryset=BoxDetails.objects.all(),
                     to_attr='prefetched_box_details'),
            Prefetch('boxdetails__items', queryset=BoxItem.objects.all(
            ), to_attr='prefetched_box_items'),
            Prefetch('awbstatus_set', queryset=AWBStatus.objects.all(),
                     to_attr='prefetched_status'),
            Prefetch('runawb_set', queryset=RunAWB.objects.select_related(
                'run'), to_attr='prefetched_runawbs'),
        )
        .get(awbno=awb_no)
    )

    next_awb = AWBDetail.objects.filter(
        id__gt=awb.id
    ).order_by('id').first()
    prev_awb = AWBDetail.objects.filter(
        id__lt=awb.id
    ).order_by('-id').first()
    if request.user.role == 'agencyuser' and awb.agency != request.user.agency:
        messages.error(
            request, "You don't have permission to view this AWB.")
        return redirect("awb:pages:awb:list")
    # Get box details and items
    box_details = getattr(awb, 'prefetched_box_details', [])
    box_items = []

    # Process box items for each box
    for box in box_details:
        box_items.extend(box.items.all())

    # Calculate totals
    totals = {
        'total_quantity': sum(item.quantity or 0 for item in box_items),
        'total_unit_weight': sum(item.unit_weight or 0 for item in box_items),
        'total_unit_rate': sum(item.unit_rate or 0 for item in box_items)
    }
    shipment_status = getattr(awb, 'prefetched_status', [])
    shipment_status_form = AWBStatusForm(request.POST or None, awb=awb)

    run = getattr(awb, 'prefetched_runawbs', [])

    forwarding_number_form = UpdateForwardingNumberForm(
        initial={'awbno': awb.awbno, 'forwarding_number': awb.forwarding_number if awb.forwarding_number else ''})
    UpdateBoxAWBNumberFormSet = forms.formset_factory(
        UpdateBoxAWBNumberForm, extra=len(box_details))
    UpdateBoxAWBNumberFormSet.initial = [
        {
            'box_id': box.id,
            'box_awb_no': box.box_awb_no
        }
        for box in box_details
    ]

    context = {
        "awb": awb,
        "consignee": awb.consignee,
        "consignor": awb.consignor,
        "box_details": box_details,
        "box_items": box_items,
        "shipment_status": shipment_status,
        "shipment_status_form": shipment_status_form,
        'next_awb': next_awb,
        'prev_awb': prev_awb,
        'run': run,
        'totals': totals,
        'forwarding_number_form': forwarding_number_form,
        'box_awb_number_formset': UpdateBoxAWBNumberFormSet,
        "title": "Shipment"
    }

    return render(request, "awb/awb/detail.html", context)
    # except Exception as e:
    #     messages.error(request, f"An error occurred: {e}")
    #     try:
    #         return redirect(request.META['HTTP_REFERER'])
    #     except Exception as e:
    #         return redirect('awb:pages:awb:list')


@login_required
@has_roles(['admin'])
def add_shipment_status(request, awb_no):
    awb = get_object_or_404(AWBDetail, awbno=awb_no)
    shipment_status_form = AWBStatusForm(request.POST or None, awb=awb)
    if request.method == "POST" and shipment_status_form.is_valid():
        try:
            with transaction.atomic():
                # Use the form's save method instead of manually creating
                shipment_status_form.save()
                messages.success(
                    request, "Shipment status updated successfully!")
                return redirect(request.META['HTTP_REFERER'])
        except Exception as e:
            messages.error(request, str(e))
    else:
        messages.error(
            request, "Invalid form data. Please check and try again.")
        messages.error(request, shipment_status_form.errors)
    return redirect(request.META['HTTP_REFERER'])


@login_required
@has_roles(['admin', 'agencyuser'])
def delete_shipment_status(request, status_id):
    shipment_status = AWBStatus.objects.get(id=status_id)
    awb_no = shipment_status.awb.awbno
    shipment_status.delete()
    messages.success(request, "Shipment status deleted successfully!")
    return redirect("awb:pages:awb:detail", awb_no=awb_no)


def redirect_to_awb_detail(request):
    awb_no = request.GET.get('awb_no').strip()
    try:
        awb_obj = get_object_or_404(AWBDetail, awbno=awb_no)
        if awb_obj.is_verified:
            return redirect('awb:pages:awb:update', awb_no=awb_no)
        return redirect('awb:pages:awb:update', awb_no=awb_no)
    except Exception:
        messages.error(request, f"AWB not found")
        return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
@has_roles(['admin', 'agencyuser'])
def shipment_list(request):
    shipment_without_invoice = request.GET.get('shipment_without_invoice')
    print(shipment_without_invoice)
    if re.search(r'^text/html', request.META.get('HTTP_ACCEPT', '')):
        # Determine active tab based on the filter
        active_tab = 'shipment_without_invoice' if shipment_without_invoice else 'shipment'
        return render(request, 'awb/awb/list.html', {
            "title": "Shipment",
            "shipment_without_invoice": shipment_without_invoice,
            "active_tab": active_tab
        })

    post_dict = QueryDict(request.POST.urlencode(), mutable=True)
    param_start = max(int(request.POST.get('start', 0)), 0)
    param_limit = max(int(request.POST.get('length', 50)), 1)

    column_map = {
        0: None,
        1: 'created_at',
        2: 'awbno',
        3: 'agency__company_name',
        4: 'consignee__person_name',
        5: 'destination__name',
        6: 'total_box',
        7: None,
        8: 'is_verified',
        9: 'forwarding_number',
        10: 'company__name',
        11: 'runawb__run__run_no',
        12: 'product_type__name',
        13: 'hub__name',
        14: 'origin__name',
        15: 'service__name',
        16: 'shipment_value',
        17: 'consignor__person_name',
        18: None,
        19: 'consignor__phone_number'
    }

    qs = AWBDetail.objects.select_related(
        'consignee', 'consignor', 'hub', 'company', 'origin', 'destination',
        'service', 'agency', 'product_type', 'runawb__run'
    ).only(
        'id', 'awbno', 'hub__name', 'company__name', 'origin__name',
        'destination__name', 'service__name', 'shipment_value',
        'consignee__person_name', 'consignor__person_name',
        'is_cancelled', 'created_at', 'agency__company_name', 'total_box',
        'is_verified', 'total_actual_weight', 'total_volumetric_weight',
        'total_charged_weight', 'forwarding_number', 'runawb__run__run_no',
        'product_type__name', 'consignor__company', 'consignor__phone_number',
        'is_invoice_generated'
    )
    if shipment_without_invoice == 'true':
        print("shipment_without_invoice")
        qs = qs.filter(is_invoice_generated=False)

    if getattr(request.user, 'role', '') == 'agencyuser':
        qs = qs.filter(agency=request.user.agency)

    search_val = request.POST.get('search[value]', '').strip()
    if search_val:
        qs = qs.filter(
            Q(awbno__icontains=search_val) |
            Q(hub__name__icontains=search_val) |
            Q(destination__name__icontains=search_val) |
            Q(service__name__icontains=search_val) |
            Q(consignee__person_name__icontains=search_val) |
            Q(consignor__person_name__icontains=search_val) |
            Q(shipment_value__icontains=search_val) |
            Q(company__name__icontains=search_val) |
            Q(consignor__phone_number__icontains=search_val)
        )

    filtered_total = qs.count()

    order_idx = int(request.POST.get('order[0][column]', 1))
    order_dir = request.POST.get('order[0][dir]', 'desc')
    order_field = column_map.get(order_idx)
    if order_field:
        qs = qs.order_by(order_field if order_dir ==
                         'asc' else f'-{order_field}')
    else:
        qs = qs.order_by('-created_at')

    for index, field in column_map.items():
        column_search_value = request.POST.get(
            f'columns[{index}][search][value]', '').strip()
        if column_search_value:
            kwargs = {f"{field}__icontains": column_search_value}
            qs = qs.filter(**kwargs)

    filtered_total = qs.count()
    # Always add created_at as secondary ordering for consistency
    departed_qs = AWBStatus.objects.filter(
        awb=OuterRef('pk'),
        status='SHIPMENT DEPARTURED'
    )
    qs = qs.annotate(has_departed=Exists(departed_qs))

    rows = qs[param_start:param_start + param_limit].values(
        'id', 'awbno', 'hub__name', 'company__name', 'origin__name',
        'destination__name', 'service__name', 'shipment_value',
        'consignee__person_name', 'consignor__person_name',
        'is_cancelled', 'created_at', 'agency__company_name', 'total_box',
        'is_verified', 'total_actual_weight', 'total_volumetric_weight',
        'total_charged_weight', 'forwarding_number', 'runawb__run__run_no',
        'product_type__name', 'consignor__company',
        'has_departed', 'consignor__phone_number'  # ← this annotation

    )

    return JsonResponse({
        'draw': int(post_dict.get('draw', 0)),
        'recordsTotal': AWBDetail.objects.count(),
        'recordsFiltered': filtered_total,
        'data': list(rows),
    }, safe=False)


@login_required
@has_roles(['admin', 'agencyuser'])
@csrf_exempt
def agency_detail(request, pk):
    # Retrieve the agency by primary key, or return a 404 if not found
    agency = get_object_or_404(Agency, pk=pk)
    try:
        leadger = Ledger.objects.filter(agency=agency).first()
        last_balance = leadger.get_last_balance(agency=agency)
    except Exception as e:
        last_balance = 0

    if request.user.role == "agencyuser" and agency != request.user.agency:
        messages.error(
            request, "You don't have permission to access this agency.")
        return JsonResponse({"error": "You don't have permission to access this agency."}, status=403)

    # Create a JSON-friendly dictionary for the agency's data
    agency_data = {
        "id": agency.id,
        "name": agency.company_name,
        "agency_code": agency.agency_code,
        "owner_name": agency.owner_name,
        "country_name": agency.country.name,
        "country_id": agency.country.id,
        "zip_code": agency.zip_code,
        "email": agency.email,
        "state": agency.state,  # Assuming agency.state exists
        "city": agency.city,
        "address1": agency.address1,
        "address2": agency.address2,
        "contact_no_1": agency.contact_no_1,
        "contact_no_2": agency.contact_no_2,
        "credit_limit": agency.credit_limit,
        "max_user": agency.max_user,
        "is_blocked": agency.is_blocked,
        "last_balance": last_balance
    }

    # Return the data as a JSON response
    return JsonResponse(agency_data)


@login_required
@has_roles(['admin', 'agencyuser'])
def create_shipment_view(request):
    awb_no = request.GET.get('awb_no', '')

    total_box_count = 1  # default if not specified
    if request.method == "POST":
        if request.user.role == 'agencyuser':
            try:
                last_balance = (
                    float(Ledger.objects.filter(agency=request.user.agency).first().get_last_balance(agency=request.user.agency)))*(-1)
            except Exception as e:
                last_balance = 0
            credit_limit = float(request.user.agency.credit_limit)
            if request.user.agency.is_blocked:
                messages.error(
                    request, "Your agency is blocked. Please contact the admin.")
                return redirect(request.META.get('HTTP_REFERER'))
            if last_balance > credit_limit:
                messages.error(
                    request, "Your agency has no credit limit. Please contact the admin.")
                return redirect(request.META.get('HTTP_REFERER'))
        post_data = request.POST.copy()
        if post_data.get("awb-is_custom", "off") != "on":
            post_data["awb_awbno"] = ""
        total_box_count = int(request.POST.get("total_box_count", 1))
        BoxDetailsFormSet = formset_factory(
            BoxDetailsForm, extra=total_box_count)
        BoxItemFormSet = formset_factory(BoxItemForm, extra=1)

        # Initialize forms and formsets with POST data
        awb_form = AWBForm(post_data, prefix="awb", user=request.user)
        consignee_form = ConsigneeForm(
            post_data, request.FILES, prefix="consignee")
        consignor_form = ConsignorForm(
            post_data, request.FILES, prefix="consignor")
        box_details_formset = BoxDetailsFormSet(
            post_data, prefix="box_details")
        # Create the box item formset (choices will be updated below)
        box_item_formset = BoxItemFormSet(post_data, prefix="box_item")
        create_shipment_invoice = request.POST.get(
            "create_shipment_invoice", False)

        # Determine available box choices.
        # If the box_details_formset is valid, use its cleaned_data; otherwise, fall back to default numbering.
        if box_details_formset.is_valid():
            # Only include forms with data.
            box_choices = [
                (form.cleaned_data.get('box_number'),
                 form.cleaned_data.get('box_number'))
                for form in box_details_formset if form.cleaned_data.get('box_number')
            ]
        else:
            # Fallback: generate default choices based on total_box_count.
            box_choices = [(f"Box-{i+1}", f"Box-{i+1}")
                           for i in range(total_box_count)]

        # Update each box item form with the available box choices.
        for form in box_item_formset:
            form.fields['box_number'].widget.choices = box_choices

        if (awb_form.is_valid() and
            consignee_form.is_valid() and
            consignor_form.is_valid() and
            box_details_formset.is_valid() and
                (not create_shipment_invoice or box_item_formset.is_valid())):
            try:
                with transaction.atomic():
                    awb = awb_form.save()
                    consignee = consignee_form.save(commit=False)
                    consignee.awb = awb
                    consignee.save()
                    consignor = consignor_form.save(commit=False)
                    consignor.awb = awb
                    consignor.save()
                    # Save box details and build a mapping for box items.
                    box_map = {}
                    for form in box_details_formset:
                        if form.cleaned_data:
                            box_detail = form.save(commit=False)
                            box_detail.awb = awb
                            box_detail.save()
                            box_number = form.cleaned_data.get('box_number')
                            box_map[box_number] = box_detail

                    print(f"Box map keys: {box_map.keys()}")

                    # Save box items if creating shipment    invoice.
                    if create_shipment_invoice and box_map:
                        print(
                            f"Box item formset data: {[form.cleaned_data for form in box_item_formset if form.cleaned_data]}")
                        for form in box_item_formset:
                            if form.cleaned_data:
                                selected_box = form.cleaned_data.get(
                                    'box_number')
                                if selected_box in box_map:
                                    box_item = form.save(commit=False)
                                    box_item.box = box_map[selected_box]
                                    box_item.save()
                                else:
                                    print(
                                        f"Box number {selected_box} not found in box_map: {box_map.keys()}")
                    messages.success(request, "Shipment created successfully!")
                    return redirect('awb:pages:awb:update', awb_no=awb.awbno)
            except Exception as e:
                messages.error(request, f"Error creating shipment: {e}")
        else:
            print(awb_form.errors)
            messages.error(request, "Please correct the errors below.")
    else:
        # For GET requests: use the default total_box_count.
        BoxDetailsFormSet = formset_factory(
            BoxDetailsForm, extra=total_box_count)
        BoxItemFormSet = formset_factory(BoxItemForm, extra=1)
        awb_form = AWBForm(prefix="awb", user=request.user)
        if request.user.role == 'agencyuser':
            awb_form.fields['agency'].queryset = Agency.objects.filter(
                id=request.user.agency.id)
            awb_form.fields['agency'].initial = request.user.agency
        consignee_form = ConsigneeForm(prefix="consignee")
        consignor_form = ConsignorForm(prefix="consignor")
        if awb_no:
            awb = get_object_or_404(AWBDetail, awbno=awb_no)
            consignee_form.initial = model_to_dict(awb.consignee)
            # Add country code to the initial data
            if awb.destination and awb.destination.code:
                consignee_form.initial['country_code'] = awb.destination.code
            consignor_form.initial = model_to_dict(awb.consignor)
            awb_initial = model_to_dict(awb)
            if 'awbno' in awb_initial:
                del awb_initial['awbno']
            awb_form.initial = awb_initial
        box_details_formset = BoxDetailsFormSet(prefix="box_details")
        # For GET, generate default box choices (e.g., if only one box is expected initially).
        box_choices = [(f"Box-{i+1}", f"Box-{i+1}")
                       for i in range(total_box_count)]
        box_item_formset = BoxItemFormSet(prefix="box_item")
        for form in box_item_formset:
            form.fields['box_number'].widget.choices = box_choices
        create_shipment_invoice = False
        print(box_item_formset.errors)

    context = {
        "awb_form": awb_form,
        "consignee_form": consignee_form,
        "consignor_form": consignor_form,
        "units": UnitType.objects.all(),
        "box_details_formset": box_details_formset,
        "total_box_count": total_box_count,
        "create_shipment_invoice": create_shipment_invoice,
        "box_item_formset": box_item_formset,
        'is_error': bool(awb_form.errors or consignee_form.errors or consignor_form.errors or box_details_formset.errors or box_item_formset.errors),
        'is_updating': False,
        'is_cloned': True if awb_no else False,
        "title": "Shipment"
    }
    return render(request, "awb/awb/create.html", context)


@login_required
@has_roles(['admin', 'agencyuser'])
def update_shipment_view(request, awb_no):
    total_box_count = 1  # default
    awb = get_object_or_404(
        AWBDetail.objects.prefetch_related('runawb_set', 'runawb_set__run'),
        awbno=awb_no
    )
    box_details = BoxDetails.objects.filter(awb=awb)
    box_items = BoxItem.objects.filter(box__awb=awb)
    box_details_count = box_details.count()
    box_items_count = box_items.count()

    # preventing agency user from updating awb after 5 minutes

    # Determine if shipment invoice should be created based on whether there are existing box details
    create_shipment_invoice = box_items_count > 0

    BoxDetailsFormSet = modelformset_factory(
        BoxDetails, form=BoxUpdateForm, extra=0, can_delete=True)
    BoxItemFormSet = modelformset_factory(
        BoxItem, form=BoxItemUpdateForm, extra=1, can_delete=True)

    if request.method == "POST":
        if awb.is_invoice_generated:
            messages.error(
                request, "You cannot update this AWB after it is billed.")
            return redirect("awb:pages:awb:detail", awb_no=awb_no)
        if request.user.role == 'agencyuser' and awb.created_at < timezone.now() - timedelta(minutes=30):
            messages.error(
                request, "You cannot update this AWB after 30 minutes.")
            return redirect("awb:pages:awb:detail", awb_no=awb_no)
        if not awb.is_editable:
            messages.error(request, "You cannot update this AWB.")
            return redirect("awb:pages:awb:update", awb_no=awb_no)
        total_box_count = int(request.POST.get("total_box_count", 1))

        awb_form = AWBUpdateForm(request.POST, prefix="awb", instance=awb)
        consignee_form = ConsigneeForm(
            request.POST, request.FILES, prefix="consignee", instance=awb.consignee)
        consignor_form = ConsignorForm(
            request.POST, request.FILES, prefix="consignor", instance=awb.consignor)
        box_details_formset = BoxDetailsFormSet(
            request.POST, prefix="box_details", queryset=box_details)
        box_item_formset = BoxItemFormSet(
            request.POST, prefix="box_item", queryset=box_items)

        create_shipment_invoice = request.POST.get(
            "create_shipment_invoice", False)

        # Set up choices for box numbering in BoxItem forms
        if box_details_formset.is_valid():
            box_choices = [(form.cleaned_data.get('box_number'),
                            form.cleaned_data.get('box_number'))
                           for form in box_details_formset if form.cleaned_data.get('box_number')]
        else:
            box_choices = [(f"Box-{i+1}", f"Box-{i+1}")
                           for i in range(total_box_count)]

        for form in box_item_formset:
            form.fields['box_number'].widget.choices = box_choices

        # if awb.is_verified:
        #     messages.error(
        #         request, "You cannot update this AWB after it is verified.")
        #     return redirect(request.META.get('HTTP_REFERER'))

        if (awb_form.is_valid() and
            consignee_form.is_valid() and
            consignor_form.is_valid() and
            box_details_formset.is_valid() and
                (not create_shipment_invoice or box_item_formset.is_valid())):

            try:
                with transaction.atomic():
                    # Save AWB and related consignee/consignor details
                    awb = awb_form.save()
                    consignee = consignee_form.save(commit=False)
                    consignee.awb = awb
                    consignee.save()
                    consignor = consignor_form.save(commit=False)
                    consignor.awb = awb
                    consignor.save()

                    # Create a mapping for box details and later assign BoxItems to them
                    box_and_box_item_detail = {}
                    for i, form in enumerate(box_details_formset):
                        if form.cleaned_data:
                            box_number = form.cleaned_data.get(
                                'box_number') or f"Box-{i+1}"
                            box_and_box_item_detail[box_number] = {
                                "length": form.cleaned_data.get('length'),
                                "breadth": form.cleaned_data.get('breadth'),
                                "height": form.cleaned_data.get('height'),
                                "actual_weight": form.cleaned_data.get('actual_weight'),
                                "box_awb_no": form.cleaned_data.get('box_awb_no'),
                                "bag_no": form.cleaned_data.get('bag_no'),
                                "items": []  # will be filled in next loop
                            }

                    if create_shipment_invoice:
                        # Map BoxItems to the corresponding box_number
                        for form in box_item_formset:
                            if form.cleaned_data and not form.cleaned_data.get('DELETE'):
                                box_number = form.cleaned_data.get(
                                    'box_number')
                                if box_number in box_and_box_item_detail:
                                    box_and_box_item_detail[box_number]["items"].append({
                                        "description": form.cleaned_data.get('description'),
                                        "hs_code": form.cleaned_data.get('hs_code'),
                                        "quantity": form.cleaned_data.get('quantity'),
                                        "unit_type": form.cleaned_data.get('unit_type'),
                                        "unit_weight": form.cleaned_data.get('unit_weight'),
                                        "unit_rate": form.cleaned_data.get('unit_rate')
                                    })

                    # Delete existing related BoxDetails and BoxItems
                    BoxItem.objects.filter(box__awb=awb).delete()
                    BoxDetails.objects.filter(awb=awb).delete()

                    # Save BoxDetails one by one (so that the pre_save signal is triggered)
                    for box_number, details in box_and_box_item_detail.items():
                        box_detail = BoxDetails(
                            awb=awb,
                            length=details['length'],
                            breadth=details['breadth'],
                            height=details['height'],
                            actual_weight=details['actual_weight'],
                            box_awb_no=details['box_awb_no'],
                            bag_no=details['bag_no']
                        )
                        box_detail.save()  # Triggers pre_save signal for box_awb_no and weight calculations
                        details["instance"] = box_detail

                    # Prepare and create BoxItem instances (bulk_create can be used if no signals are needed)
                    box_item_instances = []
                    for box_data in box_and_box_item_detail.values():
                        box_instance = box_data["instance"]
                        for item in box_data.get("items", []):
                            amount = float(item.get("quantity")) * \
                                float(item.get("unit_rate"))
                            box_item = BoxItem(
                                box=box_instance,
                                description=item.get("description"),
                                hs_code=item.get("hs_code"),
                                quantity=item.get("quantity"),
                                unit_type=item.get("unit_type"),
                                unit_weight=item.get("unit_weight"),
                                unit_rate=item.get("unit_rate"),
                                amount=amount
                            )
                            box_item_instances.append(box_item)

                    BoxItem.objects.bulk_create(box_item_instances)
                    messages.success(request, "Shipment updated successfully!")
                    return redirect(request.META.get('HTTP_REFERER'))
            except Exception as e:
                messages.error(request, f"Error updating shipment: {e}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        awb_form = AWBUpdateForm(prefix="awb", instance=awb)
        if request.user.role == 'agencyuser':
            awb_form.fields['agency'].queryset = Agency.objects.filter(
                id=request.user.agency.id)
            awb_form.fields['agency'].initial = request.user.agency

        consignee_form = ConsigneeForm(
            prefix="consignee", instance=awb.consignee)
        consignor_form = ConsignorForm(
            prefix="consignor", instance=awb.consignor)
        box_details_formset = BoxDetailsFormSet(
            prefix="box_details", queryset=box_details)

        box_id_to_position = {}
        for i, box in enumerate(box_details):
            box_id_to_position[box.id] = i + 1

        box_choices = [(f"Box-{i+1}", f"Box-{i+1}")
                       for i in range(box_details.count())]

        box_item_formset = BoxItemFormSet(
            prefix="box_item", queryset=box_items)
        for form in box_item_formset.forms:
            form.fields['box_number'].widget.choices = box_choices
            if form.instance.id and form.instance.box_id in box_id_to_position:
                position = box_id_to_position[form.instance.box_id]
                box_number = f"Box-{position}"
                form.fields['box_number'].initial = box_number
                if form.is_bound:
                    form.data = form.data.copy()
                    form.data[f'{form.prefix}-box_number'] = box_number

    context = {
        "awb_form": awb_form,
        "consignee_form": consignee_form,
        "consignor_form": consignor_form,
        "units": UnitType.objects.all(),
        "box_details_formset": box_details_formset,
        "total_box_count": box_details_count,
        "create_shipment_invoice": create_shipment_invoice,
        "box_item_formset": box_item_formset,
        "box_items_count": box_items_count,
        'is_error': bool(awb_form.errors or consignee_form.errors or consignor_form.errors or box_details_formset.errors or box_item_formset.errors),
        'is_updating': True,
        "title": "Shipment",
        "awb": awb,
        "active_tab": "awb"
    }
    return render(request, "awb/awb/update.html", context)


@login_required
@has_roles(['admin', 'agencyuser'])
def export_awb_invoice_view(request, awb_no):
    format_type = request.GET.get("format", "excel")
    awb = AWBDetail.objects.get(awbno=awb_no)
    exporter = AWBInvoiceExporter(awb)
    return exporter.export_invoice(format_type)


@login_required
@has_roles(['admin', 'agencyuser'])
def export_awb_pdf(request, pk, mode):
    awb = get_object_or_404(AWBDetail, id=pk)
    exporter = AWBInvoiceExporter(awb)
    return exporter.generate_pdf(mode=mode)


@login_required
@has_roles(['admin', 'agencyuser'])
def shipment_history(request, awb_no):
    try:
        # Fetch the AWBDetail
        awb = get_object_or_404(AWBDetail, awbno=awb_no)
        # For regular GET requests, render the template
        if request.method == 'GET':
            context = {
                "awb": awb,
                "consignee": awb.consignee,
                "consignor": awb.consignor,
                "box_details": BoxDetails.objects.filter(awb=awb),
                "box_items": BoxItem.objects.filter(box__awb=awb),
                "title": "Shipment"
            }
            return render(request, "awb/awb/history.html", context)
        # For DataTables AJAX request
        if request.method == 'POST':
            draw = int(request.POST.get('draw', 1))
            start = int(request.POST.get('start', 0))
            length = int(request.POST.get('length', 10))
            search_value = request.POST.get('search[value]', '').lower()

            # Fetch all histories in a single query for each model
            histories = {
                'Consignee': awb.consignee.history.all(),
                'Consignor': awb.consignor.history.all(),
                'BoxDetails': BoxDetails.history.filter(awb=awb),
                'BoxItem': BoxItem.history.filter(box__awb=awb)
            }

            # Combine and sort all histories
            combined_history = sorted(
                chain(*histories.values()),
                key=lambda x: x.history_date,
                reverse=True
            )

            # Process history entries
            history_data = []
            for entry in combined_history:
                changes = {}
                if hasattr(entry, 'prev_record') and entry.prev_record:
                    delta = entry.diff_against(entry.prev_record)
                    for change in delta.changes:
                        field_value = getattr(entry, change.field)
                        if isinstance(field_value, models.Model):
                            old_value = getattr(
                                entry.prev_record, change.field, None)
                            new_value = field_value
                            changes[change.field] = {
                                "old": str(old_value) if old_value else "None",
                                "new": str(new_value) if new_value else "None"
                            }
                        else:
                            changes[change.field] = {
                                "old": change.old,
                                "new": change.new
                            }

                if changes or entry.history_type == "+":
                    # Get model name from the actual model
                    model_name = entry.history_object.__class__.__name__ if hasattr(
                        entry, 'history_object') else entry.__class__.__name__.replace('Historical', '')

                    history_data.append({
                        "date": entry.history_date.strftime("%Y-%m-%d %H:%M:%S"),
                        "model": model_name,
                        "changed_by": entry.history_user.full_name if entry.history_user else "System",
                        "history_type": "Created" if entry.history_type == "+" else "Updated" if entry.history_type == "~" else "Deleted",
                        "changes": changes
                    })

            # Apply search filter if provided
            if search_value:
                history_data = [
                    entry for entry in history_data
                    if (search_value in entry['model'].lower() or
                        search_value in entry['date'].lower() or
                        search_value in entry['changed_by'].lower() or
                        search_value in entry['history_type'].lower() or
                        any(search_value in str(change).lower() for change in entry['changes'].values()))
                ]

            # Calculate totals and paginate
            total_records = len(history_data)
            filtered_total = len(history_data)
            paginated_data = history_data[start:start + length]

            # Return JSON response for DataTables
            return JsonResponse({
                'draw': draw,
                'recordsTotal': total_records,
                'recordsFiltered': filtered_total,
                'data': paginated_data
            })

    except Exception as e:
        messages.error(request, f"An error occurred: {e}")
        try:
            return redirect(request.META['HTTP_REFERER'])
        except Exception:
            return redirect('awb:pages:awb:list')


@login_required
@has_roles(['admin', 'agencyuser'])
def cancel_awb(request, awb_no):
    try:
        awb = get_object_or_404(AWBDetail, awbno=awb_no)
        if request.user.role == 'agencyuser':
            if not request.user.is_default_user or awb.agency != request.user.agency:
                messages.error(
                    request, "You don't have permission to cancel AWB")
                return redirect(request.META.get('HTTP_REFERER', '/'))

        time_window = timedelta(hours=24)
        if timezone.now() - awb.created_at > time_window:
            messages.error(
                request, "AWB cannot be cancelled after 24 hours of creation")
            return redirect(request.META.get('HTTP_REFERER', '/'))

        if awb.is_cancelled:
            messages.error(request, "AWB is already cancelled!")
            return redirect(request.META.get('HTTP_REFERER', '/'))

        try:
            with transaction.atomic():
                awb.is_cancelled = True
                awb.cancelled_at = timezone.now()
                awb.save()
                messages.success(request, "AWB cancelled successfully!")
        except Exception as e:
            messages.error(request, f"An error occured: {str(e)}")

        return redirect(request.META.get('HTTP_REFERER', '/'))

    except Exception as e:
        messages.error(request, f"An error occured: {str(e)}")
        return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
@has_roles(['admin', 'agencyuser'])
def upload_forwarding_number(request):
    form = ForwardinNumberUploadForm()
    if request.method == 'POST':
        form = ForwardinNumberUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            via = form.cleaned_data['via']
            try:
                # Validate file format
                if not file.name.endswith(('.csv', '.xlsx', '.xls')):
                    messages.error(
                        request, "Invalid file format. Please upload CSV or Excel file.")
                    return redirect(request.META.get('HTTP_REFERER', '/'))

                # Read file based on extension
                df = pd.read_csv(file) if file.name.endswith(
                    '.csv') else pd.read_excel(file)
                df.fillna('', inplace=True)

                # Add error column
                df['ERROR'] = ''

                # Initialize counters
                has_errors = False
                updated_count = 0

                # Process each row
                for index, row in df.iterrows():
                    try:
                        # Get first and second column values regardless of header names
                        house_awb = str(df.iloc[index, 0]).strip()
                        forwarding_number = str(df.iloc[index, 1]).strip()

                        # Validate required fields
                        if not all([house_awb, forwarding_number]):
                            missing_fields = []
                            if not house_awb:
                                missing_fields.append('HOUSEAWB')
                            if not forwarding_number:
                                missing_fields.append('FORWARDINGNUMBER')
                            df.at[index,
                                  'ERROR'] = f"Missing required fields: {', '.join(missing_fields)}"
                            has_errors = True
                            continue

                        # Update records based on via field
                        if via == 'PARCEL':
                            # For BoxDetails, we need to find the AWB first
                            box = BoxDetails.objects.filter(
                                box_awb_no=house_awb).first()
                            if not box:
                                df.at[index,
                                      'ERROR'] = f'No Parcel found with number: {house_awb}'
                                has_errors = True
                                continue

                            box.forwarding_number = forwarding_number
                            box.save()
                            updated_count += 1

                        else:  # SHIPMENT
                            # For AWBDetail, update directly
                            awb = AWBDetail.objects.filter(
                                awbno=house_awb).first()
                            if not awb:
                                df.at[index,
                                      'ERROR'] = f'No shipment found with AWB: {house_awb}'
                                has_errors = True
                                continue

                            awb.forwarding_number = forwarding_number
                            awb.save()
                            updated_count += 1

                    except Exception as e:
                        df.at[index,
                              'ERROR'] = f'Error processing row: {str(e)}'
                        has_errors = True

                # Handle errors if any
                if has_errors:
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Sheet1')

                    response = HttpResponse(
                        output.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename=forwarding_number_errors.xlsx'

                    messages.warning(
                        request, "Some errors were found. Please check the downloaded file for details.")
                    return response

                # Success message
                messages.success(
                    request,
                    f"Forwarding numbers updated successfully! Updated {updated_count} {via.lower()}s."
                )
                return redirect(request.META.get('HTTP_REFERER', '/'))

            except Exception as e:
                messages.error(request, f"Error reading file: {str(e)}")
                return redirect(request.META.get('HTTP_REFERER', '/'))
        else:
            messages.error(request, f"Error uploading file: {form.errors}")

    return render(
        request,
        'awb/awb/upload_forwarding_number.html',
        {'form': form, "title": "Shipment"}
    )


@login_required
@has_roles(['admin', 'agencyuser'])
def get_box_awb_no(request, awb_no):
    awb = get_object_or_404(AWBDetail, awbno=awb_no)
    boxes = BoxDetails.objects.filter(awb=awb).values('box_awb_no')
    return JsonResponse(list(boxes), safe=False)


@login_required
@has_roles(['admin', 'agencyuser'])
def download_forwarding_number_template(request):
    box_details = BoxDetails.objects.all()
    awb_details = AWBDetail.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Forwarding Number'

    ws['A1'] = 'HOUSEAWB'
    ws['B1'] = 'FORWARDINGNUMBER'

    header_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = header_font

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=forwarding_number_template.xlsx'
    wb.save(response)
    return response


@login_required
@has_roles(['admin'])
def update_awb_numbers(request):
    if request.method == 'POST':
        # Initialize forms
        forwarding_form = UpdateForwardingNumberForm(request.POST)
        box_formset = forms.formset_factory(
            UpdateBoxAWBNumberForm)(request.POST, prefix='form')

        if forwarding_form.is_valid() and box_formset.is_valid():
            try:
                # Update forwarding number
                awbno = forwarding_form.cleaned_data['awbno']
                awb = AWBDetail.objects.get(awbno=awbno)
                awb.forwarding_number = forwarding_form.cleaned_data['forwarding_number']
                awb.save()

                # Update box AWB numbers
                for form in box_formset:
                    if form.cleaned_data:
                        box_id = form.cleaned_data['box_id']
                        box_awb_no = form.cleaned_data['box_awb_no']
                        box = BoxDetails.objects.get(id=box_id)
                        box.box_awb_no = box_awb_no
                        box.save()

                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': True})
                messages.success(request, "AWB numbers updated successfully.")
                return redirect('awb:pages:awb:detail', awb_no=awbno)

            except (AWBDetail.DoesNotExist, BoxDetails.DoesNotExist):
                error_msg = "AWB or Box not found."
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': error_msg})
                messages.error(request, error_msg)
            except Exception as e:
                error_msg = f"An error occurred: {str(e)}"
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': error_msg})
                messages.error(request, error_msg)
        else:
            error_msg = "Invalid form data."
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
@has_roles(['admin', 'agencyuser'])
def get_country_code(request):
    """
    AJAX endpoint to get country code for a given country ID
    """
    if request.method == 'GET':
        country_id = request.GET.get('country_id')
        if country_id:
            try:
                country = Country.objects.get(id=country_id)
                return JsonResponse({
                    'country_code': country.code,
                    'success': True
                })
            except Country.DoesNotExist:
                return JsonResponse({
                    'error': 'Country not found',
                    'success': False
                }, status=404)
        else:
            return JsonResponse({
                'error': 'Country ID is required',
                'success': False
            }, status=400)

    return JsonResponse({
        'error': 'Invalid request method',
        'success': False
    }, status=405)
