from openpyxl.worksheet.page import PageMargins
from awb.models import AWBDetail, BoxDetails
from django.http import FileResponse
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm, inch
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Table, TableStyle, Paragraph, Spacer, SimpleDocTemplate
from reportlab.lib.pagesizes import A4, landscape, letter, portrait
from io import BytesIO
import inflect
from openpyxl.utils import get_column_letter

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER


from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import Workbook

from openpyxl.styles import Border, Side, Alignment, Font

from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib import colors

import inflect


def convert_to_words_with_cents(amount):
    p = inflect.engine()

    whole_number = int(amount)
    cents = round((amount - whole_number) * 100)

    whole_number_in_words = p.number_to_words(
        whole_number).replace(",", "").upper()
    total_amount_in_words = ""
    if whole_number > 0:
        total_amount_in_words += whole_number_in_words
    if cents > 0:
        cents_in_words = p.number_to_words(cents).replace(",", "").upper()
        total_amount_in_words += f" AND {cents_in_words} CENTS"
    total_amount_in_words += " USD ONLY"
    total_amount_in_words = total_amount_in_words.upper()
    # remove any special characters ,*/-#$
    total_amount_in_words = total_amount_in_words.replace(
        ",", " ").replace("*", " ").replace("/", " ").replace("-", " ").replace("#", " ").replace("$", " ").replace(".", " ").replace("  ", " ")
    return total_amount_in_words


class AWBInvoiceExporter:
    """
    Exports an AWB invoice as an Excel workbook, formatted to match
    the 'INVOICE & PACKING LIST' layout from your provided screenshot.
    """

    def __init__(self, awb):
        self.awb = awb
        # Prefetch related box items to reduce database queries
        self.boxes = awb.boxdetails.prefetch_related("items").all()
        self.shipper = awb.company
        self.consignor = awb.consignor
        self.consignee = awb.consignee
        self.agency = awb.agency
        self.page_width, self.page_height = A4

    def export_invoice(self, format_type):
        print("format_type", format_type)
        if format_type.lower() == "pdf":
            return self._export_pdf()
        if format_type.lower() == "excel":
            return self._export_excel()

    def _export_pdf(self):

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch
        )
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name='TitleStyle',
            parent=styles['Heading1'],
            fontSize=11,
            alignment=TA_CENTER,
            spaceAfter=6
        )
        header_style = ParagraphStyle(
            name='HeaderStyle',
            parent=styles['Heading2'],
            fontSize=9,
            fontName='Helvetica-Bold'
        )
        normal_style = styles['Normal']

        # === 1) Main Header ===
        title_data = [
            [Paragraph("INVOICE & PACKING LIST", title_style), "", "", "", "", "", "", ""]]
        title_table = Table(title_data, colWidths=[
                            1*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        title_table.setStyle(TableStyle([
            ('SPAN', (0, 0), (7, 0)),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(title_table)

        # === 2) Top Info Section ===
        top_data = [
            [
                f"COUNTRY OF ORIGIN: NEPAL",
                f"ACTUAL WEIGHT: {self.awb.total_actual_weight}"
            ],
            [
                f"INVOICE DATE: {self.awb.booking_datetime.strftime('%B %d, %Y')}",
                f"TOTAL PIECES: {self.awb.total_box}"
            ],
            [
                f"INVOICE NO: {self.awb.awbno}",
                ""
            ],
            [
                "",
                ""
            ]
        ]
        top_table = Table(top_data, colWidths=[4 * inch, 4 * inch])
        top_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LINEAFTER', (0, 0), (0, -1), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(top_table)

        # === 3) Shipper/Consignee Section ===
        shipper_consignee_headers = [["SHIPPER", "CONSIGNEE"]]
        sc_header_table = Table(shipper_consignee_headers,
                                colWidths=[4 * inch, 4 * inch])
        sc_header_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LINEAFTER', (0, 0), (0, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(sc_header_table)
        if self.awb.is_cash_user:
            shipper_info = [
                self.shipper.name.upper(),
                self.shipper.name.upper(),
                self.shipper.address1.upper(),
                self.shipper.address2 if self.shipper.address2 else "",
                str(self.shipper.post_zip_code).upper(),
                f"{self.shipper.city or ''}, {self.shipper.state_county or ''}".upper() ,
                self.shipper.country.name.upper(),
                f"EMAIL: {self.shipper.email}".upper(
                ) if self.shipper.email else "",
                f"PHONE NUMBER: +{self.shipper.phone_number}".upper(
                ) if self.shipper.phone_number else ""
            ]
        else:
            shipper_info = [
                self.agency.company_name,
                self.agency.owner_name,
                self.agency.address1,
                self.agency.address2 if self.agency.address2 else "",
                str(self.agency.zip_code),
                f"{self.agency.city or ''},{self.agency.state or ''}",
                self.agency.country.name,
                f"EMAIL: {self.agency.email}".upper(
                ) if self.agency.email else "",
                f"PHONE NUMBER: +{self.agency.contact_no_1}".upper(
                ) if self.agency.contact_no_1 else ""
            ]
        consignee_info = [
            self.consignee.company.upper(),
            self.consignee.person_name.upper(),
            self.consignee.address1.upper(),
            self.consignee.address2.upper() if self.consignee.address2 else "",
            str(self.consignee.post_zip_code).upper(),
            (f"{self.consignee.city or ''}, {self.consignee.state_county or ''}").upper(),
            self.awb.destination.name.upper(),
            f"EMAIL:{self.consignee.email_address}".upper(
            ) if self.consignee.email_address else "",
            f"PHONE NUMBER: +{self.consignee.phone_number}".upper(
            ) if self.consignee.phone_number else ""
        ]
        sc_data = [[shipper_info[i], consignee_info[i]]
                   for i in range(len(shipper_info))]
        sc_table = Table(sc_data, colWidths=[4 * inch, 4 * inch])
        sc_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LINEAFTER', (0, 0), (0, -1), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(sc_table)

        # === 4) Items Table ===
        headers = ["BOXES", "SR NO", "DESCRIPTION", "HS CODE",
                   "UNIT TYPE", "QUANTITY", "UNIT RATE", "AMOUNT (USD)"]
        col_widths = [0.8 * inch, 0.4 * inch, 2.0 * inch, 1.1 * inch,
                      0.9 * inch, 0.8 * inch, 0.9 * inch, 1.1 * inch]

        items_data = [headers]

        for box in self.boxes:
            sr_no = 1
            box_items = list(box.items.all())
            box_number = f"BOX{box.get_box_number()}"
            for idx, item in enumerate(box_items):
                box_label = box_number if idx == 0 else ""
                items_data.append([
                    box_label,
                    sr_no,
                    item.description,
                    item.hs_code,
                    item.unit_type.name,
                    item.quantity,
                    item.unit_rate,
                    item.amount
                ])
                sr_no += 1

        items_table = Table(items_data, colWidths=col_widths)

        table_style = [
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (5, 1), (-1, -1), 'CENTER'),
        ]

        current_row = 1
        for box in self.boxes:
            box_items = list(box.items.all())
            if len(box_items) > 1:
                table_style.append(
                    ('SPAN', (0, current_row), (0, current_row + len(box_items) - 1)))
            current_row += len(box_items)

        items_table.setStyle(TableStyle(table_style))
        elements.append(items_table)

        # === 5) Footer Section ===
        total_quantity = sum(
            item.quantity for box in self.boxes for item in box.items.all())
        grand_total = sum(
            item.amount for box in self.boxes for item in box.items.all())

        grandtotal_in_words = convert_to_words_with_cents(grand_total).upper()

        total_data = [["Total Quantity", "", "", "", "",
                       total_quantity, "Grand Total", grand_total]]
        total_table = Table(total_data, colWidths=col_widths)
        total_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('SPAN', (0, 0), (4, 0)),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(total_table)

        words_data = [[grandtotal_in_words, "",
                       "", "", "", "", "Total", grand_total]]
        words_table = Table(words_data, colWidths=col_widths)
        words_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('SPAN', (0, 0), (5, 0)),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(words_table)

        notes_data = [["NOTES", "", "", "", "", "SIGNATURE / STAMP", "", ""]]
        notes_table = Table(notes_data, colWidths=col_widths)
        notes_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('SPAN', (0, 0), (4, 0)),
            ('SPAN', (5, 0), (7, 0)),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(notes_table)

        declaration_text = Paragraph(
            "We declare that the above mentioned goods are made in Nepal<br/>"
            "and other descriptions are true.",
            normal_style  # or any style you want (normal_style is fine)
        )

        declaration_data = [[declaration_text, "", "", "", "", "", "", ""]]

        declaration_table = Table(declaration_data, colWidths=col_widths)
        declaration_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('SPAN', (0, 0), (4, 0)),
            ('SPAN', (5, 0), (7, 0)),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(declaration_table)

        # Build and return response
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=AWB_{self.awb.awbno}_Invoice.pdf'
        return response

    def _export_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"

        # === Define Styles ===
        bold = Font(bold=True)
        bold_large = Font(bold=True, size=14)
        # Changed from "left" to "center"
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        large_pt_side = Side(style="thin", color="000000")

        thick_border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=large_pt_side,
            bottom=large_pt_side
        )

        # Row tracker
        row = 1

        # === 1) Main Header ===
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row=row, column=1, value="INVOICE & PACKING LIST")
        cell.font = bold_large
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        # Add borders to merged cells
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thick_border
        row += 1

        # === 2) Top Info Section ===
        # row1
        # Left column info
        ws.cell(
            row=row,
            column=1,
            value="COUNTRY OF ORIGIN: NEPAL"
        ).border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=large_pt_side,
            bottom=Side(style=None)
        )
        ws.merge_cells(f'A{row}:C{row}')
        # Right column info

        ws.cell(row=row, column=4,
                value=f"ACTUAL WEIGHT :{self.awb.total_actual_weight}").border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=large_pt_side,
            bottom=Side(style=None)
        )
        ws.merge_cells(f'D{row}:H{row}')
        row += 1
        # row 3
        # Invoice Date
        ws.cell(row=row, column=1,
                value=f"INVOICE DATE. : {self.awb.booking_datetime.strftime('%B %d, %Y')}").border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=Side(style=None),
            bottom=Side(style=None)
        )
        ws.merge_cells(f'A{row}:C{row}')
        # Total Pieces
        ws.cell(row=row, column=4,
                value=f"TOTAL PIECES :{self.awb.total_box}").border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=Side(style=None),
            bottom=Side(style=None)
        )
        ws.merge_cells(f'D{row}:H{row}')
        row += 1
        # row 3
        # Invoice Number
        ws.cell(row=row, column=1,
                value=f"INVOICE NO: {self.awb.awbno}").border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=Side(style=None),
            bottom=Side(style=None)
        )
        ws.merge_cells(f'A{row}:C{row}')

        ws.cell(row=row, column=4, value="").border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=Side(style=None),
            bottom=Side(style=None)
        )

        ws.merge_cells(f'D{row}:H{row}')
        row += 1
        # row 4

        # empty row
        ws.cell(row=row, column=1, value="").border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=Side(style=None),
            bottom=large_pt_side
        )
        ws.merge_cells(f'A{row}:C{row}')

        ws.cell(row=row, column=4, value="").border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=Side(style=None),
            bottom=large_pt_side
        )
        ws.merge_cells(f'D{row}:H{row}')

        row += 1
        # new row
        # === 3) Shipper/Consignee Section ===
        # Headers
        ws.cell(row=row, column=1, value="SHIPPER").font = bold
        ws.cell(row=row, column=1).border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=large_pt_side,
            bottom=Side(style=None)
        )
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row, column=4, value="CONSIGNEE").font = bold
        ws.cell(row=row, column=4).border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=large_pt_side,
            bottom=Side(style=None)
        )
        ws.merge_cells(f'D{row}:H{row}')
        row += 1

        # Shipper and Consignee Details
        if self.awb.is_cash_user:
            shipper_info = [
                self.shipper.name,
                self.shipper.name,
                self.shipper.address1,
                self.shipper.address2 if self.shipper.address2 else "",
                str(self.shipper.post_zip_code),
                f"{self.shipper.city or ''}, {self.shipper.state_county or ''}",
                self.shipper.country.name,
                f"EMAIL:",
                f"PHONE NUMBER: +{self.shipper.phone_number}"
            ]
        else:
            shipper_info = [
                self.agency.company_name,
                self.agency.owner_name,
                self.agency.address1,
                self.agency.address2 if self.agency.address2 else "",
                str(self.agency.zip_code),
                f"{self.agency.city or ''},{self.agency.state or ''}",
                self.agency.country.name,
                f"EMAIL:",
                f"PHONE NUMBER: +{self.agency.contact_no_1}"
            ]

        consignee_info = [
            self.consignee.company,
            self.consignee.person_name,
            self.consignee.address1,
            self.consignee.address2 if self.consignee.address2 else "",
            str(self.consignee.post_zip_code),
            (f"{self.consignee.city or ''}, {self.consignee.state_county or ''}").upper(),
            self.awb.destination.name,
            "EMAIL:",
            f"PHONE NUMBER: +{self.consignee.phone_number}"
        ]

        for s_info, c_info in zip(shipper_info, consignee_info):
            cell_s = ws.cell(row=row, column=1, value=s_info)
            cell_s.border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=Side(style=None),
                bottom=Side(style=None)
            )
            cell_s.alignment = left
            ws.merge_cells(f'A{row}:C{row}')

            cell_c = ws.cell(row=row, column=4, value=c_info)
            cell_c.border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=Side(style=None),
                bottom=Side(style=None)
            )
            cell_c.alignment = left
            ws.merge_cells(f'D{row}:H{row}')
            row += 1

        # empty row
        ws.cell(row=row, column=1, value="").border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=Side(style=None),
            bottom=large_pt_side
        )
        ws.merge_cells(f'A{row}:C{row}')

        ws.cell(row=row, column=4, value="").border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=Side(style=None),
            bottom=large_pt_side
        )
        ws.merge_cells(f'D{row}:H{row}')
        row += 1

        # === 4) Items Table ===
        # Table Headers
        headers = ["BOXES", "SR NO", "DESCRIPTION", "HS CODE",
                   "UNIT TYPE", "QUANTITY", "UNIT RATE", "AMOUNT (USD)"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = bold
            cell.alignment = center
            cell.border = thick_border
        row += 1

        # Table Content

        for box in self.boxes:
            box_items = list(box.items.all())
            start_row = row

            # Write box number once
            box_cell = ws.cell(row=start_row, column=1,
                               value=f"BOX{box.get_box_number()}")
            box_cell.alignment = center
            box_cell.border = thick_border
            sr_no = 1
            for item in box_items:
                # Item details
                data = [
                    None,  # Box number handled separately
                    sr_no,
                    item.description,
                    item.hs_code,
                    item.unit_type.name,
                    item.quantity,
                    item.unit_rate,
                    item.amount
                ]

                for col_idx, value in enumerate(data, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.alignment = center
                    cell.border = thick_border

                sr_no += 1
                row += 1

            # Merge box number cells if there are multiple items
            if len(box_items) > 1:
                ws.merge_cells(start_row=start_row, start_column=1,
                               end_row=row-1, end_column=1)

        # === 5) Auto-adjust Column Widths and Row Heights ===

        # Auto-fit all column widths
        for column in range(1, 9):  # Columns A through H
            max_length = 0
            column_letter = get_column_letter(column)

            # Check all cells in the column
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=column)
                try:
                    if cell.value:
                        # Handle different types of values and calculate proper width
                        cell_value = str(cell.value).strip()

                        # For wrapped text, consider line breaks
                        if '\n' in cell_value:
                            lines = cell_value.split('\n')
                            cell_length = max(len(line) for line in lines)
                        else:
                            cell_length = len(cell_value)

                        max_length = max(max_length, cell_length)
                except:
                    continue

            # Set width with padding but no fixed limits
            adjusted_width = max_length + 3 if max_length > 0 else 10
            ws.column_dimensions[column_letter].width = adjusted_width

        # Auto-fit row heights
        for row_num in range(1, ws.max_row + 1):
            max_height = 0  # Start with 0 to let content determine height

            for col_num in range(1, 9):  # Check all columns in each row
                cell = ws.cell(row=row_num, column=col_num)

                if cell.value:
                    try:
                        cell_value = str(cell.value).strip()

                        # Calculate height based on text length and column width
                        column_letter = get_column_letter(col_num)
                        column_width = ws.column_dimensions[column_letter].width or 10

                        # Estimate lines needed based on text length and column width
                        if len(cell_value) > column_width:
                            estimated_lines = max(
                                1, len(cell_value) // int(column_width))
                            # Add extra lines for explicit line breaks
                            if '\n' in cell_value:
                                estimated_lines += cell_value.count('\n')

                            # Calculate height (approximately 15 points per line)
                            calculated_height = estimated_lines * 15
                            max_height = max(max_height, calculated_height)

                        # Handle explicit line breaks
                        elif '\n' in cell_value:
                            line_count = cell_value.count('\n') + 1
                            calculated_height = line_count * 15
                            max_height = max(max_height, calculated_height)
                        else:
                            # Single line content
                            max_height = max(max_height, 15)

                    except:
                        continue

            # Set the row height based on content (minimum 15 for readability)
            if max_height > 0:
                ws.row_dimensions[row_num].height = max(15, max_height)
            else:
                ws.row_dimensions[row_num].height = 15

        # === 6) Add Footer Section ===
        # Calculate total quantity
        total_quantity = sum(
            item.quantity for box in self.boxes for item in box.items.all())

        # Calculate grand total
        grand_total = sum(
            item.amount for box in self.boxes for item in box.items.all())

        # Add Total Quantity row
        ws.cell(row=row, column=1, value="Total Quantity").border = thick_border
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row=row, column=6, value=total_quantity).border = thick_border
        ws.cell(row=row, column=7, value="Grand Total").border = thick_border
        ws.cell(row=row, column=8, value=grand_total).border = thick_border
        row += 1

        grandtotal_in_words = convert_to_words_with_cents(grand_total)

        # Add Total Amount row
        ws.cell(row=row, column=1, value=grandtotal_in_words).border = thick_border
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=7,
                value=f"Total:{grand_total}").border = thick_border
        ws.merge_cells(f'G{row}:H{row}')
        row += 1

        # two column
        ws.cell(row=row, column=1, value="NOTES").border = thick_border
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row, column=4, value="SIGNATURE / STAMP").border = thick_border
        ws.merge_cells(f'D{row}:H{row}')
        row += 1

        # Add Declaration Note
        # Declaration row
        declaration_cell = ws.cell(
            row=row,
            column=1,
            value="WE DECLARE THAT THE ABOVE MENTIONED GOODS ARE MADE IN NEPAL AND OTHER DESCRIPTIONS ARE TRUE."
        )
        declaration_cell.border = thick_border
        declaration_cell.alignment = Alignment(
            wrap_text=True, vertical="center")
        ws.merge_cells(f'A{row}:C{row}')

        signature_cell = ws.cell(row=row, column=4, value="")
        signature_cell.border = thick_border
        signature_cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f'D{row}:H{row}')

        row += 1

        # Final adjustment for the last few rows that were added after the main auto-fit
        for row_num in range(row - 4, row):
            if row_num > 0:
                max_height = 0
                for col_num in range(1, 9):
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.value:
                        try:
                            cell_value = str(cell.value).strip()
                            column_letter = get_column_letter(col_num)
                            column_width = ws.column_dimensions[column_letter].width or 10

                            if len(cell_value) > column_width:
                                estimated_lines = max(
                                    1, len(cell_value) // int(column_width))
                                calculated_height = estimated_lines * 15
                                max_height = max(max_height, calculated_height)
                            elif '\n' in cell_value:
                                line_count = cell_value.count('\n') + 1
                                calculated_height = line_count * 15
                                max_height = max(max_height, calculated_height)
                            else:
                                max_height = max(max_height, 15)
                        except:
                            continue

                if max_height > 0:
                    ws.row_dimensions[row_num].height = max(15, max_height)
                else:
                    ws.row_dimensions[row_num].height = 15

        # Manual column width adjustments
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 48

        # === Page Setup: IMPROVED horizontal centering ===
        # Set page orientation to portrait (default)
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

        # Enable fit to page
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.fitToPage = True

        # IMPORTANT: Center horizontally on the printed page
        ws.print_options.horizontalCentered = True
        # Only center horizontally, not vertically
        ws.print_options.verticalCentered = False

        # Alternative method - also set page_setup horizontal centering
        ws.page_setup.horizontalCentered = True

        # Set custom margins (in inches)
        ws.page_margins = PageMargins(
            left=0.2,
            right=0.2,
            top=0.2,
            bottom=0.2,
            header=0.2,
            footer=0.2
        )

        # Export
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename={self.awb.awbno}_invoice.xlsx'
        return response

    def generate_pdf(self, mode="box"):
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=landscape(A4))

        total_boxes = self.boxes.count()

        if mode == "address":
            # This will generate for address mode
            self.draw_awb_address_label(p)

        p.save()
        buffer.seek(0)

        return FileResponse(
            buffer,
            as_attachment=True,
            filename=f'AWB_{self.awb.awbno}_{mode}.pdf',
            content_type='application/pdf'
        )

    def draw_awb_address_label(self, p):
        # ★  Make absolutely sure we are in portrait  ★
        p.setPageSize(portrait(A4))
        page_w, page_h = portrait(A4)

        margin = 28
        label_w = page_w - 2*margin
        # 4 labels, 5 gutters (top + 3 between + bottom)
        label_h = (page_h - 5*margin) / 4

        consignor = self.consignor
        consignee = self.consignee
        boxes = list(self.awb.boxdetails.all())
        total_boxes = len(boxes)

        box_idx = 0
        while box_idx < total_boxes:
            # y-positions of the four slots on this page
            # slot 0 (top) down to slot 3 (bottom)
            y_slots = [
                page_h - margin - label_h,
                page_h - margin - (label_h + margin)*2 + margin,
                page_h - margin - (label_h + margin)*3 + margin*2,
                page_h - margin - (label_h + margin)*4 + margin*3,
            ]

            # ------- first box on this sheet -------
            self.draw_label_section(p, margin, y_slots[0], label_w, label_h,
                                    "Shipper",   consignor,  box_idx+1, total_boxes)
            self.draw_label_section(p, margin, y_slots[1], label_w, label_h,
                                    "Consignee", consignee,  box_idx+1, total_boxes)
            box_idx += 1

            # ------- second box (if any) on same sheet -------
            if box_idx < total_boxes:
                self.draw_label_section(p, margin, y_slots[2], label_w, label_h,
                                        "Shipper",   consignor,  box_idx+1, total_boxes)
                self.draw_label_section(p, margin, y_slots[3], label_w, label_h,
                                        "Consignee", consignee,  box_idx+1, total_boxes)
                box_idx += 1

            p.showPage()                   # new physical sheet

    def draw_label_section(self, p, x, y, w, h, role, person, idx, total_boxes):
        p.setStrokeColor(colors.black)
        p.setLineWidth(1)
        p.rect(x, y, w, h, stroke=1, fill=0)

        pad_x, pad_y = 10, 12
        header_sz, body_sz = 14, 11
        awb_sz, line_gap = 12, 14
        rhs = x + w - pad_x        # right-hand column anchor

        # Header & counter
        p.setFont("Helvetica-Bold", header_sz)
        p.drawString(x + pad_x, y + h - pad_y, role.upper())
        p.drawRightString(rhs,  y + h - pad_y, f"{idx} / {total_boxes}")

        # Address block
        cur_y = y + h - pad_y - 22

        def line(lbl, val, bold=False):
            nonlocal cur_y
            if val:
                p.setFont("Helvetica", body_sz)
                p.drawString(x + pad_x, cur_y, lbl)
                p.setFont("Helvetica-Bold" if bold else "Helvetica", body_sz)
                p.drawString(x + pad_x + 58, cur_y, str(val))
                cur_y -= line_gap

        line("Name:",     person.person_name, bold=True)
        line("Address:",  person.address1)
        if person.address2:
            line("",      person.address2)
        line("Pincode:",  person.post_zip_code)
        line("City:",     person.city)
        line("State:",    person.state_county)
        country = (self.awb.origin.name if role == "Shipper"
                   else self.awb.destination.name)
        line("Country:",  country.upper(), bold=True)
        line("Phone No:", person.phone_number)

        # AWB & Barcode (right-aligned)
        awb_y = y + pad_y + 34
        barcode_y = y + pad_y

        p.setFont("Helvetica-Bold", awb_sz)
        # p.drawRightString(rhs, awb_y, str(self.awb.awbno))

        if self.awb.barcode_image:
            try:
                barcode = ImageReader(self.awb.barcode_image.path)
                p.drawImage(barcode,
                            rhs - 200, barcode_y,
                            width=240, height=56,   # increased size
                            preserveAspectRatio=True, mask='auto')

            except Exception as e:
                print("Barcode render error:", e)
