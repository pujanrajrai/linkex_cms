from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Q
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta, date
import json
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

from decorators import has_roles
from .forms import MainFilterForm, ShipmentFilterForm
from awb.models import AWBDetail, BoxDetails
from finance.models import Invoice, Ledger
from hub.models import Run, RunAWB
from accounts.models import Agency, Company


def get_date_range(date_range, from_date, to_date):
    """Helper function to get date range based on selection"""
    today = timezone.now().date()

    if date_range == 'today':
        return today, today
    elif date_range == 'yesterday':
        yesterday = today - timedelta(days=1)
        return yesterday, yesterday
    elif date_range == 'this_week':
        # Get start of week (Monday)
        start_week = today - timedelta(days=today.weekday())
        return start_week, today
    elif date_range == 'last_week':
        # Get last week Monday to Sunday
        start_last_week = today - timedelta(days=today.weekday() + 7)
        end_last_week = start_last_week + timedelta(days=6)
        return start_last_week, end_last_week
    elif date_range == 'this_month':
        return today.replace(day=1), today
    elif date_range == 'last_month':
        last_month = today.replace(day=1) - timedelta(days=1)
        return last_month.replace(day=1), last_month
    elif date_range == 'this_year':
        return today.replace(month=1, day=1), today
    elif date_range == 'last_year':
        last_year = today.year - 1
        return date(last_year, 1, 1), date(last_year, 12, 31)
    elif date_range == 'custom' and from_date and to_date:
        return from_date, to_date
    else:
        return today, today


@login_required
@has_roles(['admin'])
def dashboard(request):
    """Simplified dashboard view with three sections"""

    # Initialize forms
    main_form = MainFilterForm(request.GET)
    shipment_form = ShipmentFilterForm(request.GET)

    # Get date range
    date_range = request.GET.get('date_range', 'today')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
    if to_date:
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

    date_from, date_to = get_date_range(date_range, from_date, to_date)

    # Get main filters
    company = request.GET.get('company')

    # Get shipment filters (agency and country now in shipment section)
    agency = request.GET.get('agency')
    country = request.GET.get('country')

    # Get shipment statistics
    shipment_stats = get_shipment_statistics(
        date_from, date_to, company, agency, country)

    # Get run statistics
    run_stats = get_run_statistics(date_from, date_to, company, agency)

    # Get finance statistics
    finance_stats = get_finance_statistics(date_from, date_to, company, agency)

    context = {
        'title': 'Management Dashboard',
        'main_form': main_form,
        'shipment_form': shipment_form,
        'date_from': date_from,
        'date_to': date_to,
        'shipment_stats': shipment_stats,
        'run_stats': run_stats,
        'finance_stats': finance_stats,
    }

    return render(request, 'awb/reports/dashboard.html', context)


def get_shipment_statistics(date_from, date_to, company, agency, country):
    """Get shipment statistics for the specified date range"""

    # Base queryset
    queryset = AWBDetail.objects.filter(
        booking_datetime__date__gte=date_from,
        booking_datetime__date__lte=date_to,
        is_deleted=False
    )

    # Apply filters
    if company:
        queryset = queryset.filter(company_id=company)
    if agency:
        queryset = queryset.filter(agency_id=agency)
    if country:
        queryset = queryset.filter(
            Q(origin_id=country) | Q(destination_id=country))

    # Get aggregated data
    stats = queryset.aggregate(
        total_shipments=Count('id'),
        total_boxes=Sum('total_box'),
        total_actual_weight=Sum('total_actual_weight'),
        total_charged_weight=Sum('total_charged_weight'),
        total_volumetric_weight=Sum('total_volumetric_weight')
    )

    # Handle None values
    for key, value in stats.items():
        if value is None:
            stats[key] = 0
    return stats


def get_run_statistics(date_from, date_to, company, agency):
    """Get run statistics for the specified date range"""

    # Base queryset for runs created in date range
    created_queryset = Run.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
        is_deleted=False
    )

    # Runs departing in date range (regardless of creation date)
    departing_queryset = Run.objects.filter(
        flight_departure_date__gte=date_from,
        flight_departure_date__lte=date_to,
        is_deleted=False
    )

    # Apply company filter
    if company:
        created_queryset = created_queryset.filter(company_id=company)
        departing_queryset = departing_queryset.filter(
            company_id=company)

    # Get run counts
    total_runs_created = created_queryset.count()
    total_runs_departing = departing_queryset.count()

    # Get AWBs in created runs for weight calculations
    created_run_awbs = RunAWB.objects.filter(run__in=created_queryset)
    if agency:
        created_run_awbs = created_run_awbs.filter(awb__agency_id=agency)

    # Get AWBs in departing runs
    departing_run_awbs = RunAWB.objects.filter(
        run__in=departing_queryset)
    if agency:
        departing_run_awbs = departing_run_awbs.filter(awb__agency_id=agency)

    # Get piece count for departing runs
    departing_pieces_count = departing_run_awbs.count()

    # Get weight statistics for created runs
    created_awb_ids = created_run_awbs.values_list('awb_id', flat=True)
    created_weight_stats = AWBDetail.objects.filter(id__in=created_awb_ids).aggregate(
        total_pieces=Count('id'),
        total_actual_weight=Sum('total_actual_weight'),
        total_charged_weight=Sum('total_charged_weight'),
        total_volumetric_weight=Sum('total_volumetric_weight')
    )

    # Get weight statistics for departing runs
    departing_awb_ids = departing_run_awbs.values_list('awb_id', flat=True)
    departing_weight_stats = AWBDetail.objects.filter(id__in=departing_awb_ids).aggregate(
        total_actual_weight=Sum('total_actual_weight'),
        total_charged_weight=Sum('total_charged_weight'),
        total_volumetric_weight=Sum('total_volumetric_weight')
    )

    # Handle None values for created runs
    for key, value in created_weight_stats.items():
        if value is None:
            created_weight_stats[key] = 0

    # Handle None values for departing runs
    for key, value in departing_weight_stats.items():
        if value is None:
            departing_weight_stats[key] = 0

    stats = {
        'total_runs_created': total_runs_created,
        'total_runs_departing': total_runs_departing,
        'created_pieces': created_weight_stats['total_pieces'],
        'created_actual_weight': created_weight_stats['total_actual_weight'],
        'created_charged_weight': created_weight_stats['total_charged_weight'],
        'created_volumetric_weight': created_weight_stats['total_volumetric_weight'],
        'departing_pieces': departing_pieces_count,
        'departing_actual_weight': departing_weight_stats['total_actual_weight'],
        'departing_charged_weight': departing_weight_stats['total_charged_weight'],
        'departing_volumetric_weight': departing_weight_stats['total_volumetric_weight'],
    }

    return stats


def get_finance_statistics(date_from, date_to, company, agency):
    """Get finance statistics for the specified date range"""

    # Cash user statistics (from invoices where agency is null)
    cash_invoice_queryset = Invoice.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
        awb__agency__isnull=True,
        is_active=True
    )

    if company:
        cash_invoice_queryset = cash_invoice_queryset.filter(
            awb__company_id=company)

    cash_stats = cash_invoice_queryset.aggregate(
        cash_invoices_created=Count('id'),
        cash_debit_amount=Sum('grand_total'),
        cash_credit_amount=Sum('total_paid_amount')
    )

    # Cash shipments without invoices
    cash_no_invoice_queryset = AWBDetail.objects.filter(
        booking_datetime__date__gte=date_from,
        booking_datetime__date__lte=date_to,
        agency__isnull=True,
        is_invoice_generated=False,
        is_deleted=False
    )

    if company:
        cash_no_invoice_queryset = cash_no_invoice_queryset.filter(
            company_id=company)

    cash_shipments_without_invoice = cash_no_invoice_queryset.count()

    # Agency statistics (from ledger)
    agency_ledger_queryset = Ledger.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
        agency__isnull=False
    )

    if company:
        agency_ledger_queryset = agency_ledger_queryset.filter(
            company_id=company)
    if agency:
        agency_ledger_queryset = agency_ledger_queryset.filter(
            agency_id=agency)

    # Get debit and credit amounts
    agency_debit = agency_ledger_queryset.filter(
        ledger_type='DEBIT'
    ).aggregate(total_debit=Sum('amount'))['total_debit'] or 0

    agency_credit = agency_ledger_queryset.filter(
        ledger_type='CREDIT'
    ).aggregate(total_credit=Sum('amount'))['total_credit'] or 0

    # Agency shipments without invoices
    agency_no_invoice_queryset = AWBDetail.objects.filter(
        booking_datetime__date__gte=date_from,
        booking_datetime__date__lte=date_to,
        agency__isnull=False,
        is_invoice_generated=False,
        is_deleted=False
    )

    if company:
        agency_no_invoice_queryset = agency_no_invoice_queryset.filter(
            company_id=company)
    if agency:
        agency_no_invoice_queryset = agency_no_invoice_queryset.filter(
            agency_id=agency)

    agency_shipments_without_invoice = agency_no_invoice_queryset.count()

    # Handle None values for cash stats
    if cash_stats['cash_invoices_created'] is None:
        cash_stats['cash_invoices_created'] = 0
    if cash_stats['cash_debit_amount'] is None:
        cash_stats['cash_debit_amount'] = 0
    if cash_stats['cash_credit_amount'] is None:
        cash_stats['cash_credit_amount'] = 0
    stats = {
        # Cash user statistics
        'cash_invoices_created': cash_stats['cash_invoices_created'],
        'cash_debit_amount': cash_stats['cash_debit_amount'],
        'cash_credit_amount': cash_stats['cash_credit_amount'],
        'cash_shipments_without_invoice': cash_shipments_without_invoice,

        # Agency statistics
        'agency_total_debit': agency_debit,
        'agency_total_credit': agency_credit,
        'agency_shipments_without_invoice': agency_shipments_without_invoice,
    }

    return stats


@login_required
@has_roles(['admin'])
def export_dashboard_data(request):
    """Export dashboard data to Excel"""
    # Get filters
    date_range = request.GET.get('date_range', 'today')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
    if to_date:
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

    date_from, date_to = get_date_range(date_range, from_date, to_date)

    # Get main filters
    company = request.GET.get('company')
    agency = request.GET.get('agency')

    # Get statistics
    shipment_stats = get_shipment_statistics(
        date_from, date_to, company, agency, request.GET.get('country'))
    run_stats = get_run_statistics(date_from, date_to, company, agency)
    finance_stats = get_finance_statistics(date_from, date_to, company, agency)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard Summary"

    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC",
                              end_color="CCCCCC", fill_type="solid")

    # Add title
    ws['A1'] = f"Dashboard Report ({date_from} to {date_to})"
    ws['A1'].font = Font(bold=True, size=14)

    row = 3

    # Shipment Section
    ws[f'A{row}'] = "SHIPMENT STATISTICS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    row += 1

    ws[f'A{row}'] = "Total Shipments"
    ws[f'B{row}'] = shipment_stats['total_shipments']
    row += 1

    ws[f'A{row}'] = "Total Boxes"
    ws[f'B{row}'] = shipment_stats['total_boxes']
    row += 1

    ws[f'A{row}'] = "Total Actual Weight (kg)"
    ws[f'B{row}'] = shipment_stats['total_actual_weight']
    row += 1

    ws[f'A{row}'] = "Total Charged Weight (kg)"
    ws[f'B{row}'] = shipment_stats['total_charged_weight']
    row += 1

    ws[f'A{row}'] = "Total Volumetric Weight (kg)"
    ws[f'B{row}'] = shipment_stats['total_volumetric_weight']
    row += 2

    # Run Section
    ws[f'A{row}'] = "RUN STATISTICS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    row += 1

    ws[f'A{row}'] = "Total Runs (Created)"
    ws[f'B{row}'] = run_stats['total_runs_created']
    row += 1

    ws[f'A{row}'] = "Total Runs (Departing In Date Range)"
    ws[f'B{row}'] = run_stats['total_runs_departing']
    row += 1

    ws[f'A{row}'] = "Total Pieces (Created Runs)"
    ws[f'B{row}'] = run_stats['created_pieces']
    row += 1

    ws[f'A{row}'] = "Total Actual Weight (kg) - Created Runs"
    ws[f'B{row}'] = run_stats['created_actual_weight']
    row += 1

    ws[f'A{row}'] = "Total Charged Weight (kg) - Created Runs"
    ws[f'B{row}'] = run_stats['created_charged_weight']
    row += 1

    ws[f'A{row}'] = "Total Volumetric Weight (kg) - Created Runs"
    ws[f'B{row}'] = run_stats['created_volumetric_weight']
    row += 1

    ws[f'A{row}'] = "Total Pieces (Departing Runs)"
    ws[f'B{row}'] = run_stats['departing_pieces']
    row += 1

    ws[f'A{row}'] = "Total Actual Weight (kg) - Departing Runs"
    ws[f'B{row}'] = run_stats['departing_actual_weight']
    row += 1

    ws[f'A{row}'] = "Total Charged Weight (kg) - Departing Runs"
    ws[f'B{row}'] = run_stats['departing_charged_weight']
    row += 1

    ws[f'A{row}'] = "Total Volumetric Weight (kg) - Departing Runs"
    ws[f'B{row}'] = run_stats['departing_volumetric_weight']
    row += 2

    # Finance Section
    ws[f'A{row}'] = "FINANCE STATISTICS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    row += 1

    # Cash Users
    ws[f'A{row}'] = "Cash Users (From Invoices)"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    ws[f'A{row}'] = "Cash Invoices Created"
    ws[f'B{row}'] = finance_stats['cash_invoices_created']
    row += 1

    ws[f'A{row}'] = "Cash Debit Amount (NPR)"
    ws[f'B{row}'] = finance_stats['cash_debit_amount']
    row += 1

    ws[f'A{row}'] = "Cash Shipments Without Invoice"
    ws[f'B{row}'] = finance_stats['cash_shipments_without_invoice']
    row += 1

    # Agency
    ws[f'A{row}'] = "Agency (From Ledger)"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    ws[f'A{row}'] = "Agency Total Debit (NPR)"
    ws[f'B{row}'] = finance_stats['agency_total_debit']
    row += 1

    ws[f'A{row}'] = "Agency Total Credit (NPR)"
    ws[f'B{row}'] = finance_stats['agency_total_credit']
    row += 1

    ws[f'A{row}'] = "Agency Shipments Without Invoice"
    ws[f'B{row}'] = finance_stats['agency_shipments_without_invoice']

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="dashboard_report_{date_from}_to_{date_to}.xlsx"'

    wb.save(response)
    return response
