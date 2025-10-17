from datetime import datetime
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import Workbook
import openpyxl
import zipfile
import csv
from openpyxl.styles import Border, Side, Alignment, Font

from awb.pages.awb.utils import convert_to_words_with_cents
from awb.models.box_details import BoxDetails
from hub.models.run import Run, RunAWB
import io
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from datetime import date
from openpyxl.worksheet.page import PageMargins


class RunAWBExporter:
    def __init__(self, run, awb_details):
        self.run = run
        self.details = awb_details
    # Helper function to convert values properly

    # Helper function to convert values properly
    def convert_value(self, value, is_uppercase=False):
        try:
            if is_uppercase:
                return str(value).upper()

            if value is None or value == "":
                return ""

            # Try converting to string first
            str_value = str(value).strip()

            # Try converting to float
            try:
                float_val = float(str_value)
                if float_val.is_integer():
                    return int(float_val)
                else:
                    return float_val
            except (ValueError, TypeError):
                pass  # Not a numeric value, move on to uppercase

            # Return as uppercase string
            return str_value.upper()

        except Exception as e:
            # Catch all other unexpected errors
            return str(value).upper() if value is not None else ""

    # Add this import at the top of your file
    def cfl_excel_unx(self, billing=False, hawb=True):
        # Create a new workbook and select the active worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "CFL Data Billing" if billing else "CFL Data"

        # Base fields (headers)
        headers = [
            "manifest_number", "flight_number", "flight_date", "mawb_number", "hawb_number", "reference_number",
            "mawb_origin", "mawb_destination", "total_bags", "total_weight", "manifest_value_type",
            "mawb_shipper_name", "mawb_shipper_street_address_line_1", "mawb_shipper_street_address_line_2",
            "mawb_shipper_city", "mawb_shipper_county_or_state", "mawb_shipper_postal_code",
            "mawb_shipper_country_code", "mawb_shipper_tel", "mawb_shipper_email",
            "mawb_consignee_name", "mawb_consignee_street_address_line_1", "mawb_consignee_street_address_line_2",
            "mawb_consignee_city", "mawb_consignee_county_or_state", "mawb_consignee_postal_code",
            "mawb_consignee_country_code", "mawb_consignee_tel", "mawb_consignee_email",
            "consignment_number", "shipper_name", "shipper_street_address_line_1", "shipper_street_address_line_2",
            "shipper_city", "shipper_county_or_state", "shipper_postal_code", "shipper_country_code",
            "shipper_tel", "shipper_email", "consignee_name", "consignee_street_address_line_1",
            "consignee_street_address_line_2", "consignee_city", "consignee_county_or_state",
            "consignee_postal_code", "consignee_country_code", "consignee_tel", "consignee_email",
            "pieces", "weight", "description", "value", "value_currency_code", "service_info", "bag_numbers"
        ]

        if billing:
            headers.insert(5, "reference_number")

        # Precompute static values to avoid repeated calculations
        mawb_no = str(self.run.mawb_no)
        unique_bag_count = str(self.run.unique_bag_count)
        total_actual_weight = self.convert_value(
            str(self.run.total_actual_weight))

        # Define alignment for center positioning with margins
        center_alignment = Alignment(
            horizontal='center',
            vertical='center',
            indent=1
        )

        bold_font = Font(bold=True)

        # Write headers to the first row using batch operation
        header_values = [self.convert_value(header) for header in headers]
        for col_num, value in enumerate(header_values, 1):
            cell = worksheet.cell(row=1, column=col_num, value=value)
            cell.font = bold_font
            cell.alignment = center_alignment

        # Prepare all data rows at once for better performance
        all_row_data = []
        # Track max lengths for column width
        max_lengths = [len(str(header)) for header in headers]

        for detail in self.details:
            awb = detail.get("awb_details", {}) or {}
            consignor = detail.get("consignor", {}) or {}
            consignee = detail.get("consignee", {}) or {}
            boxes = detail.get("boxes", []) or []

            # Precompute bag numbers string
            bag_numbers_str = ", ".join(
                [str(box.get("bag_no", "")) for box in boxes])

            # Optimize destination check
            dest_short_name = awb.get("destination_short_name", "").lower()
            is_ts = dest_short_name in ("us", "ca")

            # Optimize forwarding number extraction
            forwarding_number = awb.get("forwarding_number", "")
            first_forwarding_number = forwarding_number.split(
                ",")[0] if forwarding_number else ""

            # Precompute string values to avoid repeated str() calls
            awb_no = str(awb.get("awb_no", ""))
            reference_number = str(awb.get("reference_number", ""))
            consignor_name = str(consignor.get("name", ""))
            consignor_phone = str(consignor.get("phone", ""))
            consignor_email = str(consignor.get("email", ""))
            consignee_name = str(consignee.get("name", ""))
            consignee_address1 = str(consignee.get("address1", ""))
            consignee_address2 = consignee.get("address2", "")
            consignee_address2_str = f"{consignee_address2}, " if consignee_address2 else ""
            consignee_city = str(consignee.get("city", ""))
            consignee_state = str(consignee.get("state", ""))
            consignee_postcode = str(consignee.get("postcode", ""))
            consignee_country = str(consignee.get("country_short_name", ""))
            consignee_phone = str(consignee.get("phone", ""))
            consignee_email = str(consignee.get("email", ""))
            box_count = str(awb.get("box_count", ""))
            awb_weight = self.convert_value(
                str(awb.get("total_actual_weight", "")))
            content = str(awb.get("content", ""))
            shipment_value = self.convert_value(
                str(awb.get("shipment_value", "")))
            currency = str(awb.get("currency", ""))

            row_data = [
                "",  # manifest_number
                "",  # flight_number
                "",  # flight_date
                mawb_no,  # mawb_number
                awb_no,  # hawb_number
                reference_number,  # reference_number
                "NP",  # mawb_origin
                "GB",  # mawb_destination
                unique_bag_count,  # total_bags
                total_actual_weight,  # total_weight
                "TS" if is_ts else "HV",  # manifest_value_type

                # MAWB Shipper (hardcoded APEX info)
                "IDEAL CARGO",
                "GAUSHALA,BATTISPUTALI",
                "",
                "KATHMANDU",
                "",
                "44600",
                "NP",
                "",
                "",

                # MAWB Consignee
                "WPX SERVICES LTD",
                "6 POYLE RD",
                "",
                "COLNBROOK",
                "",
                "SL3 0AA",
                "GB",
                "",
                "",

                first_forwarding_number,  # consignment_number

                # Shipment Consignor
                consignor_name,
                "GAUSHALA,BATTISPUTALI",
                "",
                "KATHMANDU",
                "",
                "44600",
                "NP",
                consignor_phone,
                consignor_email,

                # Shipment Consignee
                consignee_name,
                consignee_address1,
                consignee_address2_str,
                consignee_city,
                consignee_state,
                consignee_postcode,
                consignee_country,
                consignee_phone,
                consignee_email,

                box_count,  # pieces
                awb_weight,  # weight
                content,  # description
                shipment_value,  # value
                currency,  # value_currency_code
                "",  # service_info
                bag_numbers_str  # bag_numbers
            ]

            if billing:
                row_data.insert(5, reference_number)

            all_row_data.append(row_data)

            # Update max lengths for column width calculation
            for i, value in enumerate(row_data):
                if i < len(max_lengths):
                    value_str = str(value) if value is not None else ""
                    max_lengths[i] = max(max_lengths[i], len(value_str))
        # Write all data rows at once using batch operation
        for row_num, row_data in enumerate(all_row_data, 3):  # Start from row 3
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                cell.alignment = center_alignment

        # Optimize column width calculation using precomputed max lengths
        for col_num, max_length in enumerate(max_lengths, 1):
            column_letter = worksheet.cell(row=1, column=col_num).column_letter
            adjusted_width = max(max_length + 2, 8)  # Minimum width of 8
            adjusted_width = min(adjusted_width, 80)  # Maximum width of 80
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Create HTTP response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="CFL Excel {self.run.run_no}.xlsx"'

        return response

    def export_invoice(self):
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # === Define Styles ===
        bold = Font(bold=True)
        bold_large = Font(bold=True, size=14)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        # Updated border style to match second design
        large_pt_side = Side(style="thin", color="000000")
        thick_border = Border(
            left=large_pt_side,
            right=large_pt_side,
            top=large_pt_side,
            bottom=large_pt_side
        )

        # Iterate through the run's details to create a separate sheet for each AWB
        for idx, detail in enumerate(self.details):
            # Create a new sheet for each AWB
            awb = detail.get("awb_details", {})
            boxes = detail.get("boxes", [])
            sheet_name = f"AWB_{detail.get('awb_no', '')}"
            ws = wb.create_sheet(title=sheet_name)
            row = 1

            # === 1) Main Header ===
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row=row, column=1, value=self.convert_value(
                "INVOICE & PACKING LIST"))
            cell.font = bold_large
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border
            # Add borders to merged cells
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thick_border
            row += 1

            # === 2) Top Info Section ===
            # Left column info
            awb = detail.get("awb_details", {})
            consignee = detail.get("consignee", {}) or {}
            consignor = detail.get("consignor", {}) or {}

            # row1
            ws.cell(row=row, column=1,
                    value=self.convert_value("COUNTRY OF ORIGIN: NEPAL")).border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=large_pt_side,
                bottom=Side(style=None)
            )
            ws.merge_cells(f'A{row}:C{row}')
            # Right column info
            ws.cell(row=row, column=4,
                    value=self.convert_value(f"ACTUAL WEIGHT :{awb.get('total_actual_weight')}")).border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=large_pt_side,
                bottom=Side(style=None)
            )
            ws.merge_cells(f'D{row}:H{row}')
            row += 1

            # row 2
            # Invoice Date
            ws.cell(row=row, column=1,
                    value=self.convert_value(f"INVOICE DATE. : {awb.get('booking_datetime').strftime('%B %d, %Y')}")).border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=Side(style=None),
                bottom=Side(style=None)
            )
            ws.merge_cells(f'A{row}:C{row}')
            # Total Pieces
            ws.cell(row=row, column=4,
                    value=self.convert_value(f"TOTAL PIECES :{awb.get('box_count')}")).border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=Side(style=None),
                bottom=Side(style=None)
            )
            ws.merge_cells(f'D{row}:H{row}')
            row += 1

            # row 3
            # Invoice Number
            ws.cell(row=row, column=1,
                    value=self.convert_value(f"INVOICE NO: {detail.get('awb_no', '')}")).border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=Side(style=None),
                bottom=Side(style=None)
            )
            ws.merge_cells(f'A{row}:C{row}')

            ws.cell(row=row, column=4, value="").border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=Side(style=None),
                bottom=Side(style=None)
            )
            ws.merge_cells(f'D{row}:H{row}')
            row += 1

            # row 4 - empty row
            ws.cell(row=row, column=1, value="").border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=Side(style=None),
                bottom=large_pt_side
            )
            ws.merge_cells(f'A{row}:C{row}')

            ws.cell(row=row, column=4, value="").border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=Side(style=None),
                bottom=large_pt_side
            )
            ws.merge_cells(f'D{row}:H{row}')
            row += 1

            # === 3) Shipper/Consignee Section ===
            # Headers
            ws.cell(row=row, column=1, value=self.convert_value(
                "SHIPPER")).font = bold
            ws.cell(row=row, column=1).border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=large_pt_side,
                bottom=Side(style=None)
            )
            ws.merge_cells(f'A{row}:C{row}')
            ws.cell(row=row, column=4, value=self.convert_value(
                "CONSIGNEE")).font = bold
            ws.cell(row=row, column=4).border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=large_pt_side,
                bottom=Side(style=None)
            )
            ws.merge_cells(f'D{row}:H{row}')
            row += 1

            # Shipper and Consignee Details
            shipper_info = [
                self.convert_value(consignor.get("company")),
                self.convert_value(consignor.get("name")),
                self.convert_value(consignor.get("address1")),
                self.convert_value(consignor.get("postcode")),
                self.convert_value(
                    f"{consignor.get('city')},{consignor.get('state')}"),
                self.convert_value(consignor.get("country", "")),
                self.convert_value(f"EMAIL: {consignor.get('email')}"),
                self.convert_value(f"PHONE NUMBER: +{consignor.get('phone')}")
            ]

            consignee_info = [
                self.convert_value(consignee.get("company")),
                self.convert_value(consignee.get("name")),
                self.convert_value(consignee.get("address1")),
                self.convert_value(consignee.get("postcode")),
                self.convert_value(
                    f"{consignee.get('city')},{consignee.get('state')}"),
                self.convert_value(consignee.get("country", "")),
                self.convert_value(f"EMAIL: {consignee.get('email')}"),
                self.convert_value(f"PHONE NUMBER: +{consignee.get('phone')}")
            ]

            for s_info, c_info in zip(shipper_info, consignee_info):
                cell_s = ws.cell(row=row, column=1,
                                 value=self.convert_value(s_info))
                cell_s.border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=Side(style=None),
                    bottom=Side(style=None)
                )
                cell_s.alignment = left
                ws.merge_cells(f'A{row}:C{row}')

                cell_c = ws.cell(row=row, column=4,
                                 value=self.convert_value(c_info))
                cell_c.border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=Side(style=None),
                    bottom=Side(style=None)
                )
                cell_c.alignment = left
                ws.merge_cells(f'D{row}:H{row}')
                row += 1

            # empty row
            ws.cell(row=row, column=1, value="").border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=Side(style=None),
                bottom=large_pt_side
            )
            ws.merge_cells(f'A{row}:C{row}')

            ws.cell(row=row, column=4, value="").border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=Side(style=None),
                bottom=large_pt_side
            )
            ws.merge_cells(f'D{row}:H{row}')
            row += 1

            # === 4) Items Table ===
            # Table Headers
            headers = ["BOXES", "SR NO", "DESCRIPTION", "HS CODE",
                       "UNIT TYPE", "QUANTITY", "UNIT RATE", "AMOUNT (USD)"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx,
                               value=self.convert_value(header))
                cell.font = bold
                cell.alignment = center
                cell.border = thick_border
            row += 1

            # Table Content
            sr_no = 1

            for box in boxes:
                # Ensure box_items is a list, even if empty
                box_items = box.get("items", [])
                if not isinstance(box_items, list):
                    box_items = []

                start_row = row

                # Write box number once - fix the box number display
                box_number = box.get('box_number')
                # print("box_number", box_number)
                if box_number is None:
                    # Try alternative keys if box_number is not available
                    box_number = box.get('box_no') or box.get(
                        'box') or box.get('number') or "BOX"

                box_cell = ws.cell(row=start_row, column=1,
                                   value=self.convert_value(f"BOX{box_number}"))
                box_cell.alignment = center
                box_cell.border = thick_border

                for item in box_items:
                    # Item details
                    data = [
                        self.convert_value(
                            f'BOX {item.get("box_number", box_number)}'),
                        self.convert_value(sr_no),
                        self.convert_value(item.get("description") or ""),
                        self.convert_value(item.get("hs_code") or ""),
                        self.convert_value(item.get("unit_type") or ""),
                        self.convert_value(item.get("quantity") or ""),
                        self.convert_value(item.get("unit_rate") or ""),
                        self.convert_value(item.get("amount") or "")
                    ]

                    for col_idx, value in enumerate(data, start=1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.alignment = center
                        cell.border = thick_border

                    sr_no += 1
                    row += 1

                # Merge box number cells if there are multiple items
                if len(box_items) > 1:
                    ws.merge_cells(start_row=start_row,
                                   start_column=1, end_row=row-1, end_column=1)

            # === 5) Auto-adjust Column Widths and Row Heights ===
            # Auto-fit all column widths
            for column in range(1, 9):  # Columns A through H
                max_length = 0
                column_letter = get_column_letter(column)

                # Check all cells in the column
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=column)
                    try:
                        if cell.value:
                            # Handle different types of values and calculate proper width
                            cell_value = str(cell.value).strip()

                            # For wrapped text, consider line breaks
                            if '\n' in cell_value:
                                lines = cell_value.split('\n')
                                cell_length = max(len(line) for line in lines)
                            else:
                                cell_length = len(cell_value)

                            max_length = max(max_length, cell_length)
                    except:
                        continue

                # Set width with padding but no fixed limits
                adjusted_width = max_length + 3 if max_length > 0 else 10
                ws.column_dimensions[column_letter].width = adjusted_width

            # Auto-fit row heights
            for row_num in range(1, ws.max_row + 1):
                max_height = 0  # Start with 0 to let content determine height

                for col_num in range(1, 9):  # Check all columns in each row
                    cell = ws.cell(row=row_num, column=col_num)

                    if cell.value:
                        try:
                            cell_value = str(cell.value).strip()

                            # Calculate height based on text length and column width
                            column_letter = get_column_letter(col_num)
                            column_width = ws.column_dimensions[column_letter].width or 10

                            # Estimate lines needed based on text length and column width
                            if len(cell_value) > column_width:
                                estimated_lines = max(
                                    1, len(cell_value) // int(column_width))
                                # Add extra lines for explicit line breaks
                                if '\n' in cell_value:
                                    estimated_lines += cell_value.count('\n')

                                # Calculate height (approximately 15 points per line)
                                calculated_height = estimated_lines * 15
                                max_height = max(max_height, calculated_height)

                            # Handle explicit line breaks
                            elif '\n' in cell_value:
                                line_count = cell_value.count('\n') + 1
                                calculated_height = line_count * 15
                                max_height = max(max_height, calculated_height)
                            else:
                                # Single line content
                                max_height = max(max_height, 15)

                        except:
                            continue

                # Set the row height based on content (minimum 15 for readability)
                if max_height > 0:
                    ws.row_dimensions[row_num].height = max(15, max_height)
                else:
                    ws.row_dimensions[row_num].height = 15

            # === 6) Add Footer Section ===
            # Calculate total quantity
            total_quantity = 0
            grand_total = 0

            # Safely calculate totals
            for box in boxes:
                box_items = box.get("items", [])
                if not isinstance(box_items, list):
                    box_items = []

                for item in box_items:
                    try:
                        # Get quantity and amount with fallbacks
                        quantity = item.get("quantity", 0)
                        amount = item.get("amount", 0)

                        # If amount is not available, try to calculate it from unit_rate and quantity
                        if not amount and item.get("unit_rate") and item.get("quantity"):
                            try:
                                unit_rate = float(item.get("unit_rate", 0))
                                qty = float(item.get("quantity", 0))
                                amount = unit_rate * qty
                            except (ValueError, TypeError):
                                pass

                        # Convert to numeric values if they're strings
                        if isinstance(quantity, str):
                            quantity = float(quantity) if quantity else 0
                        if isinstance(amount, str):
                            amount = float(amount) if amount else 0

                        total_quantity += quantity
                        grand_total += amount
                    except (ValueError, TypeError):
                        # Skip invalid values
                        continue

            # If totals are still 0, try to get values from the AWB details
            if total_quantity == 0 and grand_total == 0:
                try:
                    # Try to get total quantity from AWB
                    total_quantity = float(awb.get("total_quantity", 0))
                except (ValueError, TypeError):
                    pass

                try:
                    # Try to get total amount from AWB
                    grand_total = float(awb.get("shipment_value", 0))
                except (ValueError, TypeError):
                    pass

            # Add Total Quantity row
            ws.cell(row=row, column=1,

                    value="Total Quantity").border = thick_border
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=6, value=self.convert_value(
                total_quantity)).border = thick_border
            ws.cell(row=row, column=7, value="Grand Total").border = thick_border
            ws.cell(row=row, column=8, value=self.convert_value(
                grand_total)).border = thick_border
            row += 1

            grandtotal_in_words = convert_to_words_with_cents(grand_total)
            grandtotal_in_words = grandtotal_in_words.upper()
            # Add Total Amount row
            ws.cell(row=row, column=1,
                    value=self.convert_value(grandtotal_in_words)).border = thick_border
            ws.merge_cells(f'A{row}:F{row}')
            ws.cell(row=row, column=7,
                    value=self.convert_value(f"Total:{grand_total}")).border = thick_border
            ws.merge_cells(f'G{row}:H{row}')
            row += 1

            # Add Notes Section
            ws.cell(row=row, column=1, value=self.convert_value(
                "NOTES")).border = thick_border
            ws.merge_cells(f'A{row}:C{row}')
            ws.cell(row=row, column=4,
                    value=self.convert_value("SIGNATURE / STAMP")).border = thick_border
            ws.merge_cells(f'D{row}:H{row}')
            row += 1

            # Add Declaration Note
            declaration_cell = ws.cell(
                row=row,
                column=1,
                value=self.convert_value(
                    "WE DECLARE THAT THE ABOVE MENTIONED GOODS ARE MADE IN NEPAL AND OTHER DESCRIPTIONS ARE TRUE.")
            )
            declaration_cell.border = thick_border
            declaration_cell.alignment = Alignment(
                vertical="center", wrap_text=True)
            ws.merge_cells(f'A{row}:C{row}')

            signature_cell = ws.cell(row=row, column=4, value="")
            signature_cell.border = thick_border
            signature_cell.alignment = Alignment(
                vertical="center")
            ws.merge_cells(f'D{row}:H{row}')
            row += 1

            # Final adjustment for the last few rows that were added after the main auto-fit
            for row_num in range(row - 4, row):
                if row_num > 0:
                    max_height = 0
                    for col_num in range(1, 9):
                        cell = ws.cell(row=row_num, column=col_num)
                        if cell.value:
                            try:
                                cell_value = str(cell.value).strip()
                                column_letter = get_column_letter(col_num)
                                column_width = ws.column_dimensions[column_letter].width or 10

                                if len(cell_value) > column_width:
                                    estimated_lines = max(
                                        1, len(cell_value) // int(column_width))
                                    calculated_height = estimated_lines * 15
                                    max_height = max(
                                        max_height, calculated_height)
                                elif '\n' in cell_value:
                                    line_count = cell_value.count('\n') + 1
                                    calculated_height = line_count * 15
                                    max_height = max(
                                        max_height, calculated_height)
                                else:
                                    max_height = max(max_height, 15)
                            except:
                                continue

                    if max_height > 0:
                        ws.row_dimensions[row_num].height = max(15, max_height)
                    else:
                        ws.row_dimensions[row_num].height = 15

            # Manual column width adjustments
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 48

            # === Page Setup: IMPROVED horizontal centering ===
            # Set page orientation to portrait (default)
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

            # Enable fit to page
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.page_setup.fitToPage = True

            # IMPORTANT: Center horizontally on the printed page
            ws.print_options.horizontalCentered = True
            # Only center horizontally, not vertically
            ws.print_options.verticalCentered = False

            # Alternative method - also set page_setup horizontal centering
            ws.page_setup.horizontalCentered = True

            # Set custom margins (in inches)
            ws.page_margins = PageMargins(
                left=0.2,
                right=0.2,
                top=0.2,
                bottom=0.2,
                header=0.2,
                footer=0.2
            )

        # Export
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename=AWB INVOICES{self.run}.xlsx'
        return response

    def export_invoice_zip(self):
        # Create a BytesIO object to hold the zip file
        zip_buffer = BytesIO()

        # Create the zip file
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:

            # === Define Styles (moved outside loop for efficiency) ===
            bold = Font(bold=True)
            bold_large = Font(bold=True, size=14)
            center = Alignment(horizontal="center", vertical="center")
            left = Alignment(horizontal="left", vertical="center")

            # Updated border style to match second design
            large_pt_side = Side(style="thin", color="000000")
            thick_border = Border(
                left=large_pt_side,
                right=large_pt_side,
                top=large_pt_side,
                bottom=large_pt_side
            )

            # Iterate through the run's details to create a separate file for each AWB
            for idx, detail in enumerate(self.details):
                # Create a new workbook for each AWB
                wb = Workbook()

                # Remove default sheet
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]

                # Create a single sheet for this AWB
                awb = detail.get("awb_details", {})
                boxes = detail.get("boxes", [])
                awb_no = detail.get('awb_no', f'AWB_{idx+1}')

                ws = wb.create_sheet(title=f"Invoice_{awb_no}")
                row = 1

                # === 1) Main Header ===
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws.cell(row=row, column=1, value=self.convert_value(
                    "INVOICE & PACKING LIST"))
                cell.font = bold_large
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
                cell.border = thick_border
                # Add borders to merged cells
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = thick_border
                row += 1

                # === 2) Top Info Section ===
                # Left column info
                awb = detail.get("awb_details", {})
                consignee = detail.get("consignee", {}) or {}
                consignor = detail.get("consignor", {}) or {}

                # row1
                ws.cell(row=row, column=1,
                        value=self.convert_value("COUNTRY OF ORIGIN: NEPAL")).border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=large_pt_side,
                    bottom=Side(style=None)
                )
                ws.merge_cells(f'A{row}:C{row}')
                # Right column info
                ws.cell(row=row, column=4,
                        value=self.convert_value(f"ACTUAL WEIGHT :{awb.get('total_actual_weight')}")).border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=large_pt_side,
                    bottom=Side(style=None)
                )
                ws.merge_cells(f'D{row}:H{row}')
                row += 1

                # row 2
                # Invoice Date
                ws.cell(row=row, column=1,
                        value=self.convert_value(f"INVOICE DATE. : {awb.get('booking_datetime').strftime('%B %d, %Y')}")).border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=Side(style=None),
                    bottom=Side(style=None)
                )
                ws.merge_cells(f'A{row}:C{row}')
                # Total Pieces
                ws.cell(row=row, column=4,
                        value=self.convert_value(f"TOTAL PIECES :{awb.get('box_count')}")).border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=Side(style=None),
                    bottom=Side(style=None)
                )
                ws.merge_cells(f'D{row}:H{row}')
                row += 1

                # row 3
                # Invoice Number
                ws.cell(row=row, column=1,
                        value=self.convert_value(f"INVOICE NO: {detail.get('awb_no', '')}")).border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=Side(style=None),
                    bottom=Side(style=None)
                )
                ws.merge_cells(f'A{row}:C{row}')

                ws.cell(row=row, column=4, value="").border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=Side(style=None),
                    bottom=Side(style=None)
                )
                ws.merge_cells(f'D{row}:H{row}')
                row += 1

                # row 4 - empty row
                ws.cell(row=row, column=1, value="").border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=Side(style=None),
                    bottom=large_pt_side
                )
                ws.merge_cells(f'A{row}:C{row}')

                ws.cell(row=row, column=4, value="").border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=Side(style=None),
                    bottom=large_pt_side
                )
                ws.merge_cells(f'D{row}:H{row}')
                row += 1

                # === 3) Shipper/Consignee Section ===
                # Headers
                ws.cell(row=row, column=1, value=self.convert_value(
                    "SHIPPER")).font = bold
                ws.cell(row=row, column=1).border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=large_pt_side,
                    bottom=Side(style=None)
                )
                ws.merge_cells(f'A{row}:C{row}')
                ws.cell(row=row, column=4, value=self.convert_value(
                    "CONSIGNEE")).font = bold
                ws.cell(row=row, column=4).border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=large_pt_side,
                    bottom=Side(style=None)
                )
                ws.merge_cells(f'D{row}:H{row}')
                row += 1

                # Shipper and Consignee Details
                shipper_info = [
                    self.convert_value(consignor.get("company")),
                    self.convert_value(consignor.get("name")),
                    self.convert_value(consignor.get("address1")),
                    self.convert_value(consignor.get("postcode")),
                    self.convert_value(
                        f"{consignor.get('city')},{consignor.get('state')}"),
                    self.convert_value(consignor.get("country", "")),
                    self.convert_value(f"EMAIL: {consignor.get('email')}"),
                    self.convert_value(
                        f"PHONE NUMBER: +{consignor.get('phone')}")
                ]

                consignee_info = [
                    self.convert_value(consignee.get("company")),
                    self.convert_value(consignee.get("name")),
                    self.convert_value(consignee.get("address1")),
                    self.convert_value(consignee.get("postcode")),
                    self.convert_value(
                        f"{consignee.get('city')},{consignee.get('state')}"),
                    self.convert_value(consignee.get("country", "")),
                    self.convert_value(f"EMAIL: {consignee.get('email')}"),
                    self.convert_value(
                        f"PHONE NUMBER: +{consignee.get('phone')}")
                ]

                for s_info, c_info in zip(shipper_info, consignee_info):
                    cell_s = ws.cell(row=row, column=1,
                                     value=self.convert_value(s_info))
                    cell_s.border = Border(
                        left=large_pt_side,
                        right=large_pt_side,
                        top=Side(style=None),
                        bottom=Side(style=None)
                    )
                    cell_s.alignment = left
                    ws.merge_cells(f'A{row}:C{row}')

                    cell_c = ws.cell(row=row, column=4,
                                     value=self.convert_value(c_info))
                    cell_c.border = Border(
                        left=large_pt_side,
                        right=large_pt_side,
                        top=Side(style=None),
                        bottom=Side(style=None)
                    )
                    cell_c.alignment = left
                    ws.merge_cells(f'D{row}:H{row}')
                    row += 1

                # empty row
                ws.cell(row=row, column=1, value="").border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=Side(style=None),
                    bottom=large_pt_side
                )
                ws.merge_cells(f'A{row}:C{row}')

                ws.cell(row=row, column=4, value="").border = Border(
                    left=large_pt_side,
                    right=large_pt_side,
                    top=Side(style=None),
                    bottom=large_pt_side
                )
                ws.merge_cells(f'D{row}:H{row}')
                row += 1

                # === 4) Items Table ===
                # Table Headers
                headers = ["BOXES", "SR NO", "DESCRIPTION", "HS CODE",
                           "UNIT TYPE", "QUANTITY", "UNIT RATE", "AMOUNT (USD)"]
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col_idx,
                                   value=self.convert_value(header))
                    cell.font = bold
                    cell.alignment = center
                    cell.border = thick_border
                row += 1

                # Table Content
                sr_no = 1

                for box in boxes:
                    # Ensure box_items is a list, even if empty
                    box_items = box.get("items", [])
                    if not isinstance(box_items, list):
                        box_items = []

                    start_row = row

                    # Write box number once - fix the box number display
                    box_number = box.get('box_number')
                    print("box_number", box_number)
                    if box_number is None:
                        # Try alternative keys if box_number is not available
                        box_number = box.get('box_no') or box.get(
                            'box') or box.get('number') or "BOX"

                    box_cell = ws.cell(row=start_row, column=1,
                                       value=self.convert_value(f"BOX{box_number}"))
                    box_cell.alignment = center
                    box_cell.border = thick_border

                    for item in box_items:
                        # Item details
                        data = [
                            self.convert_value(
                                f'BOX {item.get("box_number", box_number)}'),
                            self.convert_value(sr_no),
                            self.convert_value(item.get("description") or ""),
                            self.convert_value(item.get("hs_code") or ""),
                            self.convert_value(item.get("unit_type") or ""),
                            self.convert_value(item.get("quantity") or ""),
                            self.convert_value(item.get("unit_rate") or ""),
                            self.convert_value(item.get("amount") or "")
                        ]

                        for col_idx, value in enumerate(data, start=1):
                            cell = ws.cell(
                                row=row, column=col_idx, value=value)
                            cell.alignment = center
                            cell.border = thick_border

                        sr_no += 1
                        row += 1

                    # Merge box number cells if there are multiple items
                    if len(box_items) > 1:
                        ws.merge_cells(start_row=start_row,
                                       start_column=1, end_row=row-1, end_column=1)

                # === 5) Auto-adjust Column Widths and Row Heights ===
                # Auto-fit all column widths
                for column in range(1, 9):  # Columns A through H
                    max_length = 0
                    column_letter = get_column_letter(column)

                    # Check all cells in the column
                    for row_num in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row_num, column=column)
                        try:
                            if cell.value:
                                # Handle different types of values and calculate proper width
                                cell_value = str(cell.value).strip()

                                # For wrapped text, consider line breaks
                                if '\n' in cell_value:
                                    lines = cell_value.split('\n')
                                    cell_length = max(len(line)
                                                      for line in lines)
                                else:
                                    cell_length = len(cell_value)

                                max_length = max(max_length, cell_length)
                        except:
                            continue

                    # Set width with padding but no fixed limits
                    adjusted_width = max_length + 3 if max_length > 0 else 10
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Auto-fit row heights
                for row_num in range(1, ws.max_row + 1):
                    max_height = 0  # Start with 0 to let content determine height

                    for col_num in range(1, 9):  # Check all columns in each row
                        cell = ws.cell(row=row_num, column=col_num)

                        if cell.value:
                            try:
                                cell_value = str(cell.value).strip()

                                # Calculate height based on text length and column width
                                column_letter = get_column_letter(col_num)
                                column_width = ws.column_dimensions[column_letter].width or 10

                                # Estimate lines needed based on text length and column width
                                if len(cell_value) > column_width:
                                    estimated_lines = max(
                                        1, len(cell_value) // int(column_width))
                                    # Add extra lines for explicit line breaks
                                    if '\n' in cell_value:
                                        estimated_lines += cell_value.count(
                                            '\n')

                                    # Calculate height (approximately 15 points per line)
                                    calculated_height = estimated_lines * 15
                                    max_height = max(
                                        max_height, calculated_height)

                                # Handle explicit line breaks
                                elif '\n' in cell_value:
                                    line_count = cell_value.count('\n') + 1
                                    calculated_height = line_count * 15
                                    max_height = max(
                                        max_height, calculated_height)
                                else:
                                    # Single line content
                                    max_height = max(max_height, 15)

                            except:
                                continue

                    # Set the row height based on content (minimum 15 for readability)
                    if max_height > 0:
                        ws.row_dimensions[row_num].height = max(15, max_height)
                    else:
                        ws.row_dimensions[row_num].height = 15

                # === 6) Add Footer Section ===
                # Calculate total quantity
                total_quantity = 0
                grand_total = 0

                # Safely calculate totals
                for box in boxes:
                    box_items = box.get("items", [])
                    if not isinstance(box_items, list):
                        box_items = []

                    for item in box_items:
                        try:
                            # Get quantity and amount with fallbacks
                            quantity = item.get("quantity", 0)
                            amount = item.get("amount", 0)

                            # If amount is not available, try to calculate it from unit_rate and quantity
                            if not amount and item.get("unit_rate") and item.get("quantity"):
                                try:
                                    unit_rate = float(item.get("unit_rate", 0))
                                    qty = float(item.get("quantity", 0))
                                    amount = unit_rate * qty
                                except (ValueError, TypeError):
                                    pass

                            # Convert to numeric values if they're strings
                            if isinstance(quantity, str):
                                quantity = float(quantity) if quantity else 0
                            if isinstance(amount, str):
                                amount = float(amount) if amount else 0

                            total_quantity += quantity
                            grand_total += amount
                        except (ValueError, TypeError):
                            # Skip invalid values
                            continue

                # If totals are still 0, try to get values from the AWB details
                if total_quantity == 0 and grand_total == 0:
                    try:
                        # Try to get total quantity from AWB
                        total_quantity = float(awb.get("total_quantity", 0))
                    except (ValueError, TypeError):
                        pass

                    try:
                        # Try to get total amount from AWB
                        grand_total = float(awb.get("shipment_value", 0))
                    except (ValueError, TypeError):
                        pass

                # Add Total Quantity row
                ws.cell(row=row, column=1,
                        value="Total Quantity").border = thick_border
                ws.merge_cells(f'A{row}:E{row}')
                ws.cell(row=row, column=6, value=self.convert_value(
                    total_quantity)).border = thick_border
                ws.cell(row=row, column=7,
                        value="Grand Total").border = thick_border
                ws.cell(row=row, column=8, value=self.convert_value(
                    grand_total)).border = thick_border
                row += 1

                grandtotal_in_words = convert_to_words_with_cents(grand_total)
                grandtotal_in_words = grandtotal_in_words.upper()
                # Add Total Amount row
                ws.cell(row=row, column=1,
                        value=self.convert_value(grandtotal_in_words)).border = thick_border
                ws.merge_cells(f'A{row}:F{row}')
                ws.cell(row=row, column=7,
                        value=self.convert_value(f"Total:{grand_total}")).border = thick_border
                ws.merge_cells(f'G{row}:H{row}')
                row += 1

                # Add Notes Section
                ws.cell(row=row, column=1, value=self.convert_value(
                    "NOTES")).border = thick_border
                ws.merge_cells(f'A{row}:C{row}')
                ws.cell(row=row, column=4,
                        value=self.convert_value("SIGNATURE / STAMP")).border = thick_border
                ws.merge_cells(f'D{row}:H{row}')
                row += 1

                # Add Declaration Note
                declaration_cell = ws.cell(
                    row=row,
                    column=1,
                    value=self.convert_value(
                        "WE DECLARE THAT THE ABOVE MENTIONED GOODS ARE MADE IN NEPAL AND OTHER DESCRIPTIONS ARE TRUE.")
                )
                declaration_cell.border = thick_border
                declaration_cell.alignment = Alignment(
                    vertical="center", wrap_text=True)
                ws.merge_cells(f'A{row}:C{row}')

                signature_cell = ws.cell(row=row, column=4, value="")
                signature_cell.border = thick_border
                signature_cell.alignment = Alignment(
                    vertical="center")
                ws.merge_cells(f'D{row}:H{row}')
                row += 1

                # Final adjustment for the last few rows that were added after the main auto-fit
                for row_num in range(row - 4, row):
                    if row_num > 0:
                        max_height = 0
                        for col_num in range(1, 9):
                            cell = ws.cell(row=row_num, column=col_num)
                            if cell.value:
                                try:
                                    cell_value = str(cell.value).strip()
                                    column_letter = get_column_letter(col_num)
                                    column_width = ws.column_dimensions[column_letter].width or 10

                                    if len(cell_value) > column_width:
                                        estimated_lines = max(
                                            1, len(cell_value) // int(column_width))
                                        calculated_height = estimated_lines * 15
                                        max_height = max(
                                            max_height, calculated_height)
                                    elif '\n' in cell_value:
                                        line_count = cell_value.count('\n') + 1
                                        calculated_height = line_count * 15
                                        max_height = max(
                                            max_height, calculated_height)
                                    else:
                                        max_height = max(max_height, 15)
                                except:
                                    continue

                        if max_height > 0:
                            ws.row_dimensions[row_num].height = max(
                                15, max_height)
                        else:
                            ws.row_dimensions[row_num].height = 15

                # Manual column width adjustments
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 8
                ws.column_dimensions['C'].width = 48

                # === Page Setup: IMPROVED horizontal centering ===
                # Set page orientation to portrait (default)
                ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

                # Enable fit to page
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 1
                ws.page_setup.fitToPage = True

                # IMPORTANT: Center horizontally on the printed page
                ws.print_options.horizontalCentered = True
                # Only center horizontally, not vertically
                ws.print_options.verticalCentered = False

                # Alternative method - also set page_setup horizontal centering
                ws.page_setup.horizontalCentered = True

                # Set custom margins (in inches)
                ws.page_margins = PageMargins(
                    left=0.2,
                    right=0.2,
                    top=0.2,
                    bottom=0.2,
                    header=0.2,
                    footer=0.2
                )

                # Save each workbook to a BytesIO buffer
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)

                # Create filename for this AWB
                filename = f"Invoice_{awb_no}.xlsx"

                # Add the Excel file to the zip
                zip_file.writestr(filename, excel_buffer.getvalue())

                # Close the excel buffer
                excel_buffer.close()

        # Prepare the zip file for download
        zip_buffer.seek(0)

        response = HttpResponse(
            zip_buffer.getvalue(),
            content_type="application/zip"
        )
        response["Content-Disposition"] = f'attachment; filename=AWB_INVOICES_{self.run}.zip'

        return response

    def export_bag_details(self):
        run = self.run
        run_awbs = RunAWB.objects.filter(
            run=run).values_list('awb_id', flat=True)
        box_details = BoxDetails.objects.filter(
            awb_id__in=run_awbs).exclude(bag_no__isnull=True)

        bag_map = defaultdict(list)
        for box in box_details:
            if box.awb.vendor.name == "SGCA":
                bag_map[box.bag_no].append(str(box.awb.reference_number))
            else:
                bag_map[box.bag_no].append(str(box.awb.awbno))

        # Get the range of bag numbers (from min to max)
        if bag_map:
            min_bag = min(bag_map.keys())
            max_bag = max(bag_map.keys())

            # Create sequential bag list with empty lists for missing bags
            sequential_bags = []
            for bag_no in range(min_bag, max_bag + 1):
                if bag_no in bag_map:
                    sequential_bags.append((bag_no, bag_map[bag_no]))
                else:
                    # Empty list for missing bags
                    sequential_bags.append((bag_no, []))
        else:
            sequential_bags = []

        # Chunk the sequential bags
        bag_chunks = [sequential_bags[i:i + 10]
                      for i in range(0, len(sequential_bags), 10)]

        wb = Workbook()
        ws = wb.active
        ws.title = "Bag Details"

        # Styles
        header_font = Font(bold=True)
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        current_row = 1
        start_col = 1  # Start from column B
        grand_total = 0

        # Title row
        current_row += 1

        for chunk in bag_chunks:
            header_row = current_row
            data_start_row = header_row + 1
            max_awbs = max(len(awbs) for _, awbs in chunk) if chunk else 0
            total_col = start_col + 12

            # Header row
            for idx, (bag_no, _) in enumerate(chunk):
                col = start_col + idx
                cell = ws.cell(row=header_row, column=col)
                cell.value = self.convert_value(f"BAG # {bag_no}")
                cell.font = header_font
                cell.fill = yellow_fill
                cell.alignment = center
                cell.border = thin_border

            # TOTAL PCS header
            if chunk == bag_chunks[0]:
                total_header_cell = ws.cell(row=header_row, column=total_col)
                total_header_cell.value = self.convert_value("TOTAL PCS")
                total_header_cell.font = header_font
                total_header_cell.fill = yellow_fill
                total_header_cell.alignment = center
                total_header_cell.border = thin_border

            # Data rows
            for i in range(max_awbs):
                row = data_start_row + i
                row_total = 0
                for idx, (_, awbs) in enumerate(chunk):
                    col = start_col + idx
                    if i < len(awbs):
                        ws.cell(row=row, column=col).value = self.convert_value(
                            awbs[i])
                        row_total += 1
                    # If awbs is empty (missing bag), the cell will remain empty
                # Row total
                if row_total > 0:  # Only add to total if there are actual AWBs
                    ws.cell(row=row, column=total_col).value = row_total
                    grand_total += row_total

            current_row = data_start_row + max_awbs + 1  # Leave a space between chunks

        # Title with merged cells (C to F)
        title_cell = ws.cell(row=1, column=start_col + 2)
        title_cell.value = self.convert_value(
            f"BAG DETAILS #{run.run_no} #{run.mawb_no} / {grand_total}")
        title_cell.font = header_font
        title_cell.alignment = center
        ws.merge_cells(start_row=1, start_column=start_col + 2,
                       end_row=1, end_column=start_col + 5)  # Merge C to F
        ws.cell(row=current_row, column=start_col+2).font = header_font
        # Final grand total at bottom
        ws.cell(row=current_row, column=start_col + 12).value = grand_total
        ws.cell(row=current_row, column=start_col + 12).font = header_font

        # Auto-adjust column widths (excluding first row)
        for column in ws.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                try:
                    # Skip first row and merged cells
                    if cell.row > 1 and hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter

                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass

            # Only set width if we have a valid column letter
            if column_letter:
                # Set minimum width and add some padding
                # Max width of 50 to prevent excessive widths
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = max(
                    adjusted_width, 10)  # Minimum width of 10

        # Save and return
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"bag_details_run_{run.run_no}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        output = io.BytesIO()
        wb.save(output)
        response.write(output.getvalue())
        return response

    def export_us_bag_details(self):
        run = self.run

        run_awbs = RunAWB.objects.filter(
            run=run).values_list("awb_id", flat=True)
        box_details = BoxDetails.objects.filter(
            awb_id__in=run_awbs
        ).exclude(bag_no__isnull=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "US Bag Details"

        # Styles
        title_font = Font(bold=True, size=22)
        bold_font = Font(bold=True, size=11)
        center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Row 1: Title (no merge, just in A1)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "BAG DETAILS"
        title_cell.font = title_font
        title_cell.alignment = center
        # wrap text
        title_cell.alignment = Alignment(wrap_text=False)

        # Row 2: empty

        # Row 3: Headers (B, C, D, E, F)
        headers = {
            2: "S.N",
            3: "HAWB",
            4: "",  # Empty column with border
            5: "RECEIVER NAME",
            6: "BAG NO.",
        }
        for col, header in headers.items():
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center
            cell.border = thin_border
            # font 14
            cell.font = Font(size=14)
            # bold
            cell.font = Font(bold=True)

        # Data rows start from row 4
        current_row = 4
        for idx, box in enumerate(box_details, start=1):
            # Column B: S.N

            sn_cell = ws.cell(row=current_row, column=2, value=idx)
            sn_cell.font = bold_font
            sn_cell.alignment = center
            sn_cell.border = thin_border

            # Column C: HAWB
            hawb = (
                str(box.awb.reference_number)
                if box.awb.vendor.name == "SGUS"
                else str(box.awb.awbno)
            )
            hawb_cell = ws.cell(row=current_row, column=3, value=hawb)
            hawb_cell.font = bold_font
            hawb_cell.alignment = center
            hawb_cell.border = thin_border

            # Column D: empty but with border
            empty_cell = ws.cell(row=current_row, column=4, value="")
            empty_cell.font = bold_font
            empty_cell.border = thin_border

            # Column E: RECEIVER NAME (empty)
            recv_cell = ws.cell(row=current_row, column=5,
                                value=box.awb.consignee.person_name)
            recv_cell.font = bold_font
            recv_cell.border = thin_border

            # Column F: BAG NO.
            bag_cell = ws.cell(row=current_row, column=6, value=box.bag_no)
            bag_cell.font = bold_font
            bag_cell.alignment = center
            bag_cell.border = thin_border

            current_row += 1

        # Auto-adjust widths for B–F
        for col_letter in ["B", "C", "D", "E", "F"]:
            max_length = 0
            for cell in ws[col_letter]:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            ws.column_dimensions[col_letter].width = max(
                min(max_length + 2, 50), 10)

        # Save and return
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"us_bag_details_run_{run.run_no}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        output = io.BytesIO()
        wb.save(output)
        response.write(output.getvalue())
        return response

    def export_jfk_bom_custom(self):

        response = HttpResponse(content_type='text/csv')

        response["Content-Disposition"] = f'attachment; filename=Run_{self.run.run_no}_JFK_BOM_Custom.csv'

        writer = csv.writer(response)

        # Style
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=False)

        # All headers from CSV
        headers = [
            'SiteId', 'ArrivalAirport', 'WaybillOriginator', 'AirlinePrefix', 'AwbSerialNumber',
            'HouseAwb', 'MasterAwbIndicator', 'OriginAirport', 'Pieces', 'WeightCode', 'Weight',
            'Description', 'FdaIndicator', 'ImportingCarrier', 'FlightNumber', 'ArrivalDay',
            'ArrivalMonth', 'ShipperName', 'ShipperStreetAddress', 'ShipperCity',
            'ShipperStateOrProvince', 'ShipperPostalCode', 'ShipperCountry', 'ShipperTelephone',
            'Consignee', 'ConsigneeStreetAddress', 'ConsigneeCity', 'ConsigneeStateOrProvince',
            'ConsigneePostalCode', 'ConsigneeCountry', 'ConsigneeTelephone',
            'AmendmentFlag', 'AmendmentCode', 'AmendmentReason', 'PtpDestination',
            'PtpDestinationDay', 'PtpDestinationMonth', 'BoardedPieces', 'BoardedWeightCode',
            'BoardedWeight', 'PartialShipmentRef', 'BrokerCode', 'InbondDestination',
            'InbondDestinationType', 'BondedCarrierId', 'OnwardCarrier', 'BondedPremisesId',
            'TransferControlNumber', 'EntryType', 'EntryNumber', 'CountryOfOrigin', 'CustomsValue',
            'CurrencyCode', 'HtsNumber', 'ExpressRelease'
        ]

        writer.writerow(headers)

        # Header Row
        # for col_num, header in enumerate(headers, 1):
        #     cell = ws.cell(row=1, column=col_num, value=header)
        #     cell.font = Font(bold=True)
        #     cell.alignment = header_alignment
        #     cell.border = thin_border

        # Data Rows
        for detail in self.details:
            awb = detail.get("awb_details", {})
            consignor = detail.get("consignor", {}) or {}
            consignee = detail.get("consignee", {}) or {}

            row = [
                "",
                self.convert_value("JFK"),
                "",
                self.convert_value("157"),
                "",
                self.convert_value(detail.get("awb_no", "")),
                self.convert_value(self.run.mawb_no),
                self.convert_value("KTM"),
                self.convert_value(awb.get("box_count")),
                self.convert_value("K"),
                self.convert_value(awb.get("total_actual_weight")),
                self.convert_value(awb.get("content", "")),
                "",
                "",
                self.convert_value(self.run.flight_no),
                self.convert_value(
                    self.run.flight_departure_date.day) if self.run.flight_departure_date else "",
                self.convert_value(self.run.flight_departure_date.strftime(
                    "%B")) if self.run.flight_departure_date else "",

                self.convert_value(consignor.get("company")),
                self.convert_value(consignor.get("address1")),
                self.convert_value(consignor.get("city")),
                self.convert_value(consignor.get("state")),
                self.convert_value(consignor.get("postcode")),
                self.convert_value(consignor.get("country_short_name")),
                self.convert_value(consignor.get("phone")),

                self.convert_value(consignee.get("name")),
                self.convert_value(consignee.get("address1")),
                self.convert_value(consignee.get("city")),
                self.convert_value(consignee.get("state_short_name")),
                self.convert_value(consignee.get("postcode")),
                self.convert_value(consignee.get("country_short_name")),
                self.convert_value(consignee.get("phone")),
                self.convert_value("A"),
                self.convert_value("21"),
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                self.convert_value("NP"),
                self.convert_value(awb.get("shipment_value")),
                self.convert_value(awb.get("currency", "USD")),
                "",
                self.convert_value("Y")
            ]
            writer.writerow(row)

        return response

    def export_jfk_manifest(self):
        # Define headers exactly as per your format
        headers = [
            "Arrival Airport", "Airline Prefix", "AWB Serial Number", "House AWB",
            "Origin Airport", "Pieces", "Weight", "Description", "Importing Carrier",
            "Shipper Name", "Shipper Street Address", "Shipper City", "Shipper Country",
            "Consignee Name", "Consignee Street Address", "Consignee City", "Consignee State",
            "Consignee Postal Code", "Consignee Country", "Customs Value", "Currency Code",
            "HTS Code", "Barcode", "Barcode Transit Party", "ABV", "Quantity", "Height",
            "Width", "Length", "Shipper EORI", "Consignee Email", "Consignee Phone",
            "LMP Service", "Customer Transit Party", "Over Label Transit Party",
            "Over Label Service", "Over Label Dynamic", "Item Name", "Item Hscode",
            "Item Country", "Item Pieces", "Item Value", "Item Currency", "Item Weight",
            "Consignee Company Name", "Selling MID", "Incoterms"
        ]

        # Create buffer
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # Write header
        writer.writerow(headers)

        # Loop through details
        for detail in self.details:
            awb = detail.get("awb_details", {})
            boxes = detail.get("boxes", [])
            consignor = detail.get("consignor", {}) or {}
            consignee = detail.get("consignee", {}) or {}

            for box in boxes:
                items = box.get("items") or []

                for item in items:
                    hs_code = f"{item.get('hs_code', '')}00000"

                    row = [
                        "JFK",  # Arrival Airport
                        # Airline Prefix (default 125)
                        "QR",
                        awb.get("reference_number", ""),  # AWB Serial Number
                        awb.get("forwarding_number", ""),  # House AWB
                        "KTM",  # Origin Airport
                        "",  # Pieces
                        box.get("actual_weight", ""),  # Weight
                        item.get("description", ""),  # Description
                        "QR",  # Importing Carrier
                        consignor.get("name", ""),  # Shipper Name
                        # Shipper Street Address
                        consignor.get("address1", ""),
                        consignor.get("city", ""),  # Shipper City
                        consignor.get("country_short_name",
                                      "NP"),  # Shipper Country
                        consignee.get("name", ""),  # Consignee Name
                        # Consignee Street Address
                        consignee.get("address1", ""),
                        consignee.get("city", ""),  # Consignee City
                        # Consignee State
                        consignee.get("state_short_name", ""),
                        consignee.get("postcode", ""),  # Consignee Postal Code
                        # Consignee Country
                        consignee.get("country_short_name", ""),
                        awb.get("shipment_value", ""),  # Customs Value
                        "USD",  # Currency Code
                        hs_code[:10],  # HTS Code
                        awb.get("forwarding_number", ""),  # Barcode
                        "",  # Barcode Transit Party
                        "",  # ABV
                        "",  # Quantity
                        "", "", "",  # Height, Width, Length (blank)
                        consignor.get("eori", ""),  # Shipper EORI
                        consignee.get("email", ""),  # Consignee Email
                        consignee.get("phone", ""),  # Consignee Phone
                        "", "", "", "", "",  # LMP Service → Over Label Dynamic
                        item.get("description", ""),  # Item Name
                        hs_code[:10],  # Item Hscode
                        item.get("country", "NP"),  # Item Country
                        item.get("quantity", 1),  # Item Pieces
                        item.get("unit_rate", ""),  # Item Value
                        "USD",  # Item Currency
                        item.get("unit_weight", ""),  # Item Weight
                        consignee.get("company", ""),  # Consignee Company Name
                        "",  # Selling MID
                        awb.get("shipment_terms", ""),  # Incoterms
                    ]

                    writer.writerow(row)

        # Build HTTP response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="text/csv"
        )
        response["Content-Disposition"] = f'attachment; filename=JFK {self.run.run_no}.csv'

        return response

    def setup_worksheet(self, ws, title):
        # Apply all static formatting in one place
        self.apply_master_styles(ws, title)

    def apply_master_styles(self, ws, title):
        # Borders
        thick = Side(style='medium', color='000000')
        thin = Side(style='thin',   color='000000')
        # Fonts & fills
        center = Alignment(horizontal="center",
                           vertical="center", wrap_text=True)
        header_font = Font(bold=True)
        title_font = Font(bold=True, size=13)
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Title Row
        ws.merge_cells('A1:I1')
        c = ws['A1']
        c.value = title
        c.font = title_font
        c.alignment = center
        # Border below title
        for col in range(1, 10):
            right = thick if col == 9 else thin
            ws.cell(row=1, column=col).border = Border(
                left=thin, right=right, top=thin, bottom=thick)

        # Static label cells
        static_cells = {
            'A2': {'value': 'EXPORTER  & MANUFACTURER:', 'font': header_font},
            'A3': {'value': 'SHIP GLOBAL NEPAL', 'font': header_font},
            'A4': {'value': 'Gaushala Chowk, Battisputali Road', 'font': header_font},
            'A5': {'value': 'PH:', 'font': header_font},
            'A6': {'value': 'EMAIL: SHIPGLOBALNEPAL@GMAIL.COM', 'font': header_font},
            'F2': {'value': 'INV. NO. & DTD', 'font': header_font},
            'A8': {'value': 'Consignee', 'font': header_font},
            'F8': {'value': 'Buyer other than consignee', 'font': header_font},
            'A15': {'value': 'Pre - Carriage by', 'font': header_font},
            'C15': {'value': 'Place of Receipt by Pre-carrier', 'font': header_font},
            'F15': {'value': 'Country of Origin of Goods', 'font': header_font},
            'I15': {'value': 'Country of Final Destination', 'font': header_font},
            'A17': {'value': 'Vessel / Flight No.', 'font': header_font},
            'C17': {'value': 'Port of Loading', 'font': header_font},
            'F17': {'value': 'Terms of Delivery and Payment', 'font': header_font},
            'A19': {'value': 'Port of Discharge', 'font': header_font},
            'C19': {'value': 'Final Destination', 'font': header_font},
            'A21': {'value': 'Marks & Nos /', 'font': header_font},
            'A22': {'value': 'Container No.', 'font': header_font},
            'B21': {'value': 'Description of Goods', 'font': header_font, 'alignment': center},
            'D21': {'value': 'HSN', 'font': header_font, 'alignment': center},
            'E21': {'value': 'Bag', 'font': header_font, 'alignment': center},
            'F21': {'value': 'Quantity', 'font': header_font, 'alignment': center},
            'G21': {'value': 'Gross Weight', 'font': header_font, 'alignment': center},
            'H21': {'value': 'Rate', 'font': header_font, 'alignment': center},
            'I21': {'value': 'Amount', 'font': header_font, 'alignment': center},
            'D23': {'value': 'HSN', 'font': header_font},
            'E23': {'value': 'Total Bags', 'font': header_font},
            'F23': {'value': 'PCS', 'font': header_font},
            'H23': {'value': 'USD', 'font': header_font},
            'I23': {'value': 'USD', 'font': header_font},
        }
        for coord, props in static_cells.items():
            cell = ws[coord]
            for attr, val in props.items():
                setattr(cell, attr, val)

        # Yellow-fill section (buyer box columns F-H rows 8-13)
        for r in range(8, 13):
            for c_idx in range(6, 9):
                ws.cell(row=r, column=c_idx).fill = yellow_fill

        # Border grid for main area (rows 2-23, cols 1-9)
        separator_rows = {7, 14, 16, 18, 20, 21, 22, 23}
        for r in range(2, 24):
            for c_idx in range(1, 10):
                right = thick if c_idx in (5, 9) else thin
                bottom = thick if r in separator_rows else thin
                ws.cell(row=r, column=c_idx).border = Border(
                    left=thin, right=right, top=thin, bottom=bottom
                )

    def fill_invoice_data(self, ws):
        awb_details, consignee, consignor = self.get_common_data()
        # Dynamic assignment of fields like F3, consignee details, etc.
        # Invoice Number & Date
        inv_no = getattr(self.run, 'invoice_number', '')
        inv_dt = getattr(self.run, 'invoice_date', '')
        cell = ws['F3']
        cell.value = self.convert_value(f"{inv_no}  DT.{inv_dt}")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00",
                                end_color="FFFF00", fill_type="solid")

        # Consignee & buyer values
        ws['A9'].value = self.convert_value(consignee.get('name', ''))
        ws['A9'].font = Font(bold=True, size=13)
        ws['A10'].value = self.convert_value(consignee.get('address1', ''))
        ws['A11'].value = self.convert_value(consignee.get('address2', ''))
        ws['A12'].value = self.convert_value(
            f"EORI NO : {consignee.get('postcode', '')}")

        ws['F9'].value = self.convert_value(consignor.get('name', ''))
        ws['F10'].value = self.convert_value(consignor.get('address1', ''))
        ws['F11'].value = self.convert_value(consignor.get('address2', ''))
        ws['F12'].value = self.convert_value(
            consignor.get('country_short_name', ''))
        ws['F13'].value = self.convert_value(
            f"TEL : {consignor.get('phone', '')}")

        # AWB, origin, destination, flight, ports, terms
        ws['F16'].value = self.convert_value(awb_details.get('origin', ''))
        ws['I16'].value = self.convert_value(
            awb_details.get('destination', ''))
        ws['A18'].value = self.convert_value(
            getattr(self.run, 'flight_no', ''))
        ws['C18'].value = self.convert_value(
            getattr(self.run, 'port_loading', ''))
        ws['F18'].value = self.convert_value(
            getattr(self.run, 'terms_delivery', ''))
        ws['A20'].value = self.convert_value(
            getattr(self.run, 'port_discharge', ''))
        ws['C20'].value = self.convert_value(
            getattr(self.run, 'final_destination', ''))
        ws['C20'].fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws['F19'].value = self.convert_value(
            "**SHIPMENT UNDER TRANSIT VIA FRANKFURT TO USA!!")
        ws['F19'].fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Calculate totals and write product rows (existing logic)
        row = 24
        sr_no = 1
        total_qty = total_amt = total_gwt = 0.0
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        for detail in self.details:
            for box in detail.get('boxes', []):
                for item in box.get('items', []):
                    qty = float(item.get('quantity', 0))
                    rate = float(item.get('unit_rate', 0))
                    wt = float(item.get('unit_weight', 0))
                    amt = qty * rate
                    gwt = qty * wt
                    total_qty += qty
                    total_amt += amt
                    total_gwt += gwt
                    # Write borders
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).border = thin_border
                    # Fill cells
                    ws.cell(row=row, column=1, value=sr_no)
                    ws.cell(row=row, column=2,
                            value=self.convert_value(item.get('description', '')))
                    ws.cell(row=row, column=4, value=self.convert_value(
                        item.get('hs_code', '')))
                    ws.cell(row=row, column=6, value=self.convert_value(qty))
                    ws.cell(row=row, column=7, value=self.convert_value(gwt))
                    ws.cell(row=row, column=8, value=self.convert_value(rate))
                    ws.cell(row=row, column=9, value=self.convert_value(amt))
                    sr_no += 1
                    row += 1
        # Freight charges if any
        freight = getattr(self.run, 'freight_charges', 0)
        if freight:
            ws.cell(row=row, column=3,
                    value="FREIGHT CHARGES").font = Font(bold=True)
            ws.cell(row=row, column=9, value=self.convert_value(freight))
            total_amt += freight
            row += 1
        # Totals row
        row += 1
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=2, value=self.convert_value(
            "TOTAL")).font = Font(bold=True)
        ws.cell(row=row, column=6, value=self.convert_value(total_qty))
        ws.cell(row=row, column=7, value=self.convert_value(getattr(
            self.run, 'total_gross_weight', total_gwt)))
        ws.cell(row=row, column=8, value=self.convert_value("C&F"))
        ws.cell(row=row, column=9, value=self.convert_value(total_amt))
        # Amount in words
        row += 2
        text = convert_to_words_with_cents(total_amt).upper()
        ws.merge_cells(f'A{row}:G{row}')
        ws.cell(row=row, column=1,
                value=self.convert_value(f"Amount in words: {text}")).font = Font(bold=True)
        # Auto-adjust widths
        for col in range(1, 11):
            max_len = max((len(str(ws.cell(r, col).value)) for r in range(
                1, ws.max_row+1) if ws.cell(r, col).value), default=0)
            ws.column_dimensions[get_column_letter(col)].width = max_len + 2

    def fill_packing_data(self, ws):
        from collections import defaultdict
        bag_items = defaultdict(list)
        for d in self.details:
            for box in d.get('boxes', []):
                bag_items[box.get('bag_no', '')].extend(box.get('items', []))
        row = 24
        sr_no = 1
        total_qty = total_amt = total_gwt = 0.0
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        for bag_no, items in sorted(bag_items.items(), key=lambda x: str(x[0])):
            for item in items:
                qty = float(item.get('quantity', 0))
                rate = float(item.get('unit_rate', 0))
                wt = float(item.get('unit_weight', 0))
                amt = qty * rate
                gwt = qty * wt
                total_qty += qty
                total_amt += amt
                total_gwt += gwt
                for col in range(1, 10):
                    ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=1, value=self.convert_value(sr_no))
                ws.cell(row=row, column=2, value=self.convert_value(
                    item.get('description', '')))
                ws.cell(row=row, column=4, value=self.convert_value(
                    item.get('hs_code', '')))
                ws.cell(row=row, column=5, value=self.convert_value(bag_no))
                ws.cell(row=row, column=6, value=self.convert_value(qty))
                ws.cell(row=row, column=7, value=self.convert_value(gwt))
                ws.cell(row=row, column=8, value=self.convert_value(rate))
                ws.cell(row=row, column=9, value=self.convert_value(amt))
                sr_no += 1
                row += 1
        # Freight row
        freight = getattr(self.run, 'freight_charges', 0)
        if freight:
            ws.cell(row=row, column=3,
                    value="FREIGHT CHARGES").font = Font(bold=True)
            ws.cell(row=row, column=9, value=self.convert_value(freight))
            total_amt += freight
            row += 1
        # Totals row
        row += 1
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=2, value=self.convert_value(
            "TOTAL")).font = Font(bold=True)
        ws.cell(row=row, column=6, value=self.convert_value(total_qty))
        ws.cell(row=row, column=7, value=self.convert_value(getattr(
            self.run, 'total_gross_weight', total_gwt)))
        ws.cell(row=row, column=8, value=self.convert_value("C&F"))
        ws.cell(row=row, column=9, value=self.convert_value(total_amt))
        # Declarations
        row += 2
        ws.cell(row=row, column=1, value=self.convert_value(
            "Declaration :")).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1,
                value=self.convert_value("We declare that this invoice shows the actual price of the goods"))
        row += 1
        ws.cell(row=row, column=1,
                value=self.convert_value("described and that all particulars are true and correct."))
        # Amount in words
        row += 2
        text = convert_to_words_with_cents(total_amt).upper()
        ws.merge_cells(f'A{row}:G{row}')
        ws.cell(row=row, column=1,
                value=self.convert_value(f"Amount in words: {text}")).font = Font(bold=True)
        # Signature
        row += 2
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row, column=1,
                value=self.convert_value("For SHIP GLOBAL NEPAL")).font = Font(bold=True)

        ws.merge_cells(f'E{row}:G{row}')
        ws.cell(row=row, column=5,
                value=self.convert_value("Authorized Signatory")).font = Font(bold=True)

        # Auto-adjust column widths
        for col in range(1, 11):  # Columns A through J
            max_length = 0
            column_letter = get_column_letter(col)

            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    continue

            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

    def get_common_data(self):
        # Extract common data (consignee, consignor, etc.) from details
        awb_details = {}
        consignee = {}
        consignor = {}
        for detail in self.details:
            awb_details = detail.get("awb_details", {})
            consignee = detail.get("consignee", {}) or {}
            consignor = detail.get("consignor", {}) or {}
            break  # Assuming single AWB or taking first
        return awb_details, consignee, consignor

    def export_uk_manifest(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "UK Manifest"

        print("run", self.run)
        print("details", self.details)

        # Styles
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=False)

        # Header row
        headers = [
            "AWB NO", "PCS", "WT", "DESCRIPTION", "CONSIGNOR", "ADDRESS",
            "CONSIGNEE", "ADDRESS", "GBP", "BAG No."
        ]

        header_row = 1
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num,
                           value=self.convert_value(header))
            cell.font = Font(bold=True)
            cell.alignment = header_alignment
            cell.border = thin_border
            cell.fill = PatternFill(
                start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

        # Data rows
        for idx, detail in enumerate(self.details, start=header_row + 1):
            awb = detail.get("awb_details", {})
            consignee = detail.get("consignee", {}) or {}
            consignor = detail.get("consignor", {}) or {}

            # Format consignor name and address separately
            consignor_name = consignor.get("company", "")
            consignor_address_parts = [
                consignor.get("address1", "").upper(),
                consignor.get("city", "").upper(),
                consignor.get("postcode", "")
            ]
            consignor_address = ", ".join(
                filter(None, consignor_address_parts))

            # Format consignee name and address separately
            consignee_name = consignee.get("name", "")
            consignee_address_parts = [
                consignee.get("address1", "").upper(),
                consignee.get("city", "").upper(),
                consignee.get("postcode", "")
            ]
            consignee_address = ", ".join(
                filter(None, consignee_address_parts))

            row = [
                awb.get("reference_number"),  # AWB NO
                awb.get("box_count"),  # PCS
                awb.get("total_actual_weight"),  # WT
                awb.get("content", ""),  # DESCRIPTION
                consignor_name,  # CONSIGNOR
                consignor_address,  # CONSIGNOR ADDRESS
                consignee_name,  # CONSIGNEE
                consignee_address,  # CONSIGNEE ADDRESS
                awb.get('shipment_value'),  # GBP
                ", ".join(str(box.get("bag_no", ""))
                          for box in detail.get("boxes", [])),  # BAG No.
            ]

            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=idx, column=col_num,
                               value=self.convert_value(value))
                cell.alignment = center_alignment
                cell.border = thin_border

        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value))
                             if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = max_length + 4

        # Save and return response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename=Run_{self.run.run_no}_UK_Manifest.xlsx'
        return response

    def aus_manifest(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f" {self.run.run_no}"

        # Styles
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=False)

        headers = [
            "Parcel Barcode / HAWB Number",
            "HAWB Number",
            "Customer Name",
            "Customer Address 1",
            "Customer Address 2",
            "City",
            "State",
            "Postcode",
            "Customer Telephone (if available)",
            "Description of goods",
            "IATA Origin Port",
            "IATA Dest Port",
            "Flight No",
            "Master Bill",
            "Sub Master",
            "ETA",
            "Parcel count",
            "Weight in kg",
            "Total Value Of Parcel",
            "Goods Value Currency",
            "Origin",
            "Shipper Name",
            "Shipper Name",
            "Shipper Address 1",
            "Shipper Address 2",
            "Shipper City",
            "Shipper State",
            "Shipper Postcode",
            "Shipper Country"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num,
                           value=self.convert_value(header))
            cell.font = Font(bold=True)
            cell.alignment = header_alignment
            cell.border = thin_border
            cell.fill = PatternFill(
                start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        for idx, detail in enumerate(self.details, start=2):
            awb = detail.get("awb_details", {})
            consignee = detail.get("consignee", {}) or {}
            consignor = detail.get("consignor", {}) or {}

            awbno = awb.get("awb_no")
            if awb.get("forwarding_number"):
                awbno = awb.get("forwarding_number")

            row = [
                # Parcel Barcode / HAWB Number
                awbno,
                # HAWB Number
                awb.get("awb_no"),
                # Customer Name
                consignee.get("name"),
                # Customer Address 1
                consignee.get("address1"),
                # Customer Address 2
                consignee.get("address2", ""),
                consignee.get("city"),                                   # City
                # State
                consignee.get("state_short_name"),
                # Postcode
                consignee.get("postcode", ""),
                # Customer Telephone (if available)
                consignee.get("phone"),
                # Description of goods
                awb.get("content"),
                "KTM",                                                   # IATA Origin Port
                "SYD",                                                   # IATA Dest Port
                self.run.flight_no,                                      # Flight No
                self.run.mawb_no,                                        # Master Bill
                "",                                        # Sub Master
                "",  # ETA
                # Parcel count
                awb.get("box_count"),
                # Weight in kg
                awb.get("total_actual_weight"),
                # Total Value Of Parcel
                awb.get("shipment_value", ""),
                "AUD",                                                   # Goods Value Currency
                "KTM",                                                   # Origin
                # Shipper Name
                consignor.get("company", ""),
                consignor.get("name", ""),
                # Shipper Address 1
                consignor.get("address1"),
                # Shipper Address 2
                consignor.get("address2", ""),
                # Shipper City
                consignor.get("city"),
                # Shipper State
                consignor.get("state", ""),
                # Shipper Postcode
                consignor.get("postcode"),
                # Shipper Country
                consignor.get("country")
            ]

            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=idx, column=col_num,
                               value=self.convert_value(value))
                cell.alignment = center_alignment
                cell.border = thin_border

        # Auto-fit column widths
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value))
                             if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = max_length + 4

        # Output response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="AUSTRALIA_MANIFEST_NCB_{self.run.run_no}_RUN_{self.run.run_no}.xlsx"'
        return response

    def export_yyz(self):
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "YYZ Export"

        # Define new headers
        headers = [
            "Tracking Number", "Reference", "Internal Account Number", "Shipper",
            "Shipper Address 1", "Shipper Address 2", "Shipper Address 3", "Shipper City",
            "Shipper County/State", "Shipper Zip", "Shipper Country Code", "Consignee",
            "Address1", "Address2", "Address3", "City", "Province", "Province Code",
            "Zip", "Country Code", "Email", "Phone", "Pieces", "Total Weight",
            "Weight UOM", "Total Value", "Currency", "Incoterms", "Item Description",
            "Item HS Code", "QUANTITY", "Item Value", "Country Of Origin"
        ]

        # Define styling
        header_font = Font(bold=True, color="000000")  # Black text
        header_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )

        # Write and style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        current_row = 2

        # Write data rows
        for detail_idx, detail in enumerate(self.details):
            awb = detail.get("awb_details", {})
            boxes = detail.get("boxes", [])
            consignor = detail.get("consignor", {}) or {}
            consignee = detail.get("consignee", {}) or {}
            company = detail.get("company", {}) or {}

            detail_start_row = current_row

            for box in boxes:
                # Support multiple items per box
                items = box.get("items") or []

                for item in items:
                    hs_code = f"{item.get('hs_code')}00000"

                    row_data = [
                        # Tracking Number
                        self.convert_value(awb.get(
                            "forwarding_number", ""), is_uppercase=True),
                        "",  # Reference
                        "",  # Internal Account Number
                        self.convert_value(consignor.get("name", "")),
                        self.convert_value(consignor.get("address1", "")),
                        self.convert_value(consignor.get("address2", "")),
                        "",  # Address 3
                        self.convert_value(consignor.get("city", "")),
                        self.convert_value(consignor.get("state", "")),
                        self.convert_value(consignor.get("postcode", "")),
                        self.convert_value("NP"),
                        self.convert_value(consignee.get("name", "")),
                        self.convert_value(consignee.get("address1", "")),
                        self.convert_value(consignee.get("address2", "")),
                        "",  # Address 3
                        self.convert_value(consignee.get("city", "")),
                        self.convert_value(consignee.get("state", "")),
                        self.convert_value(
                            consignee.get("state_short_name", "")),
                        self.convert_value(consignee.get("postcode", "")),
                        self.convert_value(consignee.get(
                            "country_short_name", "")),
                        self.convert_value(consignee.get("email", "")),
                        self.convert_value(consignee.get("phone", "")),
                        self.convert_value(awb.get("box_count")),
                        self.convert_value(awb.get("total_actual_weight", "")),
                        self.convert_value("KGS"),
                        self.convert_value(awb.get("shipment_value", "")),
                        self.convert_value("CAD"),  # default currency
                        self.convert_value(awb.get("shipment_terms", "")),
                        self.convert_value(item.get("description", "")),
                        self.convert_value(hs_code[:8]),
                        self.convert_value(item.get("quantity", "")),
                        self.convert_value(item.get("unit_rate", "")),
                        self.convert_value("NP")  # Country of Origin (Nepal)
                    ]

                    # Write row data
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row,
                                       column=col, value=value)
                        cell.border = thin_border
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center")

                    current_row += 1

            # Add thick border around the entire detail block
            detail_end_row = current_row - 1
            if detail_end_row >= detail_start_row:
                # Apply thick border to the entire block
                for row in range(detail_start_row, detail_end_row + 1):
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row, column=col)

                        # Determine which sides need thick borders
                        left_thick = col == 1
                        right_thick = col == len(headers)
                        top_thick = row == detail_start_row
                        bottom_thick = row == detail_end_row

                        # Create custom border for this cell
                        cell.border = Border(
                            left=Side(style='thick' if left_thick else 'thin'),
                            right=Side(
                                style='thick' if right_thick else 'thin'),
                            top=Side(style='thick' if top_thick else 'thin'),
                            bottom=Side(
                                style='thick' if bottom_thick else 'thin')
                        )

        # Auto-adjust column widths based on actual content after all data is written
        for col in range(1, len(headers) + 1):
            column_letter = ws.cell(row=1, column=col).column_letter
            max_length = 0

            # Check all cells in this column (including header)
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)

            # Set column width with padding (extra margin on left and right)
            ws.column_dimensions[column_letter].width = min(
                max_length + 4, 50)  # Cap at 50 for very long content

        # Save to buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Create HTTP response
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename=CANADA_PRE_ALERT_YYZ_{self.run.run_no}.xlsx'

        return response

    def dxb_corierx_manifest(self):
        wb = openpyxl.Workbook()
        ws_default = wb.active
        ws_default.title = "Others"

        # Define styles
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=False)
        header_fill = PatternFill(
            start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

        headers = [
            "awbNo", "Consignee Company", "Consignee Name", "Consignee Address1", "Consignee Address2",
            "Consignee City", "Consignee Phone1", "Consignee Phone2", "Consignee Zip", "geol",
            "Orgin", "Dest", "Pieces", "Weight", "Product", "Service", "Special Inst", "Description",
            "DeclaredValue", "invoicecurr", "Reference", "Agent Code", "Agentawb", "Amount",
            "NCND Amt", "NCND Currenct", "DDP", "Shipper", "Shipper Name",
            "Shipper Address1", "Shipper Address2", "Shipper Phone1", "Shipper Phone 2"
        ]

        def write_header(ws):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num,
                               value=self.convert_value(header))
                cell.font = Font(bold=True)
                cell.alignment = header_alignment
                cell.border = thin_border
                cell.fill = header_fill

        def write_row(ws, row_idx, detail):
            awb = detail.get("awb_details", {})
            consignee = detail.get("consignee", {}) or {}
            consignor = detail.get("consignor", {}) or {}
            service = str(awb.get("service", "")).upper()
            custom_awbno = awb.get("reference_number")
            if service == "COURIERX":
                custom_forwarding_number = awb.get("reference_number")
            else:
                custom_forwarding_number = awb.get(
                    "forwarding_number").split(",")[0]

            row = [
                custom_awbno,
                consignee.get("company", ""),
                consignee.get("name", ""),
                consignee.get("address1", ""),
                consignee.get("address2", ""),
                consignee.get("city", ""),
                consignee.get("phone", ""),
                "",
                consignee.get("postcode", ""),
                "",
                "NP",
                awb.get("destination_short_name", ""),
                awb.get("box_count", ""),
                awb.get("total_actual_weight", ""),
                awb.get("product_type", ""),
                "NOR",
                "",
                awb.get("content", ""),
                awb.get("shipment_value", ""),
                awb.get("currency", ""),
                awb.get("awb_no", ""),
                "",
                custom_forwarding_number,
                "",
                "",
                "",
                "",
                consignor.get("company", ""),
                consignor.get("name", ""),
                "GAUSHALA",
                "BATTISPUTALI",
                consignor.get("phone", ""),
                "",
            ]

            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_num,
                               value=self.convert_value(value))
                cell.alignment = center_alignment
                cell.border = thin_border

        # dynamic sheet creation
        sheet_map = {"OTHERS": ws_default}
        row_counters = {"OTHERS": 2}
        write_header(ws_default)

        def get_or_create_sheet(name):
            if name not in sheet_map:
                ws = wb.create_sheet(name)
                sheet_map[name] = ws
                row_counters[name] = 2
                write_header(ws)
            return sheet_map[name]

        # Fill data into correct sheet
        for detail in self.details:
            awb = detail.get("awb_details", {})
            service = str(awb.get("service", "")).upper()

            if service in ("COURIERX", "UPS", "FEDEX"):
                # Sheet names: Courierx, Ups, Fedex
                ws = get_or_create_sheet(service.title())
                idx = row_counters[service.title()]
                write_row(ws, idx, detail)
                row_counters[service.title()] += 1
            else:
                ws = sheet_map["OTHERS"]
                idx = row_counters["OTHERS"]
                write_row(ws, idx, detail)
                row_counters["OTHERS"] += 1

        # Remove "Others" sheet if it has no data beyond header
        if row_counters["OTHERS"] == 2 and len(sheet_map) > 1:
            wb.remove(sheet_map["OTHERS"])

        # Auto-adjust column widths
        for ws in sheet_map.values():
            for column_cells in ws.columns:
                max_length = max(len(str(cell.value))
                                 if cell.value else 0 for cell in column_cells)
                ws.column_dimensions[column_cells[0]
                                     .column_letter].width = max_length + 4

        # Save and return response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename=Run_{self.run.run_no}_Vendor_Manifest.xlsx'
        return response

    def export_nepal_custom(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Nepal Custom Manifest"

        # --- Borders & alignment ---
        thin = Side(style="thin", color="000000")
        thick = Side(style="thick", color="000000")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        content_font = Font(size=9)

        # --- Header row ---
        headers = ["S.NO", "AWBNO", "REF HAWB NO", "CONSIGNOR", "CONSIGNEE", "DESTINATION",
                   "PCS", "WEIGHT", "CONTENT", "VALUE", "BAGGING"]
        header_row = 10
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.font = Font(size=9, bold=True)

        row_pointer = header_row + 1
        serial = 1

        # --- Flatten boxes ordered by bag_no ---
        all_boxes = []
        for detail in self.details:
            awb = detail.get("awb_details", {})
            consignor = detail.get("consignor", {}) or {}
            consignee = detail.get("consignee", {}) or {}
            boxes = detail.get("boxes", [])
            for box in boxes:
                all_boxes.append({
                    "awb_no": detail.get("awb_no"),
                    "consignor": consignor,
                    "consignee": consignee,
                    "awb_value": awb.get("shipment_value", ""),
                    "box": box,
                })

        all_boxes = sorted(
            all_boxes, key=lambda b: b["box"].get("bag_no", "") or "")

        prev_bag = None
        bag_start = None

        for entry in all_boxes:
            awb_no = entry["awb_no"]
            consignor = entry["consignor"]
            consignee = entry["consignee"]
            box = entry["box"]

            bagging = box.get("bag_no", "")
            pcs = 1
            weight = box.get("actual_weight", "")
            # Compute REF HAWB NO and DESTINATION from awb details
            ref_hawb_no = awb.get("reference_number", "")
            destination = awb.get("destination", "") or awb.get(
                "destination_short_name", "")

            consignor_lines = list(filter(None, [
                consignor.get("company", "") if consignor.get(
                    "company") else consignor.get("name", "")
            ]))

            consignee_lines = list(filter(None, [
                consignee.get("company", "") if consignee.get(
                    "company") else consignee.get("name", "")
            ]))

            # CONTENT
            content = ", ".join(f"{item.get('description', '')} {item.get('quantity', '')} {item.get('unit_type', '')}"
                                for item in box.get("items", []))
            content_lines = [content] if content else [""]

            # VALUE
            value = sum(float(item.get("amount", "0"))
                        for item in box.get("items", []))

            max_stack = max(len(consignor_lines), len(
                consignee_lines), len(content_lines))
            start_row = row_pointer
            end_row = row_pointer + max_stack - 1

            # --- Write data rows ---
            for i in range(max_stack):
                ws.cell(row=row_pointer, column=1,
                        value=serial if i == 0 else "")
                ws.cell(row=row_pointer, column=2,
                        value=awb_no if i == 0 else "")
                ws.cell(row=row_pointer, column=3,
                        value=ref_hawb_no if i == 0 else "")
                ws.cell(row=row_pointer, column=4, value=consignor_lines[i] if i < len(
                    consignor_lines) else "")
                ws.cell(row=row_pointer, column=5, value=consignee_lines[i] if i < len(
                    consignee_lines) else "")
                ws.cell(row=row_pointer, column=6,
                        value=destination if i == 0 else "")
                ws.cell(row=row_pointer, column=7, value=pcs if i == 0 else "")
                ws.cell(row=row_pointer, column=8,
                        value=weight if i == 0 else "")
                ws.cell(row=row_pointer, column=9,
                        value=content_lines[i] if i < len(content_lines) else "")
                ws.cell(row=row_pointer, column=10,
                        value=value if i == 0 else "")
                ws.cell(row=row_pointer, column=11,
                        value=bagging if i == 0 else "")

                # Styles
                for col in range(1, 12):
                    cell = ws.cell(row=row_pointer, column=col)
                    cell.border = thin_border
                    cell.alignment = center_alignment
                    cell.font = content_font

                row_pointer += 1

            # --- Merge columns ---
            for col in [1, 2, 3, 6, 7, 8, 9, 10, 11]:
                ws.merge_cells(start_row=start_row, start_column=col,
                               end_row=end_row, end_column=col)

            serial += 1

            # --- Empty row after each entry ---
            for col in range(1, 12):
                cell = ws.cell(row=row_pointer, column=col, value="")
                cell.border = thin_border
                cell.alignment = center_alignment
                cell.font = content_font
            row_pointer += 1

            # --- Apply thick border around bag group ---
            if prev_bag and bagging != prev_bag:
                for r in range(bag_start, start_row):
                    for c in range(1, 12):
                        cell = ws.cell(row=r, column=c)
                        cell.border = Border(
                            left=thick if c == 1 else thin,
                            right=thick if c == 11 else thin,
                            top=thick if r == bag_start else thin,
                            bottom=thick if r == start_row-1 else thin
                        )
                bag_start = start_row
            elif prev_bag is None:
                bag_start = start_row
            prev_bag = bagging

        # --- Final bag thick border ---
        if bag_start is not None:
            final_end_row = row_pointer - 2  # because last empty row
            for r in range(bag_start, final_end_row+1):
                for c in range(1, 12):
                    cell = ws.cell(row=r, column=c)
                    cell.border = Border(
                        left=thick if c == 1 else thin,
                        right=thick if c == 11 else thin,
                        top=thick if r == bag_start else thin,
                        bottom=thick if r == final_end_row else thin
                    )

        # --- Auto-size columns ---
        for column_cells in ws.columns:
            length = max(len(str(cell.value))
                         if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(
                column_cells[0].column)].width = length + 4
        total_box = self.run.total_boxes
        total_weight = self.run.total_actual_weight
        run_no = self.run.run_no

        # --- Column A & H details ---
        ws.cell(row=1, column=1, value="FROM")
        ws.cell(row=2, column=1, value="IDEIDEAL CARGO & COURIER  PVT. LTD.")
        ws.cell(row=3, column=1, value="BATTISPUTALI ROAD GAUSHALA KATHMANDU NEPAL")
        ws.cell(row=4, column=1, value=f"TEL- 01-4578809")

        ws.cell(row=5, column=1, value="")

        ws.cell(row=6, column=1, value="TO:")
        ws.cell(row=7, column=1, value=f"{self.run.vendor.legal_name}")
        ws.cell(row=8, column=1, value=f"{self.run.vendor.country}")
        ws.cell(row=9, column=1, value="")

        ws.cell(row=1, column=5, value="EXIM Code: 6023969470146NP")

        ws.cell(row=1, column=8, value=f"MAWB: {self.run.mawb_no}")
        ws.cell(row=2, column=8, value=f"TOTAL NO OF BOXES : {total_box}")
        ws.cell(row=3, column=8, value=f"GROSS WEIGHT: {total_weight} KGS")
        ws.cell(row=4, column=8, value=f"RUN NO: {run_no} ")
        ws.cell(row=5, column=8, value=f"Flight No: {self.run.flight_no}")
        ws.cell(row=6, column=8,
                value=f"AIRPORT OF SHIPMENT DATE: {self.run.flight_departure_date.strftime('%Y-%m-%d')}")

        # --- Print settings ---
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.scale = None
        ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2)
        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered = True

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename=NEPAL CUSTOM MANIFEST {self.run.run_no}.xlsx'
            }
        )

    def export_cds_manifiest_ubx_us(self, type):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CDS Manifiest"

        # Define border and alignment styles
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=False)

        headers = [

            "STAT VALUE", "GOODS DESCRIPTION", "GROSS MASS", "NET MASS", "SUPP UNITS", "SUPP UNITS CODE", "VALUE", "VALUE CURR", "VALUATION METHOD", "DESTINATION CNTRY",
            "CPC FIRST 4", "CPC ADDITIONAL", "ORIGIN CNTRY", "COMMUDITY CODE",  "TARIC CODES", "NATIONAL CODES", "PREFERENCE", "DOC CODE 1", "DOC REF 1", "DOC REASON 1",
            "DOC STATUS 1", "DOC CODE 2", "DOC REF 2", "DOC REASON 2", "DOC STATUS 2", "DOC CODE 3", "DOC REF 3", "DOC REASON 3", "DOC STATUS 3", "DOC CODE 4", "DOC REF 4",
            "DOC REASON 4", "DOC STATUS 4", "AI STATEMENT 1", "AI DESCRIPTION 1", "AI TYPE 1", "AI STATEMENT 2", "AI DESCRIPTION 2", "AI TYPE 2", "TAX TYPE 1",
            "TAX MOP 1", "PREV DOC CLASS 1", "PREV DOC REF 1", "PREV DOC TYPE 1", "PKG QTY", "PKG TYPE", "PKG MARKS", "PARTIES SHORTNAME 1", "PARTIES ID 1",
            "PARTIES TYPE 1", "PARTIES LONG NAME 1", "PARTIES ADDRESS 1", "PARTIES CITY 1", "PARTIES COUNTRY 1", "PARTIES POST CODE 1"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num,
                           value=self.convert_value(header))
            cell.font = Font(bold=True)
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for idx, detail in enumerate(self.details, start=2):
            awb = detail.get("awb_details", {})
            consignee = detail.get("consignee", {}) or {}
            consignor = detail.get("consignor", {}) or {}

            row = [
                "",
                awb.get("content"),
                awb.get("total_actual_weight"),
                awb.get("total_actual_weight"),
                awb.get("box_count"),
                "",
                awb.get('shipment_value'),
                "GBP",
                "1",
                "",
                "4000", "1RC", "NP",
                "6210500000",
                "",
                "",
                "100",
                "",
                "EVIDENCE AVALIBLE FROM TRADER",
                "",
                "JP",
                "",
                "",
                "",
                "",
                "N935",
                detail.get("awb_no", ""),
                "",
                "AC",
                "",
                "",
                "",
                "",
                "RCD01",
                "RELIEF FROM CUSTOMS DUTY CLAIMED",
                "",
                "",
                "",
                "",
                "B00",
                "E",
                "Z",
                self.run.mawb_no,
                "741", "1", "PK",

                detail.get('awb_no'),
                consignee.get("name", "").split()[0],
                "",
                "RN",
                consignee.get("name", ""),
                consignee.get("address1", ""),
                consignee.get("city", ""),

                consignee.get("country_short_name", ""),
                consignee.get("postcode", ""),
            ]

            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=idx, column=col_num,
                               value=self.convert_value(value))
                cell.alignment = center_alignment
                cell.border = thin_border

        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value))
                             if cell.value else 0 for cell in column_cells)
            adjusted_width = max_length + 4
            column_letter = column_cells[0].column_letter
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to memory and return response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename=CDS_400004_MANIFEST_{type}_UX_{self.run.run_no}.xlsx'
        return response
